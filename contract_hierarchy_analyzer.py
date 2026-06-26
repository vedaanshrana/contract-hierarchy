# ============================================================
# CONTRACT HIERARCHY ANALYZER
# Fiserv Legal Ops POC
# ============================================================
# SETUP (one-time, in terminal):
#   pip install -r requirements.txt
#   (or: pip install pymupdf openai pandas openpyxl plotly tenacity pillow docx2pdf extract-msg)
#   (docx2pdf is only required if any client folder contains .docx files;
#    it depends on MS Word being installed on Windows / Office on macOS.
#    extract-msg is only required if any client folder contains .msg files.)
#
# USAGE:
#   1. Copy .env.example to .env and set OPENAI_API_KEY (see VDI_SETUP.md)
#   2. Place contracts in subfolders: ./contracts/<ClientName>/
#      Supported file types: .pdf, .docx, .tif, .tiff, .txt, .msg
#   3. Open terminal, navigate to this folder, run:
#        python contract_hierarchy_analyzer.py
#      Or restrict to one client:
#        python contract_hierarchy_analyzer.py "Peoples Bank"
#   4. Open output/<ClientName>/contracts_hierarchy.html in any browser
# ============================================================

import os
import re
import sys
import json
import base64
import time
import math
import tempfile
from pathlib import Path
from io import BytesIO
from datetime import datetime
from collections import defaultdict

import fitz                                  # PyMuPDF
from PIL import Image, ImageSequence
from openai import OpenAI
import httpx                                  # direct HTTP for the Fiserv Foundation gateway
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# docx2pdf is only needed when a client folder contains .docx files. It relies
# on MS Word (Windows) or Office (macOS) for the underlying conversion. We import
# lazily so the script still runs on machines without Word when no .docx is present.
try:
    from docx2pdf import convert as _docx2pdf_convert
    _HAS_DOCX2PDF = True
except ImportError:
    _docx2pdf_convert = None
    _HAS_DOCX2PDF = False

# extract-msg is only needed when a client folder contains .msg files (Outlook
# email format). Same lazy-import pattern as docx2pdf — the script keeps working
# without it as long as no .msg files are present in the run.
try:
    import extract_msg as _extract_msg
    _HAS_EXTRACT_MSG = True
except ImportError:
    _extract_msg = None
    _HAS_EXTRACT_MSG = False

# ============================================================
# CONFIG — EDIT THESE LINES BEFORE RUNNING
# ============================================================
_SCRIPT_DIR     = Path(__file__).resolve().parent

# ── Secrets / endpoint (loaded from the environment, never hard-coded) ──
# The OpenAI API key and (optional) endpoint are read from the environment so
# that no secret ever lives in this file or in git history. For local and VDI
# use, put them in a ".env" file next to this script (copy .env.example → .env).
# Anything already exported in the shell or set in VDI system settings wins over
# the .env file. See VDI_SETUP.md for full instructions.
def _load_dotenv(dotenv_path):
    """Minimal .env loader — no third-party dependency. Reads KEY=VALUE lines
    and sets them in os.environ only if not already defined (real env wins)."""
    try:
        with open(dotenv_path, "r", encoding="utf-8") as _fh:
            for _raw in _fh:
                _line = _raw.strip()
                if not _line or _line.startswith("#") or "=" not in _line:
                    continue
                _k, _, _v = _line.partition("=")
                _k = _k.strip()
                _v = _v.strip().strip('"').strip("'")
                if _k and _k not in os.environ:
                    os.environ[_k] = _v
    except FileNotFoundError:
        pass

_load_dotenv(_SCRIPT_DIR / ".env")
OPENAI_API_KEY  = os.environ.get("OPENAI_API_KEY", "").strip()    # set in .env or VDI env
OPENAI_BASE_URL = os.environ.get("OPENAI_BASE_URL", "").strip()   # optional: corporate/Azure gateway; blank = default OpenAI

# ── Backend selection ──────────────────────────────────────
# OPENAI_BACKEND = "openai" (default): call api.openai.com via the OpenAI SDK
#                  Responses API (current behaviour, unchanged).
# OPENAI_BACKEND = "fiserv": call the Fiserv Foundation API gateway directly.
#                  This is an Azure-style chat/completions endpoint; the model
#                  is chosen at the gateway by the X-Purpose header, so whatever
#                  OPENAI_MODEL you set locally, the actual model is whatever the
#                  endpoint returns (we report that real name in the run metrics).
OPENAI_BACKEND      = os.environ.get("OPENAI_BACKEND", "openai").strip().lower()
FOUNDATION_API_URL  = os.environ.get("FOUNDATION_API_URL", "").strip()   # full chat/completions URL incl. ?api-version=
FISERV_EMAIL        = os.environ.get("FISERV_EMAIL", "").strip()
FISERV_PURPOSE_GPT5 = os.environ.get("FISERV_PURPOSE_GPT5", "").strip()   # X-Purpose tag for gpt-5.x models
FISERV_PURPOSE_GPT4 = os.environ.get("FISERV_PURPOSE_GPT4", "").strip()   # X-Purpose tag for everything else

OPENAI_MODEL    = "gpt-5.2"         # confirmed from team's existing code
DPI             = 600               # image render quality (team's setting)
CHUNK_SIZE      = 8                # pages per API call (team's setting)
MAX_CHUNKS      = 2                 # max chunks per contract (24 pages covers title + body; exhibits have no hierarchy signals)
CONTRACTS_ROOT  = str(_SCRIPT_DIR)                                   # root folder; one subfolder per client
OUTPUT_DIR      = str(_SCRIPT_DIR / "output")
CACHE_FILE      = str(_SCRIPT_DIR / "output" / "extraction_cache.json")
API_DELAY       = 3                 # seconds between API calls (rate limit buffer)
# Per-attempt timeout for OpenAI requests. Combined with the existing tenacity
# retry policy (3 attempts, exponential backoff), this caps the worst case at
# roughly 3 * API_TIMEOUT_SECONDS plus backoff sleeps. The SDK default is 10
# minutes, which means a single hung request used to consume up to ~30 minutes
# silently before failing. 180 s is enough for a 12-page multi-image chunk
# under normal conditions and short enough that a genuinely hung request
# bounces to the next retry promptly.
API_TIMEOUT_SECONDS = 180

# ── Product-name canonicaliser (May 21 2026) ──
# Aligns LLM-extracted product names (carried in section_header_products) to
# Fiserv's canonical product dictionary. The script scans _SCRIPT_DIR for any
# file matching PRODUCT_DICTIONARY_GLOB and picks the one with the highest
# version-number suffix (so dropping in PC-PH-ProductNames_Dictionary_v2.xlsx
# alongside v1 picks up v2 on the next run, no code change). If no file
# matches, the canonicaliser becomes a no-op and the pipeline behaves as if
# this feature didn't exist. PRODUCT_FUZZY_THRESHOLD is a difflib ratio
# (0.0-1.0); raise it for stricter matching, lower for looser.
PRODUCT_DICTIONARY_GLOB        = "PC-PH-ProductNames_Dictionary_v*.xlsx"
PRODUCT_CANONICALIZER_ENABLED  = True
PRODUCT_FUZZY_THRESHOLD        = 0.88

# Restrict the run to a subset of client folders.
#   None          → process every subfolder under CONTRACTS_ROOT (default).
#   ["X", "Y"]    → process only these client folders by exact name.
# Can also be overridden from the command line:
#   python contract_hierarchy_analyzer.py "Peoples Bank"
#   python contract_hierarchy_analyzer.py "Peoples Bank" "MetaBank"
ONLY_CLIENTS = ["EDU_TEST_FINAL"]

# File types the scanner will pick up from each client folder.
SUPPORTED_EXTS = {".pdf", ".docx", ".tif", ".tiff", ".txt", ".msg"}

# Optional pretty-name overrides for chart titles. Folder name is used as-is
# if the client isn't listed here.
CLIENT_TITLE_OVERRIDES = {
    "Peoples Bank": "The Peoples Bank",
}

# Upper bound on raw .txt length sent to the model. 50k chars ≈ 12k tokens,
# which comfortably fits within a single prompt and is far more than any
# realistic contract body text. Longer files are truncated with a note.
TXT_MAX_CHARS = 50_000

# When True, cache entries for child-type contracts (Amendment, Schedule, SOW,
# Purchase Order, etc.) that came back with an empty parent_references list
# are evicted from the cache on the next run so they get re-extracted with
# the strengthened prompt. The re-extracted entries are flagged with
# parent_rescue_attempted=True afterwards so a third re-extraction never
# happens — if the second attempt also finds no parent text, the document
# is treated as genuinely silent on its parent (Orphan).
# Set to False to disable the rescue pass and use the existing cache as-is.
RESCUE_MISSING_PARENT_REFS = True

# ============================================================
# SYSTEM PROMPT
# ============================================================
SYSTEM_PROMPT = """You are a legal contract analyzer specializing in financial services agreements.
Extract metadata from the contract pages shown.

Return ONLY a valid JSON object — no markdown, no explanation, no code blocks, nothing else.

{
  "contract_type": "MSA" or "Amendment" or "Sub-Amendment" or "SOW" or "Schedule" or "Standalone",
  // VALIDATOR RULE (May 2026): classify based on the TITLE PAGE / PREAMBLE / DOCUMENT BODY, never the filename. People uploading contracts to the repository over the last 30 years frequently picked the wrong document type at index time, so the filename's type keyword (e.g. "Services Agreement", "Amendment", "Hardware Agreement") is often misleading. If the document body opens with "AMENDMENT NO. X TO …" but the filename says "Services Agreement", the contract is an Amendment. If the filename says "Amendment" but the body is titled "Master Agreement", it is an MSA. Read the actual document and report what you see there.
  // MSA includes: Master Agreement, Master Services Agreement, Relationship Agreement, Network Membership Agreement, Membership Agreement, Amended and Restated Agreement
  // Amendment includes: any numbered amendment (First, Second, ... Sixteenth), Addendum, Rider, Consent, Waiver, Extension, Renewal
  // SOW includes: Statement of Work, Order Form, Work Order, Consulting Services Exhibit
  "amendment_number": amendment number as a string such as "16", or null if not an amendment,
  "signed_date": "YYYY-MM-DD" format — the date THIS document was signed/executed by the parties. Look for dates next to signature blocks on the signature page, "executed as of" / "fully executed" / "signed this __ day of" language, or DocuSign / e-signature timestamps. If different parties signed on different dates, use the most recent (the date the contract became fully executed). Return null if no signing date is clearly visible.,
  "effective_date": "YYYY-MM-DD" format — the date THIS document defines as its effective or commencement date (often stated in the preamble: "This Agreement, effective as of…"). May or may not equal signed_date. NOT any date referenced for a parent or predecessor contract. Return null if not stated.,
  "parties": ["Party Name 1", "Party Name 2"],
  "parent_references": ["EXACT VERBATIM QUOTE(s) from this document that reference a parent or predecessor contract. WHERE TO LOOK, in priority order: (1) the title page / preamble / first paragraph — almost every amendment, schedule, exhibit, addendum, appendix, attachment, SOW, or purchase order opens with language like 'AMENDMENT NO. X TO THE MASTER AGREEMENT dated [DATE], between [PARTIES]'; (2) any 'WHEREAS' / 'Recitals' clauses — 'the parties entered into a Master Agreement effective as of [DATE]'; (3) the first numbered section, especially anything titled 'Defined Terms', 'Background', or 'Recitals'; (4) phrases like 'pursuant to', 'governed by', 'subject to the terms of', 'as amended by', 'the Agreement (as amended)'. (5) AGREEMENT / CONTRACT NUMBERS — validator-confirmed (May 2026): the parent's internal agreement or contract number can appear printed in the body, stamped, or handwritten in the margin (e.g. 'Agreement No. 2035', 'Contract # CT-2035', a handwritten '#2035' in the top-right corner). When the same number reappears on a related document it is a strong parent link even if no date or title is shared — capture the exact text. EXAMPLES of valid extractions: 'Amendment No. 16 to the Agreement dated February 4, 2009', 'This Statement of Work is governed by the Master Services Agreement dated July 30, 2008 between Fiserv and Client', 'WHEREAS, Fiserv and Client are parties to that certain Master Agreement effective March 1, 2016 (the \"Agreement\")', 'Agreement No. 2035' (handwritten or printed). STRICT RULE: if this document's contract_type is subordinate (Amendment, Sub-Amendment, Addendum, Schedule, Exhibit, Appendix, Attachment, SOW, Statement of Work, Purchase Order, Order Form, Subsequent Order, Work Order, Termination Notice, Consent, Waiver, Rider, Extension, or Renewal), it ALMOST CERTAINLY references a parent agreement somewhere in the first two pages — read them carefully and extract every reference you find. Only return [] AFTER confirming that no parent text exists in the document; never return [] just because the first sentence didn't have one."],
  "supersedes_text": "EXACT VERBATIM QUOTE of any language stating this document supersedes, amends, modifies or replaces a prior agreement" or null,
  "internal_doc_codes": [{"value": "string — the code or number exactly as it appears", "position": "one of: handwritten_margin | handwritten_inline | printed_stamp | printed_label_box | printed_body | printed_header | printed_footer | unknown"}, "… more entries …"],
  // internal_doc_codes — STRUCTURED LIST OF OBJECTS (new format, May 12 2026).
  // Each entry MUST be an object with two fields: `value` (the literal text
  // of the code) AND `position` (where on the page you found it).
  //
  // Why position matters: the same numeric string can appear on a page as
  // (a) a handwritten margin note that the originator wrote to cross-link this
  //     document to a parent agreement — VALIDATOR-CONFIRMED parent-link signal;
  // (b) a printed header/footer that is the client's account number — appears
  //     on EVERY document for this client and is NOT a parent-link signal.
  // Without the position label we can't tell these apart. The resolver
  // downstream uses the position to decide whether to trust the code as a
  // parent-link or treat it as boilerplate.
  //
  // POSITION VOCABULARY (use exactly one of these strings — any other value
  // will be coerced to "unknown"):
  //
  //   "handwritten_margin" — number written by hand (ink / pencil) in a page
  //     margin, top corner, between paragraphs, near the signature block, or
  //     in any blank space of the page. NOT aligned with any printed label.
  //     This is the validator's HIGH-VALUE signal.
  //
  //   "handwritten_inline" — number written by hand inside a printed form
  //     field, on top of a printed line ("Contract No. ____"), or written
  //     into the body text. Same high-value signal.
  //
  //   "printed_stamp" — number appearing in a stamp impression (often with
  //     a date), e.g. "RECEIVED 03/15/2006 Contract #290". HIGH-VALUE.
  //
  //   "printed_label_box" — number printed in a clearly-labelled field box,
  //     e.g. "Contract Number: 2035", "Agreement No.: CT-2035". HIGH-VALUE.
  //
  //   "printed_body" — number printed inside the body text of a clause,
  //     e.g. "...as further described in Agreement No. 2035." HIGH-VALUE
  //     when it references another agreement; otherwise medium.
  //
  //   "printed_header" — number printed in the running page header at the
  //     top of EVERY page (e.g. Fiserv letterhead, page-number band). Often
  //     a client account number or template ID — LOW-VALUE parent-link
  //     signal, almost always boilerplate.
  //
  //   "printed_footer" — number printed in the running page footer at the
  //     BOTTOM of every page (e.g. doc revision tag, footer ID, DocuSign
  //     Envelope ID). LOW-VALUE parent-link signal, almost always
  //     boilerplate or unique-per-doc.
  //
  //   "unknown" — you found the code but cannot determine where on the page
  //     it appeared. Returned only when the rendering is ambiguous.
  //
  // EXAMPLES of valid entries:
  //   {"value": "290",                              "position": "handwritten_margin"}
  //   {"value": "Contract No. 2035",                "position": "printed_label_box"}
  //   {"value": "9904-3201",                        "position": "handwritten_margin"}
  //   {"value": "Agreement No. CT-2035",            "position": "printed_body"}
  //   {"value": "DocuSign Envelope ID: ABC-123…",   "position": "printed_footer"}
  //   {"value": "055c020209",                       "position": "printed_footer"}
  //   {"value": "Client #: 290",                    "position": "printed_header"}
  //
  // SCANNING RULES (validator's May 12 2026 call — the dominant linking signal
  // in this corpus AND the most-missed in prior runs):
  //   - Deliberately scan EVERY page margin and EVERY patch of white space
  //     for short numeric strings (3-7 digits). Do NOT skip a number because
  //     it "looks informal" or "isn't aligned with a printed label" — those
  //     are exactly the validator-flagged high-value cases.
  //   - DUAL NUMBERS: when a page shows two related numbers — e.g. a printed
  //     "Contract 9904-3201" and a handwritten "290" scrawled below it —
  //     emit BOTH as separate entries, each with its own correct position.
  //     The handwritten one is frequently the cross-document linking key.
  //   - Return [] only after deliberately scanning every page margin and
  //     every block of white space for stray numbers.
  "section_structure": [
    {
      "header":   "verbatim text of a top-level section heading in this document — e.g. 'Schedule A: ASP Services Exhibit', 'Account Processing Services Schedule', 'Termination', 'Fees'. Top-level headings are typically the boldest / largest / most prominent headings in the document layout, often starting a new schedule / exhibit / article.",
      "product":  "name of the Fiserv product/service this header refers to, or null when the header is generic (e.g. 'Termination', 'Fees', 'Definitions', 'Background', 'Recitals'). Use the most specific product name visible in the header text — e.g. for 'Configure Digital Services Schedule to ASP Services Exhibit', the product is 'Configure Digital'. For umbrella-only headers like 'ASP Services Exhibit' with no specific child product on the same line, use 'ASP Services Exhibit'.",
      "subheaders": [
        {
          "text":    "verbatim text of a sub-section heading that appears DIRECTLY UNDER this top-level header, before the next top-level header. Sub-headings are typically less prominent than the top-level header (smaller font, sub-numbered, indented) but still clearly headings — they introduce a topical block of content within the section. Example: under 'Schedule A: ASP Services Exhibit' you might see sub-headers 'Account Processing Services', 'ATM/EFT Services', 'Item Processing Services', etc.",
          "product": "name of the Fiserv product/service this sub-header refers to, or null when generic. Use the most specific name visible — for 'ATM/EFT Services' the product is 'ATM/EFT Services'; for 'Account Processing Services' it's 'Account Processing Services'.",
          "items":   ["LEAF-LEVEL product/service names listed directly beneath this sub-header, before the next sub-header. These are typically bullet points, table rows, fee-line labels, or short stand-alone labels under the sub-heading. THIS IS WHERE THE MOST SPECIFIC FISERV PRODUCT NAMES LIVE — capture each one verbatim. Examples: 'ATM Driving', 'ATM Driving - Host Connection Fees', 'ATM Driving - Network Charge', 'Check 21', 'Premier Account Processing', 'Cleartouch Account Processing', 'Configure Digital', 'Abiliti', 'Wisdom', 'Signature', 'Portico', 'DNA', 'Statement Advantage', 'Zelle Payment', 'Managed Network', 'OneSpan MFA', 'Intuit Connectivity', 'Positive Pay', 'Bill Payment', 'Mobile Banking', 'ACCEL Network'. Use the MOST DESCRIPTIVE form visible (e.g. 'ATM Driving - Host Connection Fees' rather than just 'ATM Driving' when the document specifies it). Preserve document order. Return [] when no leaf items appear under this sub-header."]
        }
      ]
    }
  ],
  "is_active": true if this document appears to be currently in force, false if terminated or superseded, null if unclear,
  "extraction_confidence": "high" if contract_type and signed_date are clearly stated (signature page legible, dates next to signatures); "medium" if inferred from body text; "low" if ambiguous or unclear,
  "extraction_confidence_score": integer 0-100 indicating your overall confidence in the extracted metadata. Anchor: 90-100 = contract type, signed date, and parties are all stated verbatim and clearly visible; 70-89 = all key fields present but one requires minor inference; 50-69 = fields inferred from body text with some ambiguity; 30-49 = significant ambiguity in type or date; 0-29 = mostly guessed / illegible pages. Return an integer, not a string.
}

Rules:
- If a field is not found, return null or an empty array []. Do NOT guess.
- parent_references must be exact verbatim quotes — copy the text character for character.
- signed_date is THIS document's own signing date only, NOT the effective date and NOT any date referenced for a parent or predecessor contract.
- effective_date is THIS document's own effective/commencement date only, never a referenced date from another contract. May coincide with signed_date.
- internal_doc_codes: STRUCTURED list of {value, position} objects (see the field comment above for the full position vocabulary). Each code MUST have a position label. The downstream resolver treats handwritten_margin / handwritten_inline / printed_stamp / printed_label_box / printed_body as high-value parent-link signals (immune to the boilerplate filter), and printed_header / printed_footer / unknown as low-value (subject to the boilerplate filter that strips client-wide identifiers). Get the position right — it is the difference between a real parent link and a noise match.
- section_structure: a NESTED list of section objects. Each top-level entry is {header, product, subheaders}. Each sub-header is {text, product, items}. Items are the leaf-level product/service names listed directly under the sub-header. Only include true section / sub-section / item-list entries — skip running page headers/footers, page numbers, and inline bold emphasis inside paragraphs. Preserve document order at every level; do not deduplicate semantically distinct headings even if they share words. Return [] when the document has no clear section structure. Use null (NOT empty string) for the `product` field when an entry is generic (Termination, Fees, Definitions, Background, Recitals, etc.). The validator's product dictionary anchors on the names at the deepest level (items inside sub-headers) — that is where the canonical Fiserv product names like "ATM Driving", "Check 21", "Premier Account Processing", "Configure Digital" reliably appear, so scan especially carefully at that level.
- extraction_confidence_score must be internally consistent with extraction_confidence (high→roughly 70-100, medium→40-69, low→0-39).
"""


# ============================================================
# PHASE 1: PDF SCANNING AND IMAGE RENDERING
# ============================================================

def pdf_to_images(pdf_path, dpi=DPI):
    """Render each page of a PDF as a PIL Image. Pattern from team's existing code."""
    doc = fitz.open(str(pdf_path))
    pages = []
    for page_index in range(len(doc)):
        page = doc.load_page(page_index)
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(BytesIO(pix.tobytes("png")))
        pages.append({"page_number": page_index + 1, "image": img})
    doc.close()
    return pages


def docx_to_images(docx_path, dpi=DPI):
    """Render each page of a .docx as a PIL Image. Converts to a temporary PDF
    via MS Word (docx2pdf) and then reuses pdf_to_images so downstream code
    sees a stream of PIL Images indistinguishable from a real PDF."""
    if not _HAS_DOCX2PDF:
        raise RuntimeError(
            "docx2pdf is not installed. Run:  pip install docx2pdf\n"
            "Note: docx2pdf requires MS Word (Windows) or Office (macOS) to perform the conversion."
        )
    with tempfile.TemporaryDirectory() as td:
        tmp_pdf = Path(td) / (Path(docx_path).stem + ".pdf")
        _docx2pdf_convert(str(docx_path), str(tmp_pdf))
        if not tmp_pdf.exists():
            raise RuntimeError(f"docx2pdf did not produce a PDF for {docx_path}")
        return pdf_to_images(tmp_pdf, dpi=dpi)


def tiff_to_images(tiff_path):
    """Load a TIFF file (possibly multi-page) and return its frames as PIL
    Images. Multi-page TIFFs are common for scanned contracts — each frame
    becomes one 'page' in the downstream pipeline."""
    with Image.open(str(tiff_path)) as img:
        pages = []
        for idx, frame in enumerate(ImageSequence.Iterator(img)):
            # Copy before we advance the iterator, and normalise to RGB so
            # palette / CMYK / 1-bit frames all serialise cleanly as PNG.
            pages.append({"page_number": idx + 1, "image": frame.copy().convert("RGB")})
        return pages


def msg_to_images(msg_path, dpi=DPI):
    """Render an Outlook .msg email as a sequence of PIL Images, so the file
    can ride the same vision-extraction pipeline as a PDF.

    Approach: parse the .msg with extract-msg, build a synthetic multi-page
    PDF in memory using PyMuPDF (already a dependency), insert the email's
    printable view (header block + body) using a simple line-paginated text
    layout, and finally reuse pdf_to_images to rasterise each page. The
    downstream LLM call therefore sees the email as a stream of page images
    indistinguishable from a real scanned PDF."""
    if not _HAS_EXTRACT_MSG:
        raise RuntimeError(
            "extract-msg is not installed. Run:  pip install extract-msg\n"
            "Required to parse Outlook .msg files."
        )

    # ── 1. Parse the .msg into a printable text blob ──────────────────
    msg = _extract_msg.Message(str(msg_path))
    try:
        header_lines = []
        if msg.subject: header_lines.append(f"Subject: {msg.subject}")
        if msg.sender:  header_lines.append(f"From: {msg.sender}")
        if msg.to:      header_lines.append(f"To: {msg.to}")
        if msg.cc:      header_lines.append(f"Cc: {msg.cc}")
        if msg.date:    header_lines.append(f"Date: {msg.date}")
        if getattr(msg, "attachments", None):
            names = []
            for a in msg.attachments:
                n = getattr(a, "longFilename", None) or getattr(a, "shortFilename", None) or "(unnamed)"
                names.append(str(n))
            if names:
                header_lines.append(f"Attachments: {', '.join(names)}")
        body = (msg.body or "").strip()
    finally:
        try:
            msg.close()
        except Exception:
            pass

    full_text = "\n".join(header_lines)
    if body:
        full_text = full_text + "\n\n" + ("─" * 60) + "\n\n" + body

    # ── 2. Wrap the text into fixed-width lines for predictable pagination ─
    import textwrap
    wrap_width = 95   # ~chars per line at 11pt courier on a US Letter width
    lines = []
    for paragraph in full_text.split("\n"):
        if not paragraph:
            lines.append("")
            continue
        wrapped = textwrap.wrap(
            paragraph,
            width=wrap_width,
            replace_whitespace=False,
            drop_whitespace=False,
            break_long_words=True,
            break_on_hyphens=False,
        )
        lines.extend(wrapped or [""])

    # ── 3. Build a synthetic PDF, one page worth of lines at a time ───
    page_w, page_h = 612, 792           # US Letter, points (1/72 in)
    margin         = 50
    fontsize       = 11
    line_height    = 15
    fontname       = "courier"           # monospaced → wrap_width is honest
    lines_per_page = max(1, (page_h - 2 * margin) // line_height)

    # PyMuPDF's built-in fonts (courier/helv/times) are Latin-1 only. Sanitise
    # so non-Latin1 chars (smart quotes, em-dashes, etc.) don't crash insert_text.
    def _to_latin1(s):
        return s.encode("latin-1", errors="replace").decode("latin-1")

    doc = fitz.open()
    if not lines:
        lines = [""]
    for i in range(0, len(lines), lines_per_page):
        page  = doc.new_page(width=page_w, height=page_h)
        chunk = lines[i:i + lines_per_page]
        y     = margin + fontsize
        for line in chunk:
            # PyMuPDF's insert_text crashes on empty / whitespace-only strings
            # (it does max(ord(c) for c in text) without an empty-seq guard).
            # Skip those lines explicitly — they only affect vertical spacing,
            # which is preserved by advancing y unconditionally below.
            sanitised = _to_latin1(line)
            if sanitised.strip():
                try:
                    page.insert_text(
                        (margin, y), sanitised,
                        fontsize=fontsize, fontname=fontname,
                    )
                except Exception:
                    # Last-resort net for any other PyMuPDF edge case.
                    pass
            y += line_height

    # ── 4. Rasterise pages exactly the same way pdf_to_images does ────
    pages = []
    for page_index in range(len(doc)):
        page = doc.load_page(page_index)
        mat  = fitz.Matrix(dpi / 72, dpi / 72)
        pix  = page.get_pixmap(matrix=mat)
        img  = Image.open(BytesIO(pix.tobytes("png")))
        pages.append({"page_number": page_index + 1, "image": img})
    doc.close()
    return pages


def txt_to_text(txt_path, max_chars=TXT_MAX_CHARS):
    """Read a .txt file as a string, falling back through common encodings,
    and truncate if over max_chars so the prompt stays bounded."""
    raw = None
    for enc in ("utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            with open(txt_path, "r", encoding=enc) as f:
                raw = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if raw is None:
        # Last resort — bytes read with replacement characters
        with open(txt_path, "rb") as f:
            raw = f.read().decode("utf-8", errors="replace")

    if len(raw) > max_chars:
        raw = raw[:max_chars] + f"\n\n[TRUNCATED — original was {len(raw):,} chars, kept first {max_chars:,}]"
    return raw


def image_to_base64(image):
    """Convert PIL Image to base64 string. Pattern from team's existing code."""
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def extract_date_from_filename(filename):
    """Extract date from filename patterns like MM-DD-YYYY as a fallback."""
    match = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', filename)
    if match:
        m, d, y = match.groups()
        try:
            return datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def scan_contracts(root, only_clients=None):
    """Walk root folder, return list of contract dicts. Client name = subfolder name.
    If only_clients is a non-empty iterable of folder names, client folders whose
    name is not in that set are skipped entirely. If a requested name doesn't
    exist as a subfolder, a warning is printed so typos are visible."""
    contracts = []
    root = Path(root)
    if not root.exists():
        print(f"ERROR: Contracts folder '{root}' not found. Check CONTRACTS_ROOT.")
        return contracts

    # Normalise the filter to a set for O(1) lookups; None / empty = no filter.
    allow = set(only_clients) if only_clients else None
    if allow:
        existing = {p.name for p in root.iterdir() if p.is_dir()}
        missing  = allow - existing
        if missing:
            print(f"  Warning: client filter includes folders that don't exist under "
                  f"'{root}': {sorted(missing)}")

    for client_dir in sorted(root.iterdir()):
        if not client_dir.is_dir():
            continue
        client_name = client_dir.name
        if allow is not None and client_name not in allow:
            continue
        # Accept every supported extension (case-insensitive) from this client folder.
        docs = sorted(
            p for p in client_dir.iterdir()
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS
        )
        if not docs:
            print(f"  Warning: no supported documents found in '{client_dir}' "
                  f"(looked for {sorted(SUPPORTED_EXTS)})")
            continue
        for doc_path in docs:
            contracts.append({
                "client":        client_name,
                "filename":      doc_path.name,
                "filepath":      str(doc_path),
                "file_ext":      doc_path.suffix.lower(),
                "filename_date": extract_date_from_filename(doc_path.name),
                "cache_key":     f"{client_name}/{doc_path.name}",  # scoped to prevent cross-client collisions
            })

    return contracts


# ============================================================
# PHASE 2: LLM METADATA EXTRACTION
# ============================================================

def load_cache(cache_path):
    if Path(cache_path).exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache, cache_path):
    Path(cache_path).parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def parse_llm_json(text):
    """Robustly extract JSON from LLM response — handles markdown code fences."""
    # Strip markdown code fences if present (GPT sometimes wraps output in ```json...```)
    text = re.sub(r'```(?:json)?\s*', '', text).strip().rstrip('`').strip()
    # Find outermost JSON object
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        return json.loads(match.group(0))
    raise ValueError(f"No JSON object found in LLM response: {text[:300]}")


# ============================================================
# RUN METRICS
# ============================================================
# Accumulates token usage / API-call counts across the whole run and tracks
# wall-clock time. Token counts come straight from each API response's `usage`
# block; the model name is read from the response too (not from OPENAI_MODEL),
# so when the Fiserv gateway substitutes its own deployment we report the real
# one. Printed as a summary by print_run_metrics() at the end of main().
class RunMetrics:
    def __init__(self):
        self.input_tokens        = 0
        self.output_tokens       = 0
        self.api_calls           = 0
        self.models_used         = []   # actual model name(s) returned by the API, in first-seen order
        self._start              = None # set by start(); None until the run begins
        self.text_payload_bytes  = 0   # cumulative bytes from text/JSON portions of all requests
        self.image_payload_bytes = 0   # cumulative bytes from base64 image portions of all requests

    def start(self):
        """Begin the run timer (call once, at the top of main())."""
        self._start = time.time()

    def record(self, input_tokens, output_tokens, model, text_bytes=0, image_bytes=0):
        """Fold one successful API response into the running totals."""
        self.api_calls             += 1
        self.input_tokens          += int(input_tokens or 0)
        self.output_tokens         += int(output_tokens or 0)
        self.text_payload_bytes    += int(text_bytes or 0)
        self.image_payload_bytes   += int(image_bytes or 0)
        if model and model not in self.models_used:
            self.models_used.append(model)

    def elapsed_seconds(self):
        return 0.0 if self._start is None else time.time() - self._start


METRICS = RunMetrics()


def _fmt_duration(seconds):
    """Human-readable HH:MM:SS for the run-time line."""
    total = int(round(seconds))
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _fmt_bytes(n):
    """Human-readable byte count: B, KB, or MB."""
    if n < 1_024:
        return f"{n:.0f} B"
    if n < 1_048_576:
        return f"{n / 1_024:.1f} KB"
    return f"{n / 1_048_576:.2f} MB"


def print_run_metrics():
    """Print the end-of-run metric summary requested by the team."""
    if METRICS.models_used:
        model_label = ", ".join(METRICS.models_used)
    else:
        model_label = f"{OPENAI_MODEL} (configured; no live API calls this run)"
    elapsed    = METRICS.elapsed_seconds()
    n_calls    = max(1, METRICS.api_calls)
    total_pay  = METRICS.text_payload_bytes + METRICS.image_payload_bytes
    avg_text   = METRICS.text_payload_bytes  / n_calls
    avg_image  = METRICS.image_payload_bytes / n_calls
    avg_total  = total_pay                   / n_calls
    print(f"\n{'=' * 62}")
    print("  RUN METRICS")
    print(f"{'=' * 62}")
    print(f"  Model used (from API):  {model_label}")
    print(f"  API calls:              {METRICS.api_calls:,}")
    print(f"  Input tokens:           {METRICS.input_tokens:,}")
    print(f"  Output tokens:          {METRICS.output_tokens:,}")
    print(f"  Total tokens:           {METRICS.input_tokens + METRICS.output_tokens:,}")
    print(f"  Payload per API call (text/JSON + images):")
    print(f"    Text/JSON:            {_fmt_bytes(avg_text)}/call  (total: {_fmt_bytes(METRICS.text_payload_bytes)})")
    print(f"    Images:               {_fmt_bytes(avg_image)}/call  (total: {_fmt_bytes(METRICS.image_payload_bytes)})")
    print(f"    Combined:             {_fmt_bytes(avg_total)}/call  (total: {_fmt_bytes(total_pay)})")
    print(f"  Backend:                {OPENAI_BACKEND}")
    print(f"  Run time:               {_fmt_duration(elapsed)}  ({elapsed:.1f}s)")
    print(f"{'=' * 62}")


_openai_client = None

def get_openai_client():
    """Return the shared OpenAI client, lazy-initialised.

    The `timeout` argument is set explicitly so a hung or unusually-slow
    request bounces to the next tenacity retry within a bounded time
    rather than consuming the SDK default (10 minutes per attempt → up
    to ~30 minutes silently if all three retries hang)."""
    global _openai_client
    if _openai_client is None:
        _client_kwargs = {
            "api_key": OPENAI_API_KEY,
            "timeout": API_TIMEOUT_SECONDS,
        }
        # Route through a corporate / Azure-compatible gateway when configured
        # (OPENAI_BASE_URL in .env). Blank → SDK default OpenAI endpoint.
        if OPENAI_BASE_URL:
            _client_kwargs["base_url"] = OPENAI_BASE_URL
        _openai_client = OpenAI(**_client_kwargs)
    return _openai_client


# ============================================================
# BACKEND DISPATCH  (OpenAI Responses API  ↔  Fiserv Foundation gateway)
# ============================================================
# Both entry points (call_llm_chunk / call_llm_text) funnel through
# _llm_complete(), which routes to the active backend, records run metrics
# (tokens / call count / the ACTUAL model name from the response), and returns
# the raw text output.

def _fiserv_purpose_for_model(model_name):
    """Pick the X-Purpose tag for the Foundation gateway. gpt-5.x → GPT5 tag;
    everything else (gpt-4.1, gpt-4o-mini, …) → GPT4 tag."""
    return FISERV_PURPOSE_GPT5 if (model_name or "").lower().startswith("gpt-5") else FISERV_PURPOSE_GPT4


def _fiserv_headers():
    """Headers for the Fiserv Foundation API.

    NOTE: confirm the exact header names against Fiserv's Foundation API docs.
    X-Purpose is required (it selects the model family). The email and auth
    headers below are best-guess defaults — adjust the header *names* here if
    the gateway expects something different (e.g. 'Authorization: Bearer <token>'
    instead of 'api-key', or 'X-User-Email' instead of 'X-Email')."""
    headers = {
        "Content-Type": "application/json",
        "X-Purpose": _fiserv_purpose_for_model(OPENAI_MODEL),
    }
    if FISERV_EMAIL:
        headers["X-Email"] = FISERV_EMAIL
    if OPENAI_API_KEY:
        headers["api-key"] = OPENAI_API_KEY          # Azure-style key header
    return headers


def _fiserv_complete(user_content_text, images_b64):
    """One call to the Fiserv Foundation chat/completions gateway."""
    user_content = [{"type": "text", "text": user_content_text}]
    for b64 in images_b64:
        user_content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{b64}"},
        })
    payload = {
        "model": OPENAI_MODEL,            # gateway may override the model via X-Purpose
        "temperature": 0,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_content},
        ],
    }
    text_bytes  = len(SYSTEM_PROMPT.encode("utf-8")) + len(user_content_text.encode("utf-8"))
    image_bytes = sum(len(b64) for b64 in images_b64)
    resp = httpx.post(
        FOUNDATION_API_URL,
        json=payload,
        headers=_fiserv_headers(),
        timeout=API_TIMEOUT_SECONDS,
    )
    resp.raise_for_status()
    data  = resp.json()
    usage = data.get("usage") or {}
    METRICS.record(usage.get("prompt_tokens", 0), usage.get("completion_tokens", 0),
                   data.get("model"), text_bytes, image_bytes)
    return data["choices"][0]["message"]["content"]


def _openai_responses_complete(user_content_text, images_b64):
    """One call to the OpenAI Responses API (default backend) — unchanged shape."""
    user_content = [{"type": "input_text", "text": user_content_text}]
    for b64 in images_b64:
        user_content.append({
            "type": "input_image",
            "image_url": f"data:image/png;base64,{b64}",
        })
    text_bytes  = len(SYSTEM_PROMPT.encode("utf-8")) + len(user_content_text.encode("utf-8"))
    image_bytes = sum(len(b64) for b64 in images_b64)
    client = get_openai_client()
    response = client.responses.create(
        model=OPENAI_MODEL,
        temperature=0,
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_content},
        ],
    )
    usage   = getattr(response, "usage", None)
    in_tok  = getattr(usage, "input_tokens", 0) if usage else 0
    out_tok = getattr(usage, "output_tokens", 0) if usage else 0
    METRICS.record(in_tok, out_tok, getattr(response, "model", OPENAI_MODEL), text_bytes, image_bytes)
    return response.output_text


def _llm_complete(user_content_text, images_b64=()):
    """Backend-agnostic single completion. Returns the model's text output and
    records run metrics as a side effect."""
    if OPENAI_BACKEND == "fiserv":
        return _fiserv_complete(user_content_text, list(images_b64))
    return _openai_responses_complete(user_content_text, list(images_b64))


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, min=4, max=30),
    retry=retry_if_exception_type(Exception),
)
def call_llm_chunk(pages_b64, chunk_num, total_chunks, contract_name):
    """Send a chunk of contract pages (as images) to the active backend."""
    user_text = (
        f"Analyze this legal contract: '{contract_name}'. "
        f"These are pages {(chunk_num - 1) * CHUNK_SIZE + 1}–{chunk_num * CHUNK_SIZE} "
        f"(chunk {chunk_num} of {total_chunks}). "
        "Extract the metadata as specified. Return ONLY the JSON object."
    )
    return _llm_complete(user_text, pages_b64)


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, min=4, max=30),
    retry=retry_if_exception_type(Exception),
)
def call_llm_text(contract_text, contract_name):
    """Text-only variant for .txt documents. Passes the full file body (already
    truncated by txt_to_text) as a single text block; no images are sent."""
    user_text = (
        f"Analyze this legal contract: '{contract_name}'. "
        "The complete contract text follows. Extract the metadata as specified. "
        "Return ONLY the JSON object.\n\n"
        f"--- BEGIN CONTRACT TEXT ---\n{contract_text}\n--- END CONTRACT TEXT ---"
    )
    return _llm_complete(user_text)


# ── internal_doc_codes — structured-format helpers ───────────────────────
# As of May 12 2026 the LLM is asked to return each code as
#   {"value": "<string>", "position": "<one of the position labels>"}
# so the resolver can tell a handwritten margin note (validator-flagged
# parent-link signal) from a printed-everywhere header / footer (boilerplate).
#
# Old caches still contain bare-string codes from earlier runs; for backward
# compat the normalizer below accepts both shapes and emits a uniform list
# of {value, position} dicts. Bare strings become position="unknown" so the
# boilerplate filter (which only spares high-value positions) still treats
# them conservatively.

_HIGH_VALUE_CODE_POSITIONS = {
    "handwritten_margin",
    "handwritten_inline",
    "printed_stamp",
    "printed_label_box",
    "printed_body",
}
_LOW_VALUE_CODE_POSITIONS = {
    "printed_header",
    "printed_footer",
    "unknown",
}
_VALID_CODE_POSITIONS = _HIGH_VALUE_CODE_POSITIONS | _LOW_VALUE_CODE_POSITIONS


def _normalize_doc_code_entries(raw):
    """Normalise an internal_doc_codes payload to a list of
    {"value": str, "position": str} dicts.

    Accepts:
      • The new structured format: list of dicts already shaped that way.
        Invalid `position` values are coerced to "unknown" so a hallucinated
        label can't break downstream logic.
      • The old flat-string format: list of plain strings. Each becomes
        {"value": s, "position": "unknown"}.
      • Anything else returns []."""
    if not raw or not isinstance(raw, list):
        return []
    out = []
    for entry in raw:
        if isinstance(entry, str):
            v = entry.strip()
            if v:
                out.append({"value": v, "position": "unknown"})
            continue
        if isinstance(entry, dict):
            v = entry.get("value")
            if not isinstance(v, str):
                continue
            v = v.strip()
            if not v:
                continue
            p = entry.get("position")
            if not isinstance(p, str):
                p = "unknown"
            p = p.strip().lower()
            if p not in _VALID_CODE_POSITIONS:
                p = "unknown"
            out.append({"value": v, "position": p})
    return out


# ── Party normalisation + agreement-number token extraction ──────────────
# Used by the two May 22 2026 refinements:
#   • find_parent_by_references partition-by-parties: normalised parties so
#     "Fiserv Solutions, Inc." and "Fiserv Solutions, LLC" land in the same
#     partition (Kearny Bank Bucket A — Clifton MSA cluster).
#   • find_parent_by_doc_codes numeric-token matching: linking high-value-
#     position codes like "Contract 2035" to "Data Processing Services
#     Agreement #2035" via the shared 4-digit agreement number (Kearny Bank
#     Bucket B — Millington Migrated chain).

_CORPORATE_SUFFIX_RE = re.compile(
    r"\b("
    r"solutions|incorporated|incorporation|inc|llc|ltd|limited|corp|corporation|"
    r"company|co|n\.a\.|na|l\.l\.c\.|p\.c\.|pllc|plc|gmbh|s\.a\.|s\.r\.l\.|"
    r"holdings|group|services"
    r")\b",
    re.IGNORECASE,
)
_PARTY_PUNCT_RE = re.compile(r"[,.\-_/]+")


def _normalize_party_for_match(s):
    """Aggressive party normalisation for set-membership comparison. Strips
    common corporate-form suffixes (LLC, Inc., Solutions, Corp, …) plus
    light punctuation, lowercases, collapses whitespace. The goal is that
    'Fiserv Solutions, Inc.' and 'Fiserv Solutions, LLC' and 'Fiserv'
    all map to the same canonical token. Returns '' when input is empty
    or non-string."""
    if not isinstance(s, str):
        return ""
    n = s.lower().strip()
    n = _PARTY_PUNCT_RE.sub(" ", n)
    n = _CORPORATE_SUFFIX_RE.sub(" ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def _norm_party_set(parties):
    """Return the frozenset of normalised, non-empty party tokens from a
    raw parties list. Skips empty / non-string elements."""
    return frozenset(
        n for n in (_normalize_party_for_match(p) for p in (parties or []))
        if n
    )


def _extract_high_value_numeric_tokens(entries):
    """Pull 3+ digit numeric runs out of internal_doc_codes ENTRIES, but
    only from entries whose position is in _HIGH_VALUE_CODE_POSITIONS
    (handwritten_margin / handwritten_inline / printed_stamp /
    printed_label_box / printed_body). These positions are the validator-
    confirmed parent-link signal — the LLM saw the number where it was
    meaningfully written, not as part of a running header/footer
    boilerplate. Low-value positions (printed_header / printed_footer /
    unknown) are skipped so that page numbers, DocuSign envelope IDs, and
    client-wide identifiers don't pollute the numeric-token keyspace.

    Returns a set of digit-only strings (e.g. {'2035', '1115102'}).

    A previous attempt to enable fuzzy numeric matching across ALL code
    positions (May 12 2026) was reverted because boilerplate codes like
    'Client #290' in page headers collided with legitimate handwritten
    '290' agreement numbers in margins. The position filter here is what
    makes the same approach safe to enable now."""
    tokens = set()
    for entry in entries or []:
        if not isinstance(entry, dict):
            continue
        if entry.get("position") not in _HIGH_VALUE_CODE_POSITIONS:
            continue
        v = entry.get("value")
        if not isinstance(v, str):
            continue
        for m in re.finditer(r"\d{3,}", v):
            tokens.add(m.group(0))
    return tokens


# ── Section-structure helpers (May 22 2026) ──────────────────────────────
# As of this date the LLM is asked to return section information as a nested
# 3-level structure:
#   section_structure = [
#     {"header": str, "product": str|None,
#      "subheaders": [{"text": str, "product": str|None, "items": [str, …]}, …]},
#     …
#   ]
# This lets the product canonicaliser see the leaf-level product names (the
# bullet-points and table-row labels under each sub-header) that the old
# flat top-level-only structure missed.
#
# Older cache entries store the flat 2-array form:
#   section_headers          = [str, …]
#   section_header_products  = [str|None, …]   (parallel)
# The normalisation helper below accepts both shapes and always returns the
# nested form, so downstream code (resolver, Excel export, hover tooltips)
# can work off a single representation.


def _normalize_section_structure(meta):
    """Return a list of section objects in the nested 3-level shape:
        [
          {"header": str,
           "product": str|None,
           "subheaders": [
             {"text": str, "product": str|None, "items": [str, ...]},
             ...
           ]},
          ...
        ]

    Accepts:
      • The new structured `section_structure` field (returned as-is, with
        validation/coercion so each level is a list / dict / string as expected).
      • The old flat `section_headers` + `section_header_products` pair —
        synthesised into the nested shape with empty subheaders for each entry.

    Returns [] when neither is present."""
    # Prefer the new field
    raw = meta.get("section_structure") if isinstance(meta, dict) else None
    if isinstance(raw, list) and raw:
        out = []
        for entry in raw:
            if not isinstance(entry, dict):
                continue
            header  = entry.get("header") or ""
            if not isinstance(header, str) or not header.strip():
                continue
            product = entry.get("product")
            if isinstance(product, str):
                product = product.strip() or None
                if product and product.lower() in ("null", "none", "n/a"):
                    product = None
            elif product is not None and not isinstance(product, str):
                product = None
            subs_raw = entry.get("subheaders") or []
            subs = []
            if isinstance(subs_raw, list):
                for sub in subs_raw:
                    if not isinstance(sub, dict):
                        continue
                    text = sub.get("text") or ""
                    if not isinstance(text, str) or not text.strip():
                        continue
                    sub_product = sub.get("product")
                    if isinstance(sub_product, str):
                        sub_product = sub_product.strip() or None
                        if sub_product and sub_product.lower() in ("null", "none", "n/a"):
                            sub_product = None
                    elif sub_product is not None and not isinstance(sub_product, str):
                        sub_product = None
                    items_raw = sub.get("items") or []
                    items = []
                    if isinstance(items_raw, list):
                        for it in items_raw:
                            if isinstance(it, str) and it.strip():
                                items.append(it.strip())
                    subs.append({"text": text.strip(), "product": sub_product, "items": items})
            out.append({"header": header.strip(), "product": product, "subheaders": subs})
        return out

    # Fallback — synthesise from the legacy flat arrays so existing cache
    # entries continue to work without re-extraction.
    headers  = (meta or {}).get("section_headers") or []
    products = (meta or {}).get("section_header_products") or []
    if not isinstance(headers, list):
        return []
    if not isinstance(products, list):
        products = []
    out = []
    seen = set()
    for i, h in enumerate(headers):
        if not isinstance(h, str) or not h.strip():
            continue
        h_clean = h.strip()
        if h_clean in seen:
            continue
        seen.add(h_clean)
        p = products[i] if i < len(products) else None
        if isinstance(p, str):
            p = p.strip() or None
            if p and p.lower() in ("null", "none", "n/a"):
                p = None
        elif p is not None and not isinstance(p, str):
            p = None
        # Legacy entries have no sub-header / item information — supply an
        # empty list so downstream code sees the same shape.
        out.append({"header": h_clean, "product": p, "subheaders": []})
    return out


def _walk_section_products(structure):
    """Walk a normalised section_structure tree and yield every product
    name found at any level — top-level header product, sub-header product,
    and leaf-item text. Used to feed the product canonicaliser, which then
    aligns each name against the Fiserv dictionary."""
    if not structure:
        return
    for entry in structure:
        if not isinstance(entry, dict):
            continue
        p = entry.get("product")
        if isinstance(p, str) and p.strip():
            yield p.strip()
        for sub in entry.get("subheaders") or []:
            if not isinstance(sub, dict):
                continue
            sp = sub.get("product")
            if isinstance(sp, str) and sp.strip():
                yield sp.strip()
            # Leaf-item names are themselves product names — the deepest
            # and most specific signal in this corpus, per the validator.
            for it in sub.get("items") or []:
                if isinstance(it, str) and it.strip():
                    yield it.strip()


def _render_section_structure_text(structure):
    """Render a normalised section_structure as a single human-readable
    multi-line string, suitable for the Excel `Section_Headers` column and
    the HTML hover tooltip.

    Shape:
        • Top-level header [Product]
            — Sub-header [Sub-product] :: item1, item2, item3
            — Sub-header without items
        • Next top-level header
    Products in [brackets] are the canonicalised dictionary names; items
    after the '::' are the leaf-level product names extracted from the
    document. Empty-product brackets are dropped so the line stays clean."""
    if not structure:
        return ""
    lines = []
    for entry in structure:
        if not isinstance(entry, dict):
            continue
        header  = entry.get("header") or ""
        product = entry.get("product")
        line = f"• {header}"
        if isinstance(product, str) and product.strip():
            line += f" [{product.strip()}]"
        lines.append(line)
        for sub in entry.get("subheaders") or []:
            if not isinstance(sub, dict):
                continue
            text     = sub.get("text") or ""
            sub_prod = sub.get("product")
            sub_line = f"    — {text}"
            if isinstance(sub_prod, str) and sub_prod.strip():
                sub_line += f" [{sub_prod.strip()}]"
            items = [it.strip() for it in (sub.get("items") or [])
                     if isinstance(it, str) and it.strip()]
            if items:
                sub_line += " :: " + ", ".join(items)
            lines.append(sub_line)
    return "\n".join(lines)


def _dedupe_doc_code_entries(entries):
    """Dedupe a list of {value, position} dicts by `value`, keeping the
    highest-value position seen for each value. When a code shows up multiple
    times across chunks with different positions, we keep the most
    informative one (handwritten_margin > printed_label_box > printed_body
    > printed_header > unknown)."""
    rank = {
        "handwritten_margin":   0,
        "handwritten_inline":   1,
        "printed_stamp":        2,
        "printed_label_box":    3,
        "printed_body":         4,
        "printed_header":       5,
        "printed_footer":       6,
        "unknown":              7,
    }
    best_by_value = {}
    for e in entries:
        v = e["value"]
        cur = best_by_value.get(v)
        if cur is None or rank.get(e["position"], 99) < rank.get(cur["position"], 99):
            best_by_value[v] = e
    # Preserve first-seen order
    seen = set()
    out = []
    for e in entries:
        if e["value"] in seen:
            continue
        seen.add(e["value"])
        out.append(best_by_value[e["value"]])
    return out


def merge_chunk_results(results):
    """Merge metadata from multiple chunks into one record.
    First chunk wins for type/date/parties (preamble); all chunks aggregated for hierarchy signals."""
    if not results:
        return {}

    merged = dict(results[0])

    # Aggregate signals that can appear anywhere in the contract body
    all_refs       = []
    all_supersedes = []
    all_doc_codes  = []

    # Section structure — accept both the new nested 3-level format
    # (section_structure: list of {header, product, subheaders}) AND the old
    # flat 2-array format (section_headers / section_header_products).
    # Within a chunk-merge we key sub-clusters by top-level header text so
    # repeated headers across chunks are deduped, and within a header the
    # sub-headers and items are deduped by their own text (preserving order).
    section_map = {}                 # header_text → {"product": .., "subs": {sub_text → {"product": .., "items": [..]}}}
    section_header_order = []        # ordered list of header_texts as first seen
    section_sub_order    = {}        # header_text → ordered list of sub_texts

    for r in results:
        refs = r.get("parent_references") or []
        if isinstance(refs, list):
            all_refs.extend(refs)

        sup = r.get("supersedes_text")
        if sup and isinstance(sup, str) and sup.strip():
            all_supersedes.append(sup)

        # internal_doc_codes — accept both the new structured format (list of
        # {value, position} dicts) AND the old flat-string format. Normalise
        # to the structured format for downstream merging.
        all_doc_codes.extend(_normalize_doc_code_entries(r.get("internal_doc_codes")))

        # Normalise the chunk's section info into the nested 3-level form
        # (handles both new nested and legacy flat formats) and merge.
        chunk_sections = _normalize_section_structure(r)
        for entry in chunk_sections:
            header  = entry["header"]
            product = entry.get("product")
            if header not in section_map:
                section_map[header] = {"product": product, "subs": {}}
                section_header_order.append(header)
                section_sub_order[header] = []
            else:
                # Upgrade null → product if a later chunk found one
                if section_map[header]["product"] is None and product:
                    section_map[header]["product"] = product
            sub_dict = section_map[header]["subs"]
            sub_order = section_sub_order[header]
            for sub in entry.get("subheaders") or []:
                text     = sub["text"]
                sub_prod = sub.get("product")
                items    = sub.get("items") or []
                if text not in sub_dict:
                    sub_dict[text] = {"product": sub_prod, "items": []}
                    sub_order.append(text)
                else:
                    if sub_dict[text]["product"] is None and sub_prod:
                        sub_dict[text]["product"] = sub_prod
                # Items — extend with dedup (preserve first-seen order)
                seen_items = set(sub_dict[text]["items"])
                for it in items:
                    if it not in seen_items:
                        sub_dict[text]["items"].append(it)
                        seen_items.add(it)

    merged["parent_references"]       = list(dict.fromkeys(filter(None, all_refs)))  # dedupe, preserve order
    merged["supersedes_text"]         = all_supersedes[0] if all_supersedes else None
    merged["internal_doc_codes"]      = _dedupe_doc_code_entries(all_doc_codes)

    # Emit the merged structure as section_structure (the new authoritative
    # field) and ALSO populate the legacy flat fields for backward compat —
    # so any code path still reading section_headers / section_header_products
    # keeps working without change.
    merged_structure = []
    legacy_headers   = []
    legacy_products  = []
    for header in section_header_order:
        slot = section_map[header]
        subs = []
        for sub_text in section_sub_order[header]:
            s = slot["subs"][sub_text]
            subs.append({"text": sub_text, "product": s["product"], "items": list(s["items"])})
        merged_structure.append({"header": header, "product": slot["product"], "subheaders": subs})
        legacy_headers.append(header)
        legacy_products.append(slot["product"])

    merged["section_structure"]       = merged_structure
    merged["section_headers"]         = legacy_headers     # legacy mirror
    merged["section_header_products"] = legacy_products    # legacy mirror

    # Take highest confidence found across any chunk
    conf_order = {"high": 3, "medium": 2, "low": 1, None: 0}
    merged["extraction_confidence"] = max(
        (r.get("extraction_confidence") for r in results),
        key=lambda c: conf_order.get(c, 0)
    )

    # Take max of extraction_confidence_score across chunks. Missing or
    # non-numeric values are ignored; if none of the chunks returned a valid
    # score, the field is omitted and the downstream code falls back to a
    # categorical→midpoint mapping via _get_numeric_extraction_score().
    numeric_scores = []
    for r in results:
        s = r.get("extraction_confidence_score")
        if isinstance(s, (int, float)) and 0 <= s <= 100:
            numeric_scores.append(int(s))
    if numeric_scores:
        merged["extraction_confidence_score"] = max(numeric_scores)

    return merged


def extract_metadata(contract, cache):
    """Extract metadata for one contract, using cache if available.

    Dispatches by file extension:
      .pdf                 → render pages as images, call vision model in chunks
      .docx                → convert to temp PDF (docx2pdf), same pipeline as PDF
      .tif / .tiff         → read frames as images, same pipeline as PDF
      .msg                 → render email header+body as a synthetic PDF, then
                             same pipeline as PDF (vision model sees the email
                             as a sequence of page images)
      .txt                 → read text directly, call text-only LLM variant (1 call)
    """
    key = contract["cache_key"]
    if key in cache:
        return cache[key]

    ext = (contract.get("file_ext") or Path(contract["filepath"]).suffix).lower()

    try:
        if ext == ".txt":
            # --- Text-only path ---------------------------------------------
            text = txt_to_text(contract["filepath"])
            raw  = call_llm_text(text, contract["filename"])
            parsed = parse_llm_json(raw)
            result = merge_chunk_results([parsed])
            # Treat each ~3000 chars as roughly one page for reporting parity.
            result["num_pages"]         = max(1, math.ceil(len(text) / 3000))
            result["extraction_failed"] = False
        else:
            # --- Image-based path (PDF / DOCX / TIFF) -----------------------
            if ext == ".pdf":
                pages = pdf_to_images(contract["filepath"])
            elif ext == ".docx":
                pages = docx_to_images(contract["filepath"])
            elif ext in (".tif", ".tiff"):
                pages = tiff_to_images(contract["filepath"])
            elif ext == ".msg":
                pages = msg_to_images(contract["filepath"])
            else:
                raise ValueError(f"Unsupported file type for extraction: {ext}")

            num_pages = len(pages)

            # Cap at MAX_CHUNKS — exhibit pages (Schedule A/B/C) have no hierarchy signals
            total_chunks  = min(math.ceil(num_pages / CHUNK_SIZE), MAX_CHUNKS)
            chunk_results = []

            for chunk_idx in range(total_chunks):
                start     = chunk_idx * CHUNK_SIZE
                end       = min(start + CHUNK_SIZE, num_pages)
                pages_b64 = [image_to_base64(p["image"]) for p in pages[start:end]]

                raw    = call_llm_chunk(pages_b64, chunk_idx + 1, total_chunks, contract["filename"])
                parsed = parse_llm_json(raw)
                chunk_results.append(parsed)

                if chunk_idx < total_chunks - 1:
                    time.sleep(API_DELAY)

            result = merge_chunk_results(chunk_results)
            result["num_pages"]         = num_pages
            result["extraction_failed"] = False

    except Exception as e:
        print(f"\n    ERROR extracting {contract['filename']}: {e}")
        result = {
            "contract_type":        "Unknown",
            "amendment_number":     None,
            "signed_date":          None,
            "effective_date":       None,
            "parties":              [],
            "parent_references":    [],
            "supersedes_text":      None,
            "internal_doc_codes":   [],
            "section_structure":    [],
            "section_headers":      [],
            "section_header_products": [],
            "is_active":            None,
            "extraction_confidence": "low",
            "extraction_failed":    True,
            "error_message":        str(e),
        }

    # Mark this extraction as having been through the strengthened prompt.
    # Both rescue triggers use these flags to avoid re-evicting an entry
    # that has already been through a rescue attempt.
    result["parent_rescue_attempted"]             = True
    result["type_rescue_attempted"]               = True
    result["doc_codes_format_rescue_attempted"]   = True
    result["section_structure_rescue_attempted"] = True

    cache[key] = result
    return result


def _is_stale_child_extraction(cached_entry):
    """Return True if a cached extraction is a candidate for parent-references
    rescue: the contract is a known child-only type, the previous extraction
    succeeded but found no parent text, and we haven't already attempted a
    rescue. Used by extract_metadata to evict stale entries so they get
    re-extracted with the strengthened prompt."""
    if not isinstance(cached_entry, dict):
        return False
    if cached_entry.get("extraction_failed"):
        return False
    if cached_entry.get("parent_rescue_attempted"):
        return False
    if cached_entry.get("parent_references"):
        return False
    raw_type = cached_entry.get("contract_type") or ""
    return normalize_type(raw_type) in _CHILD_ONLY_TYPES


# ── Type-mismatch rescue (validator-recommended, May 2026) ───────────────────
# When a single cache entry's contract_type is wrong, it cascades: every child
# downstream of that contract loses its anchor and ends up Orphan. Concrete
# example seen during the last validator-agreement audit:
#
#   • PEOPLES BANK THE_Master Agreement_03-01-2016_1389673.pdf
#     - filename keyword:  MSA
#     - LLM contract_type: "Amendment"   ← extraction error
#     - Effect: ~20 downstream amendments cite this doc by date but can't
#       resolve it as a parent, because the resolver only accepts MSAs (and
#       root-eligible siblings) as parents of amendments.
#
# Fixing one cached entry like this fixes a dozen mis-resolved children, so
# we trigger a targeted re-extraction whenever the LLM's contract_type
# strongly disagrees with a high-confidence filename keyword. Conservative
# trigger conditions:
#
#   1. Filename keyword maps to MSA (any of "Master Agreement",
#      "Master Services Agreement", "Relationship Agreement",
#      "Network Membership Agreement", "Membership Agreement",
#      "Amended and Restated…") AND the LLM-extracted type is a child-only
#      type (Amendment / Sub-Amendment / Appendix / Attachment / SOW /
#      Schedule / Order Form / etc.). This is the cleanest signal — a
#      filename including "Master Agreement" is almost never wrong, and a
#      child-only LLM verdict on top of it is exactly the cascade failure
#      mode we care about.
#
#   2. Filename keyword maps to a specific root-eligible NON-MSA subtype
#      (Software / Hardware / Services / License / Maintenance / Marketing /
#      Referral / Purchase / Network / Membership Agreement) AND the LLM
#      type is a child-only type. Same rationale — these filename keywords
#      are unambiguous in this corpus, and the LLM dropping to a child-only
#      verdict is suspicious enough to warrant a re-read.
#
# Deliberately NOT triggered when:
#   • Filename keyword is "Amendment" and LLM said MSA — that's the
#     opposite cascade and re-extraction tends to flip-flop. The validator
#     specifically called out this case (file 1494127) as one the LLM
#     should figure out on its own with the strengthened prompt.
#   • The cache entry already has type_rescue_attempted = True — never
#     loop on the same row.
#   • Extraction failed previously — no point asking the LLM again.
_MSA_FILENAME_KEYWORDS = {"MSA"}    # filename keyword set that maps to MSA
_TYPE_RESCUE_FILENAME_KEYWORDS = (
    _MSA_FILENAME_KEYWORDS | {
        "Services Agreement", "Software Agreement", "License Agreement",
        "Hardware Agreement", "Maintenance Agreement", "Marketing Agreement",
        "Referral Agreement", "Purchase Agreement", "Network Agreement",
        "Membership Agreement",
    }
)


def _is_flat_doc_codes_extraction(cached_entry):
    """Return True if the cache entry has internal_doc_codes in the OLD
    flat-string format (list of strings) instead of the structured
    {value, position} format introduced May 12 2026. Targets entries
    that need re-extraction so the resolver can use position-aware
    matching (handwritten margin notes ARE valid parent-link signals
    even when frequent; client-wide printed headers ARE NOT).

    Triggers only when the entry has a non-empty list of plain strings.
    Empty lists and entries already in the structured format are left
    alone — empty means the LLM genuinely found no codes (re-extraction
    won't help), and the structured format is already what we want.

    A separate flag (`doc_codes_format_rescue_attempted`) prevents
    re-eviction once this rescue has been attempted on a given entry."""
    if not isinstance(cached_entry, dict):
        return False
    if cached_entry.get("extraction_failed"):
        return False
    if cached_entry.get("doc_codes_format_rescue_attempted"):
        return False
    codes = cached_entry.get("internal_doc_codes")
    if not isinstance(codes, list) or not codes:
        return False
    # Old format: every entry is a string. New format: every entry is a dict.
    has_string = any(isinstance(c, str) for c in codes)
    has_dict   = any(isinstance(c, dict) for c in codes)
    return has_string and not has_dict


def _is_flat_section_structure_extraction(cached_entry):
    """Return True if the cache entry was extracted under the OLD flat
    section format (parallel `section_headers` + `section_header_products`
    arrays only) and has never been re-extracted under the NEW nested
    3-level format (`section_structure`) introduced May 22 2026.

    The flat format only captured top-level section headings and asked the
    LLM what product each one referred to. The nested format additionally
    captures sub-headers and the leaf items underneath them — which is
    where the most specific Fiserv product names (ATM Driving, Configure
    Digital, Premier Account Processing, etc.) actually appear in this
    corpus per the validator's-team confirmation. Evicting flat-format
    entries forces re-extraction under the richer schema so the canonical
    product list, the product-overlap parent finder, and the Excel /
    HTML rendering all see the deeper signal.

    Triggers only when the entry already produced SOME section data under
    the old format (i.e. `section_headers` is a non-empty list). Empty
    section data means the LLM genuinely found no sections — re-extraction
    won't change that. Entries already carrying the nested
    `section_structure` field or marked with
    `section_structure_rescue_attempted` are left alone.

    A separate flag (`section_structure_rescue_attempted`) prevents
    re-eviction once this rescue has been attempted on a given entry."""
    if not isinstance(cached_entry, dict):
        return False
    if cached_entry.get("extraction_failed"):
        return False
    if cached_entry.get("section_structure_rescue_attempted"):
        return False
    # If the entry already has the new nested field, no rescue needed.
    if isinstance(cached_entry.get("section_structure"), list) and cached_entry.get("section_structure"):
        return False
    # Trigger only when the entry actually has flat section data — empty
    # extractions get nothing from re-extraction.
    headers = cached_entry.get("section_headers")
    if not isinstance(headers, list) or not headers:
        return False
    return True


def _is_type_mismatched_extraction(cached_entry, cache_key):
    """Return True if the LLM's cached contract_type strongly disagrees with
    a confident filename keyword, AND we haven't already attempted a
    type-rescue on this row. Targets the cascade-failure mode where one
    mis-typed root agreement orphans every downstream amendment.

    cache_key is "<client>/<filename>"; the filename portion is what we
    inspect with infer_type_from_filename.
    """
    if not isinstance(cached_entry, dict):
        return False
    if cached_entry.get("extraction_failed"):
        return False
    if cached_entry.get("type_rescue_attempted"):
        return False

    # Extract the filename out of "<client>/<filename>"
    filename = cache_key.split("/", 1)[1] if "/" in cache_key else cache_key
    fn_type  = infer_type_from_filename(filename)
    if fn_type not in _TYPE_RESCUE_FILENAME_KEYWORDS:
        return False

    llm_type = normalize_type(cached_entry.get("contract_type") or "")
    # Trigger condition: filename strongly suggests a root-eligible type,
    # but the LLM returned a child-only type.
    if llm_type in _CHILD_ONLY_TYPES:
        return True

    # Trigger condition: filename says MSA but the LLM returned anything
    # other than MSA (catches the 1389673 case where LLM said "Amendment"
    # AND the case where LLM said "Standalone" for a clear MSA filename).
    if fn_type in _MSA_FILENAME_KEYWORDS and llm_type != "MSA":
        return True

    return False


# ============================================================
# PHASE 3: HIERARCHY RESOLUTION
# ============================================================

# ---- Numeric confidence scoring ---------------------------------------------
# The LLM now returns an integer 0-100 in `extraction_confidence_score`, but
# old cache entries predate that field. `_get_numeric_extraction_score` falls
# back to a midpoint mapping from the categorical `extraction_confidence` so
# both shapes of cache are supported without forced re-extraction.
_CATEGORICAL_TO_NUMERIC = {
    "high":        85,
    "medium-high": 72,
    "medium":      55,
    "low":         25,
    None:          0,
}

# Base score per hierarchy-resolution method. Averaged with the extraction
# score to produce the composite `hierarchy_confidence_score`.
_METHOD_BASE_SCORES = {
    "root":          95,    # MSA — unconditional root
    "miscellaneous": 95,    # non-MSA root-eligible type with no parent in
                            # dataset (renamed from 'standalone' May 26 2026
                            # so the Excel Hierarchy_Method column matches the
                            # renamed Contract_Type label)
    "reference":     90,    # parent resolved via textual reference
    "doc_code":      60,    # parent resolved via shared internal document codes
    "products":      45,    # gray-area: product-overlap fallback
    "duplicate":     90,    # canonical-duplicate linkage (Pass 3); high because the
                             # rule is deterministic — identical type + date + parties
    "date":          30,    # date proximity (no longer used; retained for legacy cache scoring)
    "orphan":         0,    # no parent could be resolved
}

# Bucket thresholds matching the target screenshot's legend
# ("High ≥70%", "Medium 40–69%", "Low <40%").
_BUCKET_HIGH_MIN   = 70
_BUCKET_MEDIUM_MIN = 40


def _get_numeric_extraction_score(metadata):
    """Return extraction score as an int 0-100. Prefers the LLM-supplied
    `extraction_confidence_score`; falls back to a midpoint derived from the
    categorical `extraction_confidence` for cache entries that lack the score."""
    if not metadata:
        return 0
    s = metadata.get("extraction_confidence_score")
    if isinstance(s, (int, float)) and 0 <= s <= 100:
        return int(s)
    return _CATEGORICAL_TO_NUMERIC.get(metadata.get("extraction_confidence"), 0)


def _score_to_bucket(score):
    """Map a 0-100 score to the three-level confidence bucket used by the
    visualization colour scheme. Returns None for missing scores."""
    if score is None:
        return None
    if score >= _BUCKET_HIGH_MIN:
        return "high"
    if score >= _BUCKET_MEDIUM_MIN:
        return "medium"
    return "low"


def parse_date_str(date_str):
    """Parse YYYY-MM-DD string to Python date object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def extract_dates_from_text(text):
    """Extract all recognizable dates from a text string (for fuzzy parent matching)."""
    if not text:
        return []
    dates = []
    patterns_formats = [
        # re.IGNORECASE: catch "february 4, 2009" lowercase in parent reference quotes
        (r'\b([A-Za-z]+ \d{1,2},\s*\d{4})\b',     ['%B %d, %Y', '%b %d, %Y'], re.IGNORECASE),
        (r'\b(\d{1,2}/\d{1,2}/\d{4})\b',           ['%m/%d/%Y'],               0),
        (r'\b(\d{4}-\d{2}-\d{2})\b',               ['%Y-%m-%d'],               0),
        (r'\b(\d{1,2}-\d{1,2}-\d{4})\b',           ['%m-%d-%Y'],               0),  # filename-style: 7-1-2012
    ]
    for pattern, formats, flags in patterns_formats:
        for m in re.finditer(pattern, text, flags):
            raw = m.group(1).strip()
            # Normalize case: strptime requires title-case month names
            raw_norm = raw[0].upper() + raw[1:] if raw else raw
            for fmt in formats:
                try:
                    dates.append(datetime.strptime(raw_norm, fmt).date())
                    break
                except ValueError:
                    continue
    return dates


def dates_close(d1, d2, tolerance_days=7):
    """Return True if two dates are within tolerance_days of each other.
    Required: confirmed date discrepancy in real contracts (Feb 1 effective vs Feb 4 referenced)."""
    if d1 is None or d2 is None:
        return False
    return abs((d1 - d2).days) <= tolerance_days


def find_parent_by_references(contract, by_client):
    """Priority 1: match LLM-extracted parent_references text to other contracts by date.
    Collects ALL date matches then picks the best by type preference (MSA > Amendment > ...).
    Returns (parent_key, confidence, rationale, is_ambiguous).

    Tie-break rule (validator-confirmed, May 2026):
        When multiple DISTINCT contracts share the top type-rank (e.g. two
        Master Agreements with the same effective date both match the cited
        date), this function REFUSES to pick — the contract is left
        unparented and will fall through to Orphan. Silently choosing one of
        the candidates was identified by the validator as a frequent source
        of wrong-parent errors ("Incorrect Agreement connected as Parent.
        Parent MSA missing. There can be multiple agreements with same
        effective date.")."""
    refs = contract.get("parent_references") or []
    sup  = contract.get("supersedes_text")

    # Build (text, source) pairs so we know which quote triggered the match
    search_pairs = [(t, "parent_reference") for t in refs]
    if sup:
        search_pairs.append((sup, "supersedes_clause"))

    client = contract["client"]
    matches = []  # list of (type_rank, other, text, source)

    # ── Self-date filter (extraction-quality guard, May 14 2026) ──
    # The LLM frequently extracts THIS contract's own opening line into the
    # parent_references list — e.g. an amendment dated March 28 2016 yields
    # the quote "AMENDMENT ('Amendment') dated as of March 28, 2016 …" inside
    # its own parent_references. Date-matching against that quote then runs
    # against every other contract in the dataset and produces a spurious
    # link whenever any other doc happens to share the child's own effective
    # or signed date. To prevent that, any candidate date in the quote that
    # matches THIS contract's own effective_date or signed_date (within the
    # standard 7-day tolerance) is discarded before matching — those dates
    # are the contract describing itself, not a parent reference.
    my_self_dates = []
    for fld in ("effective_date", "signed_date"):
        d = parse_date_str(contract.get(fld))
        if d:
            my_self_dates.append(d)

    for text, source in search_pairs:
        candidate_dates = extract_dates_from_text(text)
        # Filter out self-dates — see comment above.
        candidate_dates = [
            cd for cd in candidate_dates
            if not any(dates_close(cd, my_d) for my_d in my_self_dates)
        ]
        if not candidate_dates:
            continue
        for other in by_client.get(client, []):
            if other["cache_key"] == contract["cache_key"]:
                continue
            other_date = parse_date_str(other.get("effective_date"))
            if other_date is None:
                continue
            for cd in candidate_dates:
                if dates_close(other_date, cd):
                    rank = _PARENT_TYPE_PREFERENCE.get(
                        normalize_type(other.get("contract_type", "")), 99
                    )
                    matches.append((rank, other, text, source))

    if not matches:
        return None, None, None, False

    # Pick best match: lowest rank = highest in hierarchy (MSA preferred over Amendment)
    matches.sort(key=lambda x: x[0])
    top_rank = matches[0][0]
    top_distinct = {m[1]["cache_key"] for m in matches if m[0] == top_rank}

    # Tie-break handling (May 20 2026 — refined per validator's Kearny Bank call):
    # When 2+ DISTINCT candidates tie at the top type-rank for the cited date,
    # the situation can be either:
    #   (a) the candidates are all mutual duplicates of one underlying
    #       agreement — pick any one (the canonical) and let Pass 3 mark
    #       the rest as duplicates of it; OR
    #   (b) the candidates split across more than one underlying agreement
    #       (e.g. on Kearny Bank, four "Fiserv + Kearny Federal" MSAs and
    #       one "Fiserv + Clifton Savings Bank" MSA all dated the same day,
    #       because two acquired entities each had their own MSA signed on
    #       that day). In that case we partition the tied candidates by
    #       exact party-set and check whether the CONTRACT'S OWN parties
    #       uniquely select one partition. If yes, pick the canonical of
    #       that partition. If no (no matching partition, or the contract
    #       has no parties extracted and multiple partitions exist), refuse
    #       — silently picking would mis-attribute the parent across an
    #       acquired-entity boundary.
    if len(top_distinct) > 1:
        # Build the unique top-rank candidate list (a single candidate may
        # appear multiple times in `matches` if more than one body quote
        # matched its date).
        seen_keys = set()
        unique_top = []
        for m in matches:
            if m[0] != top_rank:
                continue
            other = m[1]
            if other["cache_key"] in seen_keys:
                continue
            seen_keys.add(other["cache_key"])
            unique_top.append((other, m[2], m[3]))   # (contract, text, source)

        # Partition the candidates by exact party-set (lowercase + strip).
        # Candidates with no parties extracted go into a no-parties bucket.
        # Partition by NORMALISED party set (corporate-suffix stripped,
        # punctuation collapsed) so that "Fiserv Solutions, Inc." and
        # "Fiserv Solutions, LLC" land in the same partition. Without the
        # normalisation the Clifton-2014-06-06 MSA cluster fragments by
        # corporate suffix and the child amendment (which writes "LLC"
        # where the MSAs write "Inc.") finds no matching partition.
        party_partitions: dict = {}      # frozenset → list[(contract, text, source)]
        no_party_bucket: list  = []
        for tup in unique_top:
            cand = tup[0]
            cand_parties = _norm_party_set(cand.get("parties"))
            if cand_parties:
                party_partitions.setdefault(cand_parties, []).append(tup)
            else:
                no_party_bucket.append(tup)

        contract_parties = _norm_party_set(contract.get("parties"))

        chosen_partition = None
        chosen_partition_label = None

        if contract_parties and contract_parties in party_partitions:
            # The contract's own parties uniquely identify which candidate
            # partition to pick from — this handles the acquired-entity case.
            chosen_partition = party_partitions[contract_parties]
            chosen_partition_label = (
                f"party-set match against contract's own parties"
            )
        elif len(party_partitions) == 1 and not no_party_bucket:
            # Only one party-set among all tied candidates — they're all
            # mutual duplicates regardless of the contract's parties.
            chosen_partition = next(iter(party_partitions.values()))
            chosen_partition_label = (
                f"all {len(unique_top)} candidates share one party-set "
                f"(single duplicate cluster)"
            )
        elif not party_partitions and len(no_party_bucket) == len(unique_top):
            # Every candidate has empty parties — fall back to the no-parties
            # bucket as a single cluster (matches Pass 3's no-parties fallback).
            chosen_partition = no_party_bucket
            chosen_partition_label = (
                f"all {len(unique_top)} candidates have no extracted parties "
                f"(single fallback cluster)"
            )

        if chosen_partition is not None:
            canonical_idx = min(
                range(len(chosen_partition)),
                key=lambda i: chosen_partition[i][0]["filename"],
            )
            canonical, c_text, c_source = chosen_partition[canonical_idx]
            conf = "high" if contract.get("extraction_confidence") == "high" else "medium-high"
            quote = (c_text[:120] + "…") if c_text and len(c_text) > 120 else (c_text or "")
            other_names = sorted(
                c["filename"] for c, _t, _s in chosen_partition
                if c["cache_key"] != canonical["cache_key"]
            )
            others_preview = (", ".join(other_names[:2])
                + (f" (+{len(other_names) - 2} more)" if len(other_names) > 2 else "")) \
                if other_names else "—"
            other_partition_count = sum(
                1 for p, members in party_partitions.items()
                if members is not chosen_partition
            ) + (1 if no_party_bucket and no_party_bucket is not chosen_partition else 0)
            rationale = (
                f'Matched via {c_source.replace("_", " ") if c_source else "duplicate-cluster"}: '
                f'"{quote}" — {len(unique_top)} candidate parents tied at the cited date; '
                f'resolved by {chosen_partition_label}. Picked canonical '
                f'"{canonical["filename"]}" (alphabetically earliest in its party-set '
                f'cluster of {len(chosen_partition)}).'
                + (f" Other duplicates in this cluster: {others_preview}." if other_names else "")
                + (f" {other_partition_count} other distinct candidate party-set(s) "
                   f"present but ignored — they belong to different agreements."
                   if other_partition_count > 0 else "")
            )
            return canonical["cache_key"], conf, rationale, False

        # Genuinely ambiguous tie — refuse, fall through to Orphan.
        names = sorted({c["filename"] for c, _t, _s in unique_top})
        names_preview = ", ".join(names[:3]) + (f" (+{len(names) - 3} more)" if len(names) > 3 else "")
        ambiguity_note = (
            f"Tie-break refused — {len(unique_top)} candidate parents tied at "
            f"the top type-rank for the cited date: {names_preview}. "
            f"Candidates split across {len(party_partitions)} distinct party-set(s)"
            + (f" plus a no-parties bucket" if no_party_bucket else "")
            + (f" and the contract's own parties do not uniquely match any one of them."
               if contract_parties else
               f" and the contract has no parties extracted to disambiguate.")
            + " Marked Orphan per validator-confirmed rule."
        )
        return None, None, ambiguity_note, True

    # Single top-ranked candidate — safe to pick.
    _, best_other, best_text, best_source = matches[0]
    conf = "high" if contract.get("extraction_confidence") == "high" else "medium-high"
    quote = (best_text[:120] + "…") if len(best_text) > 120 else best_text
    rationale = f'Matched via {best_source.replace("_", " ")}: "{quote}"'
    return best_other["cache_key"], conf, rationale, False


def _are_mutual_duplicates(candidates):
    """Return True if every candidate in the list is a duplicate of every
    other candidate, by the same criteria Pass 3 uses to cluster duplicates:
      • all share the same contract_type,
      • all share the same effective_date (non-empty), and
      • when both sides have a non-empty parties list, their party sets are
        EXACTLY equal (after lowercase + strip). When either side has no
        parties extracted, the party check is skipped (matching Pass 3's
        fall-back behaviour).

    Used by find_parent_by_references to recognise the validator-confirmed
    case where a child cites a date that resolves to multiple Master
    Agreements that are themselves duplicate copies of each other (e.g. one
    MSA uploaded once per Fiserv department). In that case the script should
    pick the canonical rather than refuse, because Pass 3 will collapse the
    other copies as duplicates of the same canonical anyway.

    The exact-equality requirement on parties is deliberately strict — the
    earlier 50%-Jaccard rule clustered multi-entity MSAs (e.g. Kearny Bank's
    'Fiserv + Kearny Federal' and 'Fiserv + Clifton' agreements at the same
    date) together because they shared only the Fiserv vendor party. Real
    multi-department-upload duplicates have identical parties by
    construction, so exact equality keeps them while excluding cross-entity
    matches."""
    if len(candidates) < 2:
        return False
    first = candidates[0]
    first_type = first.get("contract_type")
    first_eff  = first.get("effective_date")
    if not first_type or not first_eff:
        return False
    first_parties = {
        p.strip().lower()
        for p in (first.get("parties") or [])
        if isinstance(p, str) and p.strip()
    }
    for other in candidates[1:]:
        if other.get("contract_type") != first_type:
            return False
        if other.get("effective_date") != first_eff:
            return False
        other_parties = {
            p.strip().lower()
            for p in (other.get("parties") or [])
            if isinstance(p, str) and p.strip()
        }
        if first_parties and other_parties:
            if first_parties != other_parties:
                return False
    return True


def find_parent_by_doc_codes(contract, by_client):
    """Priority 2: match by shared internal document codes.
    Returns (parent_key, confidence, rationale, is_ambiguous).

    Validator-confirmed signal (May 2026 call):
        "Sometimes in the agreements, they mention the agreement number or
         contract number. […] A service agreement contract number 2035 […]
         and the subsequent child, which are coming, […] do mention that
         contract number 2035."

    The strengthened SYSTEM_PROMPT now instructs the LLM to extract these
    agreement / contract numbers — printed, stamped, or handwritten — into
    `internal_doc_codes`.

    ── Per-client boilerplate filter (May 12 2026) ──
    Some codes the LLM extracts are NOT parent-link signals at all: they're
    client-wide identifiers (e.g. the client account number "290" stamped
    or printed on every Peoples Bank document) or per-document boilerplate
    ("DocuSign Envelope ID: …", "Tracking # 28336", page numbers, etc.).
    Naively matching on these builds spurious chains and false children —
    the original symptom was a 21-deep chain of POs/Amendments all linked
    to each other via "290" because that code appears on ~99 of the ~298
    Peoples Bank documents in this corpus.
    Fix: before matching, compute the per-client frequency of each code
    and drop codes that appear in MORE than 25% of the client's documents
    (or in more than 8 documents, whichever is larger). Anything more
    frequent than that is almost certainly identity / boilerplate, not a
    chain-specific parent-link signal.
    Note: this does sacrifice the validator's "handwritten Contract # 290"
    cases — the LLM is currently over-extracting "290" everywhere (because
    it appears as a printed client identifier on most pages), so we cannot
    distinguish the validator's intended-handwritten "290" link from the
    printed-everywhere "290" without per-position context the LLM doesn't
    capture. The 19 yellow validator cases will not match via this rule;
    they require a more discriminating extraction pipeline.

    ── Root-preference sort (May 12 2026) ──
    Among the candidates that survive the boilerplate filter, prefer
    candidates that are ALREADY at hierarchy_level 0 (i.e. root-eligible
    Standalones / MSAs resolved in Pass 1) over chained Pass-2-resolved
    Amendments. This collapses the original 21-deep chain into a flat
    "all children of the same root" structure for the remaining matched
    codes, which is what the validator's spreadsheet actually represents
    (children pointing at the same Software Agreement / MSA root, not
    chained through each other)."""
    # Normalise to {value, position} entries. For matching downstream we
    # work with two distinct sets:
    #   - my_high_value_values: codes the LLM saw in handwritten / stamp /
    #     label-box / printed-body positions. These are the validator-
    #     flagged parent-link signals and are IMMUNE to the boilerplate
    #     filter — even if the same value appears in many docs (e.g. the
    #     "290" agreement-number signal on the 2006 Software Agreement
    #     chain), the handwritten-margin / label-box position tells us
    #     it really IS being used as a cross-document link, not as a
    #     client-wide header identifier.
    #   - my_low_value_values: codes seen only in printed header / footer
    #     / unknown positions. These get the boilerplate filter applied
    #     so client-wide identifiers and DocuSign envelope IDs don't
    #     build spurious chains.
    my_entries = _normalize_doc_code_entries(contract.get("internal_doc_codes"))
    if not my_entries:
        return None, None, None, False

    my_high_value_values = {e["value"] for e in my_entries
                             if e["position"] in _HIGH_VALUE_CODE_POSITIONS}
    my_low_value_values  = {e["value"] for e in my_entries
                             if e["position"] in _LOW_VALUE_CODE_POSITIONS}

    client      = contract["client"]
    my_date     = parse_date_str(contract.get("effective_date"))
    client_docs = by_client.get(client, [])
    n_client    = len(client_docs)

    # Build a per-document map of {value → position} so we can look up the
    # position of a candidate's matching code (for the rationale). Also
    # compute per-client value-frequency to drive the boilerplate filter
    # on low-value-position codes only.
    other_entries_by_key = {}
    other_tokens_by_key  = {}      # NEW: high-value numeric tokens per candidate
    value_freq           = {}
    for c in client_docs:
        ents = _normalize_doc_code_entries(c.get("internal_doc_codes"))
        other_entries_by_key[c["cache_key"]] = {e["value"]: e["position"] for e in ents}
        other_tokens_by_key[c["cache_key"]]  = _extract_high_value_numeric_tokens(ents)
        for e in {e["value"] for e in ents}:
            value_freq[e] = value_freq.get(e, 0) + 1

    boilerplate_threshold = max(8, int(round(n_client * 0.25)))
    # Apply the boilerplate filter to low-value-position codes only.
    my_low_value_useful = {
        v for v in my_low_value_values
        if value_freq.get(v, 0) < boilerplate_threshold
    }
    # Effective set of codes we will match on: every high-value-position
    # code (unfiltered) plus the low-value-position codes that survived
    # the boilerplate threshold.
    my_match_values = my_high_value_values | my_low_value_useful

    # ── NEW (May 22 2026) — numeric-token matching for high-value codes ──
    # The exact-string match misses cases like
    #   parent code:  'Contract 2035'                               (printed_body)
    #   child code:   'Data Processing Services Agreement #2035'    (printed_body)
    # because the surrounding text differs even though the agreement
    # number '2035' is the actual linking signal. We extract 3+ digit
    # numeric runs from HIGH-VALUE-POSITION codes ONLY (handwritten,
    # stamp, label-box, body) and use them as an additional match channel.
    # Header / footer / unknown positions are excluded — that's where
    # boilerplate page numbers and client-wide identifiers live.
    my_high_value_tokens = _extract_high_value_numeric_tokens(my_entries)

    # ── NEW (May 22 2026) — cross-entity guard ────────────────────────────
    # Even after the boilerplate filter, doc-code matches can occasionally
    # cross acquired-entity boundaries inside a single client folder
    # (e.g. Kearny Bank has Clifton / Atlas / Millington / Central Jersey
    # / Kearny Federal sub-entities). Before accepting a candidate, check
    # that its normalised party set intersects with the child's — same
    # vendor + different acquired client = no real parent link. The check
    # is skipped (existing behaviour preserved) when either side has no
    # parties extracted.
    my_parties_norm = _norm_party_set(contract.get("parties"))

    if not my_match_values and not my_high_value_tokens:
        return None, None, None, False

    candidates = []
    for other in client_docs:
        if other["cache_key"] == contract["cache_key"]:
            continue
        other_value_positions = other_entries_by_key.get(other["cache_key"], {})
        other_tokens          = other_tokens_by_key.get(other["cache_key"], set())
        if not other_value_positions and not other_tokens:
            continue
        # Channel 1 — exact-string value match (existing behaviour)
        shared_values = my_match_values & set(other_value_positions.keys())
        # Channel 2 — high-value numeric-token match (new)
        shared_tokens = (my_high_value_tokens & other_tokens) if my_high_value_tokens else set()
        if not shared_values and not shared_tokens:
            continue
        # Cross-entity guard: require party overlap when both sides have parties.
        if my_parties_norm:
            other_parties_norm = _norm_party_set(other.get("parties"))
            if other_parties_norm and not (my_parties_norm & other_parties_norm):
                continue  # different acquired entities — skip silently
        other_date = parse_date_str(other.get("effective_date"))
        if other_date and my_date and other_date < my_date:
            # Build a richer label that records WHERE each matched code was
            # seen on the candidate parent side. Useful trace for the
            # rationale and for debugging.
            label_parts = []
            for v in sorted(shared_values):
                pos = other_value_positions.get(v, "unknown")
                label_parts.append(f"{v} [parent: {pos}]")
            for t in sorted(shared_tokens):
                label_parts.append(f"agreement # {t} [matched via numeric token in high-value position]")
            candidates.append((other, other_date, shared_values | shared_tokens, label_parts))

    if not candidates:
        return None, None, None, False

    # Root-preference sort:
    #   (1) lower hierarchy_level first (level=0 roots win over Pass-2 chained Amendments)
    #   (2) lower _PARENT_TYPE_PREFERENCE rank (MSA > Amendment > SOW > …)
    #   (3) most recent older as final tiebreaker
    def _sort_key(item):
        other, other_date, _shared, _label = item
        lvl = other.get("hierarchy_level")
        # None means Pass 2 hasn't resolved it yet — treat as level 99 so it
        # always loses to a level-0 root.
        lvl_key = lvl if isinstance(lvl, int) and lvl >= 0 else 99
        type_rank = _PARENT_TYPE_PREFERENCE.get(
            normalize_type(other.get("contract_type", "")), 99
        )
        return (lvl_key, type_rank, -other_date.toordinal())

    candidates.sort(key=_sort_key)
    other, _, _shared_values, label_parts = candidates[0]
    # Confidence: when ANY of the matched values landed on the child side
    # in a high-value position, bump to medium-high. This signals to the
    # downstream score bucketing that the link is supported by the
    # validator-flagged signal rather than just template overlap.
    matched_values = {p.split(" ", 1)[0] for p in label_parts}  # extract values from "VALUE [parent: …]"
    has_high_value_match = bool(matched_values & my_high_value_values)
    conf = "medium-high" if has_high_value_match else "medium"
    rationale = (
        "Matched via shared document code(s): " + ", ".join(label_parts)
        + (" — handwritten / stamped / label-box position is the validator's "
           "high-value parent-link signal." if has_high_value_match else "")
    )
    return other["cache_key"], conf, rationale, False


# NOTE (May 14 2026): `find_parent_by_active_msa` was REMOVED here. It
# previously inferred a parent MSA for Amendment-class children whose body
# cited "the Master Agreement" generically without a resolvable date, by
# linking to the most-recent older MSA in the same client (with a ±365-day
# refusal window to mitigate ambiguity). That logic was a soft form of
# date proximity — the validator's verbal rule on the April 30 call was
# explicit that date proximity cannot be used because multiple MSAs are
# typically in effect simultaneously. The function was added in service of
# matching the validator's updated spreadsheet, which itself was later
# confirmed to be unreliable; with the spreadsheet no longer the target,
# there is no remaining justification for the rule. Children whose body
# names a parent generically without a resolvable date are now correctly
# left as Orphans, awaiting better extraction of a specific date / agreement
# number cite from the body.


def find_parent_by_products(contract, by_client):
    """Priority 3 (validator-flagged gray area, May 2026): match by product
    overlap in section headers.

    Context for why this exists (verbatim from the validator):
        "if you have the product and services matched with that again, but
        that's a gray area […]. One limitation is that at the time of Jan
        1990, there were certain products for which Fiserv got into a
        contract with the client. But in say 2020, they added new more
        services / products for the same agreement. So those products may
        not be there in the master agreements altogether."

    So product overlap CAN link a child to its parent, but the signal is
    soft and old MSAs may not enumerate later-added products. We therefore
    apply it conservatively:

      • Only root-eligible level-0 candidates (Hierarchy_Status='Parent')
        are eligible parents — no chaining through children.
      • The candidate must be strictly older than the contract.
      • Overlap must be ≥1 distinct product AND ≥50% of the smaller
        product set (avoids spurious matches on a single generic term).
      • If two or more candidates tie at the top overlap, we refuse to
        pick (validator-confirmed: gray-area, do not silently choose).

    Returns (parent_key, confidence, rationale, is_ambiguous)."""
    my_products = {
        p.strip().lower()
        for p in (contract.get("products_in_headers") or [])
        if isinstance(p, str) and p.strip()
    }
    if not my_products:
        return None, None, None, False

    client  = contract["client"]
    my_date = parse_date_str(contract.get("effective_date"))
    if my_date is None:
        return None, None, None, False

    candidates = []
    for other in by_client.get(client, []):
        if other["cache_key"] == contract["cache_key"]:
            continue
        # Only root-eligible (level 0) docs may parent via product overlap.
        if other.get("hierarchy_level") != 0:
            continue
        other_products = {
            p.strip().lower()
            for p in (other.get("products_in_headers") or [])
            if isinstance(p, str) and p.strip()
        }
        if not other_products:
            continue
        shared = my_products & other_products
        if not shared:
            continue
        min_size      = min(len(my_products), len(other_products))
        overlap_ratio = len(shared) / min_size if min_size else 0
        if overlap_ratio < 0.5:
            continue
        other_date = parse_date_str(other.get("effective_date"))
        if other_date is None or other_date >= my_date:
            continue
        candidates.append((other, shared, overlap_ratio))

    if not candidates:
        return None, None, None, False

    # Sort by overlap ratio (descending), then by shared-count (descending),
    # then by older date first (older more likely to be the originating MSA).
    candidates.sort(key=lambda x: (-x[2], -len(x[1])))

    top_ratio = candidates[0][2]
    top_count = len(candidates[0][1])
    tied = [c for c in candidates if c[2] == top_ratio and len(c[1]) == top_count]

    if len(tied) > 1:
        names = sorted({c[0]["filename"] for c in tied})
        names_preview = ", ".join(names[:3]) + (f" (+{len(names) - 3} more)" if len(names) > 3 else "")
        ambig_note = (
            f"Product-overlap tie-break refused — {len(tied)} root-eligible "
            f"candidates share the same product set with this contract: "
            f"{names_preview}. Validator-confirmed gray-area rule: do not "
            f"silently pick when more than one candidate ties."
        )
        return None, None, ambig_note, True

    other, shared_products, _ratio = candidates[0]
    shared_preview = ", ".join(sorted(shared_products)[:5])
    rationale = (
        f"Gray-area product-overlap match: shares {len(shared_products)} "
        f"product(s) ({shared_preview}) in section headers with "
        f"{other['filename']}, which is an older root-eligible document. "
        f"Validator-flagged secondary signal — used only after parent_references "
        f"and shared doc-codes both fail."
    )
    return other["cache_key"], "low", rationale, False


def find_parent_by_date(contract, by_client):
    """Priority 3 (weakest): nearest older contract that is higher in the type hierarchy.
    MSAs look for nothing (handled as roots). Amendments look for MSAs.
    Sub-Amendments look for Amendments first, then MSAs.
    Returns (parent_key, confidence, rationale, is_ambiguous).
    is_ambiguous is True when more than one candidate at the top rank has an
    effective date within 90 days of the chosen parent — the date-proximity
    tie-breaker picked one silently."""
    my_date = parse_date_str(contract.get("effective_date"))
    if my_date is None:
        return None, None, None, False

    my_type = normalize_type(contract.get("contract_type", ""))
    my_rank = _TYPE_HIERARCHY_RANK.get(my_type, 99)
    if my_rank == 0:
        # MSAs are roots — date proximity doesn't apply
        return None, None, None, False

    client     = contract["client"]
    candidates = []
    for other in by_client.get(client, []):
        if other["cache_key"] == contract["cache_key"]:
            continue
        other_type = normalize_type(other.get("contract_type", ""))
        other_rank = _TYPE_HIERARCHY_RANK.get(other_type, 99)
        # Only consider contracts that are strictly higher in the hierarchy
        if other_rank >= my_rank:
            continue
        other_date = parse_date_str(other.get("effective_date"))
        if other_date and other_date < my_date:
            candidates.append((other, other_date, other_rank))

    if not candidates:
        return None, None, None, False

    # Sort: highest other_rank (closest parent type) first, then most recent date (latest ordinal)
    candidates.sort(key=lambda x: (-x[2], -x[1].toordinal()))
    other, other_date, top_rank = candidates[0]

    # Ambiguity signal: another candidate at the same rank within 90 days.
    competing = [
        cand for cand in candidates[1:]
        if cand[2] == top_rank and abs((cand[1] - other_date).days) <= 90
    ]
    is_ambiguous = len(competing) > 0

    rationale = (f"Inferred by date proximity — nearest older {other['contract_type']} is "
                 f"{other['filename']} ({other_date}). No explicit textual signal found.")
    if is_ambiguous:
        rationale += f" (ambiguous: {len(competing) + 1} candidates within 90 days at same rank)"
    return other["cache_key"], "low", rationale, is_ambiguous


# Map every variant the LLM might return to our canonical type names.
# Expanded with types seen in Boulder Dam CU, Abri CU, and common FinServ contract language.
_TYPE_MAP = {
    # MSA variants
    "msa":                               "MSA",
    "master agreement":                  "MSA",
    "master service agreement":          "MSA",
    "master services agreement":         "MSA",
    "relationship agreement":            "MSA",   # Boulder Dam CU pattern
    "network membership agreement":      "MSA",   # Accel/network style
    "accel network membership agreement":"MSA",
    "membership agreement":              "MSA",
    "amended and restated":              "MSA",   # full restatements are new roots
    "amended and restated agreement":    "MSA",
    # Amendment variants
    "amendment":                         "Amendment",
    "addendum":                          "Amendment",
    "first amendment":                   "Amendment",
    "second amendment":                  "Amendment",
    "third amendment":                   "Amendment",
    "fourth amendment":                  "Amendment",
    "fifth amendment":                   "Amendment",
    "sixth amendment":                   "Amendment",
    "seventh amendment":                 "Amendment",
    "eighth amendment":                  "Amendment",
    "ninth amendment":                   "Amendment",
    "tenth amendment":                   "Amendment",
    "amendment and restatement":         "Amendment",
    "rider":                             "Amendment",
    "consent":                           "Amendment",
    "waiver":                            "Amendment",
    "extension":                         "Amendment",
    "renewal":                           "Amendment",
    "change order":                      "Amendment",
    # Sub-Amendment variants
    "sub-amendment":                     "Sub-Amendment",
    "sub amendment":                     "Sub-Amendment",
    # SOW — strictly reserved for documents whose filename contains
    # "Statement of Work". A post-rule in resolve_hierarchy demotes any
    # SOW-tagged contract that doesn't meet this condition.
    "sow":                               "SOW",
    "statement of work":                 "SOW",
    "consulting services exhibit":       "SOW",
    # Former SOW siblings — now tracked as distinct types so they no longer
    # collapse into "SOW" in tooltips / Excel output.
    "order form":                        "Order Form",
    "work order":                        "Work Order",
    "purchase order":                    "Purchase Order",
    "po":                                "Purchase Order",
    "subsequent order":                  "Subsequent Order",
    # Document-level agreements seen in this portfolio that aren't MSAs.
    # Hardware Agreements are root-eligible per validator confirmation (May 2026):
    # "Most often hardware agreements are root agreements — they are separate agreements."
    "services agreement":                "Services Agreement",
    "software agreement":                "Software Agreement",
    "hardware agreement":                "Hardware Agreement",
    "maintenance agreement":             "Maintenance Agreement",
    "marketing agreement":               "Marketing Agreement",
    "referral agreement":                "Referral Agreement",
    "purchase agreement":                "Purchase Agreement",
    "network agreement":                 "Network Agreement",
    "network agreements":                "Network Agreement",
    # Additional root-eligible subtypes observed in Bay Bank corpus (May 2026).
    # The LLM commonly returns generic 'Standalone' for these; we recover the
    # specific subtype from the filename keyword pass. They land at level 0
    # alongside the other root-eligible non-MSA agreements.
    "internet banking agreement":        "Internet Banking Agreement",
    "access agreement":                  "Access Agreement",
    "escrow agreement":                  "Escrow Agreement",
    "subscription agreement":            "Subscription Agreement",
    # Termination / notice documents — treated as Amendment-like in hierarchy.
    "termination notice":                "Termination Notice",
    # Schedule / Exhibit / Appendix / Attachment — all subordinate doc types
    "schedule":                          "Schedule",
    "exhibit":                           "Schedule",
    "appendix":                          "Appendix",
    "attachment":                        "Attachment",
    # Miscellaneous documents (renamed from 'Standalone' on May 26 2026 —
    # the user-visible label is now 'Miscellaneous' in both the Excel
    # Contract_Type column and the HTML Parent-type filter).
    "standalone":                        "Miscellaneous",
    "stand-alone":                       "Miscellaneous",
    "miscellaneous":                     "Miscellaneous",
    "letter agreement":                  "Miscellaneous",
    "side letter":                       "Miscellaneous",
    # License Agreement — root-eligible (validator-confirmed)
    "license agreement":                 "License Agreement",
    "software license agreement":        "License Agreement",
}

# ── Validator-confirmed taxonomy (Peoples Bank, May 2026) ───────────────────
# These two sets are the foundation of the hierarchy classifier:
#
# _ROOT_ELIGIBLE_TYPES — types that may legitimately stand alone at the root
# of a contract chain. When the document HAS a parent reference / shared doc
# code, it can still become a child. When NONE resolves, it's tagged Standalone
# (status), at hierarchy_level == 0. EXCEPTION: MSAs are unconditionally roots
# regardless of any parent reference (validator's renewal-as-root rule).
#
# _CHILD_ONLY_TYPES — types that always have a parent. They can never become
# Standalone; if the parent can't be located in the dataset they end up as
# Orphan (status), at hierarchy_level == -1.
#
# Document types not in either set are treated as ambiguous and routed
# through the child-only path (safer default — Orphan rather than Standalone).
_ROOT_ELIGIBLE_TYPES = {
    "MSA",
    "Services Agreement",
    "Software Agreement",
    "License Agreement",
    "Network Agreement",
    "Membership Agreement",
    # Hardware / Maintenance / Marketing / Referral / Purchase agreements are
    # root-eligible per validator confirmation (May 2026): they are independent
    # agreements between Fiserv and the client, not subordinate documents under
    # another contract. When the LLM identifies one of these from the document
    # body AND no parent reference resolves, it lands as Standalone (level 0).
    "Hardware Agreement",
    "Maintenance Agreement",
    "Marketing Agreement",
    "Referral Agreement",
    "Purchase Agreement",
    # Additional root-eligible subtypes confirmed for Bay Bank (May 2026):
    # these stand alone at the root of a chain rather than amending another
    # contract. Validated by filename + body inspection of the corpus.
    "Internet Banking Agreement",
    "Access Agreement",
    "Escrow Agreement",
    "Subscription Agreement",
    # 'Miscellaneous' is the catch-all root-eligible label for documents
    # whose body is a self-contained agreement but doesn't match any of the
    # specific subtype keywords (Software / Hardware / Services / etc.).
    # Renamed from 'Standalone' on May 26 2026 for output clarity.
    "Miscellaneous",
}

_CHILD_ONLY_TYPES = {
    "Amendment",          # covers Addendum, Rider, Consent, Waiver, Extension, Renewal, Change Order
    "Sub-Amendment",
    "Appendix",
    "Attachment",
    "SOW",
    "Schedule",           # also covers Exhibit
    "Purchase Order",
    "Order Form",
    "Work Order",
    "Subsequent Order",
    "Termination Notice",
}

# Used by find_parent_by_references to prefer MSA over Amendment when multiple date matches exist
_PARENT_TYPE_PREFERENCE = {
    "MSA":                 0,
    "Amendment":           1,
    "Termination Notice":  1,   # Termination modifies an existing contract
    "Sub-Amendment":       2,
    "SOW":                 3,
    "Schedule":            4,
    "Appendix":            4,
    "Attachment":          4,
    "Services Agreement":  4,
    "Software Agreement":  4,
    "License Agreement":   4,
    "Network Agreement":   4,
    # Hardware / Maintenance / Marketing / Referral / Purchase agreements are
    # at the same rank as the other root-eligible non-MSA agreements — they can
    # be cited as parents of subordinate documents (Schedule / Exhibit / Order Form / etc.).
    "Hardware Agreement":  4,
    "Maintenance Agreement": 4,
    "Marketing Agreement": 4,
    "Referral Agreement":  4,
    "Purchase Agreement":  4,
    # Same rank as the other root-eligible non-MSA agreements.
    "Internet Banking Agreement": 4,
    "Access Agreement":    4,
    "Escrow Agreement":    4,
    "Subscription Agreement": 4,
    # 'Miscellaneous' (renamed from 'Standalone' May 26 2026) is the
    # catch-all root-eligible label — ranked lowest among parents because
    # it lacks a specific subtype keyword.
    "Miscellaneous":       5,
    "Purchase Order":      6,
    "Order Form":          6,
    "Work Order":          6,
    "Subsequent Order":    6,
}

# Used by find_parent_by_date: rank in hierarchy (lower = higher up the tree).
# NOTE: as of the validator-confirmed rules, find_parent_by_date is no longer
# called — this table is retained only so the function still type-checks if
# someone re-enables it. The hierarchy classifier itself uses parent_references
# and shared doc codes only.
_TYPE_HIERARCHY_RANK = {
    "MSA":                 0,
    "Amendment":           1,
    "Termination Notice":  1,
    "Sub-Amendment":       2,
    "SOW":                 3,
    "Schedule":            3,
    "Appendix":            3,
    "Attachment":          3,
    "Miscellaneous":       3,
    "Purchase Order":      3,
    "Order Form":          3,
    "Work Order":          3,
    "Subsequent Order":    3,
    "Services Agreement":  3,
    "Software Agreement":  3,
    "License Agreement":   3,
    "Network Agreement":   3,
    "Hardware Agreement":  3,
    "Maintenance Agreement": 3,
    "Marketing Agreement": 3,
    "Referral Agreement":  3,
    "Purchase Agreement":  3,
    "Internet Banking Agreement": 3,
    "Access Agreement":    3,
    "Escrow Agreement":    3,
    "Subscription Agreement": 3,
}

def normalize_type(raw):
    if not raw:
        return "Unknown"
    return _TYPE_MAP.get(raw.lower().strip(), raw)


# Priority-ordered filename keywords → canonical contract type. Used to
# cross-validate the LLM's contract_type against a second, structured signal:
# the filename itself. Filenames in this portfolio reliably encode the
# document type as a token (e.g. PEOPLES BANK THE_Amendment_04-18-2016…pdf),
# so when they disagree with the LLM we treat the filename as ground truth.
# Order matters — more specific multi-word patterns are checked before short
# single-word patterns to avoid false matches ("Master Services Agreement"
# should win over the shorter "Services Agreement" check even if it existed).
#
# Deliberately OMITTED: bare "services agreement" — too ambiguous (could be
# a master services agreement OR a schedule/exhibit to one). When the
# filename is ambiguous, we leave the LLM's type in place rather than guessing.
_FILENAME_TYPE_KEYWORDS = [
    (re.compile(r"\bsub[-\s]?amendment\b",              re.I), "Sub-Amendment"),
    (re.compile(r"\bmaster\s+services?\s+agreement\b",  re.I), "MSA"),
    (re.compile(r"\bmaster\s+agreement\b",              re.I), "MSA"),
    (re.compile(r"\brelationship\s+agreement\b",        re.I), "MSA"),
    (re.compile(r"\bnetwork\s+membership\s+agreement\b",re.I), "MSA"),
    (re.compile(r"\bmembership\s+agreement\b",          re.I), "MSA"),
    (re.compile(r"\bamended\s+and\s+restated\b",        re.I), "MSA"),
    (re.compile(r"\btermination\s+notice\b",             re.I), "Termination Notice"),
    (re.compile(r"\bamendment\b",                        re.I), "Amendment"),
    (re.compile(r"\baddendum\b",                         re.I), "Amendment"),
    (re.compile(r"\brider\b",                            re.I), "Amendment"),
    (re.compile(r"\bwaiver\b",                           re.I), "Amendment"),
    (re.compile(r"\bconsent\b",                          re.I), "Amendment"),
    (re.compile(r"\bextension\b",                        re.I), "Amendment"),
    (re.compile(r"\brenewal\b",                          re.I), "Amendment"),
    (re.compile(r"\bchange\s+order\b",                   re.I), "Amendment"),
    # Purchase-order family MUST be checked before generic order/work patterns
    # so "Purchase Order" / "Subsequent Order" are classified distinctly.
    (re.compile(r"\bpurchase\s+order\b",                 re.I), "Purchase Order"),
    (re.compile(r"\bsubsequent\s+order\b",               re.I), "Subsequent Order"),
    # SOW is reserved for filenames containing "Statement of Work".
    (re.compile(r"\bstatement\s+of\s+work\b",            re.I), "SOW"),
    # Former SOW siblings — now their own types.
    (re.compile(r"\border\s+form\b",                     re.I), "Order Form"),
    (re.compile(r"\bwork\s+order\b",                     re.I), "Work Order"),
    (re.compile(r"\bconsulting\s+services\s+exhibit\b",  re.I), "SOW"),
    # Document-level agreements seen in the portfolio that aren't MSAs.
    # Order matters: `services agreement` must come AFTER the master-services
    # pattern above so "Master Services Agreement" wins the MSA slot.
    (re.compile(r"\bservices\s+agreement\b",             re.I), "Services Agreement"),
    (re.compile(r"\bsoftware\s+agreement\b",             re.I), "Software Agreement"),
    (re.compile(r"\bsoftware\s+license\s+agreement\b",   re.I), "License Agreement"),
    (re.compile(r"\blicense\s+agreement\b",              re.I), "License Agreement"),
    (re.compile(r"\bnetwork\s+agreements?\b",            re.I), "Network Agreement"),
    # Root-eligible agreement types confirmed by validator (May 2026).
    (re.compile(r"\bhardware\s+agreement\b",             re.I), "Hardware Agreement"),
    (re.compile(r"\bmaintenance\s+agreement\b",          re.I), "Maintenance Agreement"),
    (re.compile(r"\bmarketing\s+agreement\b",            re.I), "Marketing Agreement"),
    (re.compile(r"\breferral\s+agreement\b",             re.I), "Referral Agreement"),
    (re.compile(r"\bpurchase\s+agreement\b",             re.I), "Purchase Agreement"),
    # Additional root-eligible subtypes observed in the Bay Bank corpus.
    # 'Internet Banking Agreement' is two words before 'agreement' so it must
    # appear before any single-word fallback that might fire on 'agreement'.
    (re.compile(r"\binternet\s+banking\s+agreement\b",   re.I), "Internet Banking Agreement"),
    (re.compile(r"\baccess\s+agreement\b",               re.I), "Access Agreement"),
    (re.compile(r"\bescrow\s+agreement\b",               re.I), "Escrow Agreement"),
    (re.compile(r"\bsubscription\s+agreement\b",         re.I), "Subscription Agreement"),
    # Child-only structural types — always recognised so they can never slip
    # past the child-only guard in resolve_hierarchy.
    (re.compile(r"\bappendix\b",                         re.I), "Appendix"),
    (re.compile(r"\battachment\b",                       re.I), "Attachment"),
    (re.compile(r"\bexhibit\b",                          re.I), "Schedule"),
    (re.compile(r"\bschedule\b",                         re.I), "Schedule"),
    (re.compile(r"\bletter\s+agreement\b",               re.I), "Miscellaneous"),
    (re.compile(r"\bside\s+letter\b",                    re.I), "Miscellaneous"),
]


def infer_type_from_filename(filename):
    """Return a canonical contract type inferred from filename keywords, or
    None when no clear keyword is present. Used only as a cross-check against
    the LLM's extracted type — never the sole source of classification.

    Filenames in this portfolio use underscores as token separators
    (e.g. 'PEOPLES BANK THE_Addendum_09-04-2012_843037.pdf'). Python's
    regex `\\b` does NOT fire on an underscore-to-letter boundary because
    `_` is a word character, so we first replace underscores / dots with
    spaces to make token boundaries regex-visible."""
    if not filename:
        return None
    normalized = re.sub(r'[_.]', ' ', filename)
    for pattern, type_name in _FILENAME_TYPE_KEYWORDS:
        if pattern.search(normalized):
            return type_name
    return None


# ── Product-name canonicaliser ───────────────────────────────────────────
# Loads the Fiserv product dictionary once (cached) and provides a single
# helper to map an LLM-extracted product name to its canonical form or to
# None when no confident match exists. The dictionary is intentionally
# discovered via glob — drop a new versioned file alongside the script and
# the canonicaliser picks it up on the next run with no code change.

_product_cache = None  # set on first call to _load_product_dictionary()


def _normalize_product_name(s):
    """Aggressive normalisation used as the lookup key for the dictionary.
    Lower-cases, collapses whitespace, strips surrounding non-alphanumeric
    punctuation. Internal punctuation is preserved so that distinctive
    tokens like 'Visionplus - FICO/FITS' don't collapse to the same key as
    'Visionplus FICO FITS' — keeps the keyspace honest."""
    if not isinstance(s, str):
        return ""
    n = s.strip().lower()
    n = re.sub(r"\s+", " ", n)
    n = n.strip(".,;:|/-—_ ")
    return n


def _pick_latest_dictionary(script_dir):
    """Glob the script directory for files matching PRODUCT_DICTIONARY_GLOB
    and return the most-recent version. Version is parsed as the integer
    immediately following 'v' in the filename stem (e.g. '…_v3.xlsx' beats
    '…_v2.xlsx'). Returns the Path of the winning file, or None when no
    file matches the glob."""
    matches = list(Path(script_dir).glob(PRODUCT_DICTIONARY_GLOB))
    if not matches:
        return None
    def _version_key(p):
        m = re.search(r"_v(\d+)", p.stem, re.IGNORECASE)
        return int(m.group(1)) if m else 0
    return max(matches, key=_version_key)


def _load_product_dictionary():
    """Read the product dictionary from the latest versioned file in the
    script directory and build the canonical set + normalised lookup. Cached
    on the module so it loads exactly once per process. Returns a 3-tuple:
      (canonical_set, normalized_to_canonical, normalized_keys)
    or None when the canonicaliser is disabled, no file matches the glob,
    or the file is unreadable."""
    global _product_cache
    if _product_cache is not None:
        return _product_cache if _product_cache != "MISS" else None
    if not PRODUCT_CANONICALIZER_ENABLED:
        _product_cache = "MISS"
        return None

    dict_path = _pick_latest_dictionary(_SCRIPT_DIR)
    if dict_path is None:
        print(f"  [product-canonicaliser] No file matched '{PRODUCT_DICTIONARY_GLOB}' "
              f"in {_SCRIPT_DIR}. Canonicalisation disabled for this run.")
        _product_cache = "MISS"
        return None

    try:
        df = pd.read_excel(str(dict_path), sheet_name=0)
    except Exception as e:
        print(f"  [product-canonicaliser] Could not read {dict_path.name}: {e}. "
              f"Canonicalisation disabled for this run.")
        _product_cache = "MISS"
        return None

    # Find the column holding the terms. Prefer an exact 'Term' header, else
    # use the first non-empty column.
    if "Term" in df.columns:
        terms_col = df["Term"]
    else:
        terms_col = df.iloc[:, 0]

    canonical_set = set()
    normalized_to_canonical = {}
    for raw in terms_col.dropna().astype(str):
        canon = raw.strip()
        if not canon:
            continue
        canonical_set.add(canon)
        norm = _normalize_product_name(canon)
        if norm and norm not in normalized_to_canonical:
            normalized_to_canonical[norm] = canon

    # Sort by length descending so the head-prefix match stage in
    # canonicalize_product iterates LONGER (more specific) canonicals first
    # — e.g. when both "SecureNow" and "SecureNow Data Feed" are in the
    # dictionary, an input like "SecureNow Data Feed Schedule" picks the
    # specific "SecureNow Data Feed" rather than collapsing to the shorter
    # "SecureNow". Secondary sort key is the canonical string itself so the
    # ordering is deterministic across runs.
    normalized_keys = sorted(normalized_to_canonical.keys(), key=lambda k: (-len(k), k))
    print(f"  [product-canonicaliser] Loaded {len(canonical_set)} canonical "
          f"product names from {dict_path.name}.")
    _product_cache = (canonical_set, normalized_to_canonical, normalized_keys)
    return _product_cache


# Generic suffix words that contract authors typically append to product
# names in section headers without changing the underlying product identity
# ("SecureNow Services", "Bill Pay Schedule", "Configure Digital Module"…).
# Used by canonicalize_product to recover the canonical name when a literal
# match against the dictionary fails on the trailing word(s).
_GENERIC_PRODUCT_SUFFIXES = (
    "services", "service",
    "schedule",  "schedules",
    "exhibit",   "exhibits",
    "module",    "modules",
    "addendum",  "addenda",
)


def canonicalize_product(raw):
    """Map a raw LLM-extracted product name to a canonical dictionary entry,
    or return None when no confident match exists.

    Four-stage match (each stage only runs when the previous one fails):

      1. EXACT normalised hit — lowercase + whitespace-collapsed + trim,
         then O(1) dictionary lookup.

      2. HEAD-PREFIX match with word boundary — if the input starts with
         a canonical entry (≥3 chars, followed by a space or end-of-string),
         return that canonical. Sorted by canonical length descending so
         longer / more specific entries win first. Catches the common
         pattern where the LLM appends generic suffix text to a real
         product name — e.g. "SecureNow Services Schedule to the ASP
         Services Exhibit" → "SecureNow".

      3. SUFFIX STRIP and retry — iteratively strip a generic Fiserv-corpus
         trailing word (Services, Service, Schedule, Exhibit, Module,
         Addendum) and re-try the exact-match stage. Catches cases where
         prefix match couldn't fire because the suffix sits between the
         product name and the canonical (rare, but cheap to handle).

      4. FUZZY fallback — difflib.get_close_matches at PRODUCT_FUZZY_THRESHOLD.
         Catches typos and minor wording variations ("ach mananger" →
         "ACH Manager").

    Non-string / empty / non-matching inputs return None — callers should
    drop those entries from the canonical output."""
    if not isinstance(raw, str) or not raw.strip():
        return None
    loaded = _load_product_dictionary()
    if loaded is None:
        # Canonicaliser disabled or dictionary unavailable — pass through
        # the raw value so the pipeline behaves identically to pre-feature.
        return raw.strip() or None
    _canonical_set, normalized_to_canonical, normalized_keys = loaded

    norm = _normalize_product_name(raw)
    if not norm:
        return None

    # Stage 1 — exact normalised hit
    if norm in normalized_to_canonical:
        return normalized_to_canonical[norm]

    # Stage 2 — head-prefix match with word boundary.
    # normalized_keys is sorted longest-first by _load_product_dictionary so
    # the most specific canonical wins (e.g. "SecureNow Data Feed" beats
    # "SecureNow" when both fit). The 3-character floor avoids matching on
    # tiny canonical entries like single-letter abbreviations.
    for key in normalized_keys:
        if len(key) < 3:
            continue
        if norm.startswith(key + " "):
            return normalized_to_canonical[key]

    # Stage 3 — strip a generic trailing word and retry exact match. Repeat
    # while the trailing word remains generic; stop as soon as the trimmed
    # form matches the dictionary or no more suffixes remain.
    stripped = norm
    while True:
        # Find the last word
        parts = stripped.rsplit(" ", 1)
        if len(parts) != 2:
            break
        head, last = parts
        if last not in _GENERIC_PRODUCT_SUFFIXES:
            break
        stripped = head.rstrip(".,;:|/-—_ ").strip()
        if not stripped:
            break
        if stripped in normalized_to_canonical:
            return normalized_to_canonical[stripped]

    # Stage 4 — fuzzy fallback (difflib, stdlib)
    import difflib
    matches = difflib.get_close_matches(norm, normalized_keys, n=1, cutoff=PRODUCT_FUZZY_THRESHOLD)
    if matches:
        return normalized_to_canonical[matches[0]]
    return None


def resolve_hierarchy(contracts_flat):
    """Build parent-child hierarchy. Returns enriched contract list.
    Handles forest structure — one client can have multiple independent MSA trees."""

    # Enrich each contract with LLM-extracted fields at the top level
    for c in contracts_flat:
        meta = c.get("metadata") or {}

        # Contract-type reconciliation: LLM (document body) wins.
        #
        # Validator-confirmed rule (May 2026): the document body — what the
        # LLM is reading — is the source of truth for contract_type. Filenames
        # in this portfolio were curated by people uploading scanned documents
        # over the past 30 years, and they very frequently picked the wrong
        # type at index time. Concrete examples the validator walked us through:
        #
        #   • "PEOPLES BANK THE_Services Agreement_…_964012.tiff" → opening
        #     the PDF reveals "Appendix A1 to the v9 software license agreement",
        #     not a services agreement. The validator's verdict: Child / Appendix.
        #   • "PEOPLES BANK THE_Amendment_04-18-2016_1494127.pdf" → the body
        #     is actually a Master Agreement. Validator's verdict: Standalone MSA.
        #   • 13 hardware-agreement files where the body turned out to be an
        #     Amendment or Schedule. Validator's verdict: those rows belong on
        #     a parent chain, not as Hardware-Agreement standalones.
        #
        # We therefore trust the LLM's contract_type whenever the LLM was able
        # to identify a type at all (i.e. not "Unknown"). The filename keyword
        # is retained for two narrow uses only:
        #   (a) bootstrapping when the LLM returned "Unknown" — better to use
        #       the filename hint than to leave the document untyped;
        #   (b) audit logging when the two disagree so reviewers can spot-check
        #       LLM mis-classifications without re-running extraction.
        _llm_type = normalize_type(meta.get("contract_type", "Unknown"))
        _fn_type  = infer_type_from_filename(c.get("filename", ""))
        _type_override_note = ""

        # The LLM's enumerated contract_type vocabulary is narrow (MSA /
        # Amendment / Sub-Amendment / SOW / Schedule / Standalone), so for any
        # root-eligible non-MSA document — Software, Hardware, Services,
        # Maintenance, Marketing, Referral, Purchase, Network agreements — the
        # LLM almost always returns the generic 'Standalone' (which our
        # _TYPE_MAP renames to 'Miscellaneous' for user-facing output, as
        # of May 26 2026). The filename keyword is the only place we can
        # recover the specific subtype, and refining a generic
        # 'Miscellaneous' to e.g. 'Software Agreement' does NOT contradict
        # the validator rule: both classifications agree the document is
        # root-eligible, the filename is purely adding metadata.
        # Both label variants are recognised here so the refinement still
        # fires whether normalize_type has been re-applied or not.
        _GENERIC_LLM_TYPES = {"Standalone", "Miscellaneous", "Unknown"}
        # The set of filename-derived types that are strictly more specific
        # than 'Miscellaneous' (and still root-eligible). We only refine when
        # the filename keyword falls inside this set.
        _SPECIFIC_ROOT_ELIGIBLE = {
            "Services Agreement", "Software Agreement", "License Agreement",
            "Hardware Agreement", "Maintenance Agreement", "Marketing Agreement",
            "Referral Agreement", "Purchase Agreement", "Network Agreement",
            "Membership Agreement", "MSA",
            # Additional root-eligible subtypes observed in Bay Bank corpus.
            "Internet Banking Agreement", "Access Agreement",
            "Escrow Agreement", "Subscription Agreement",
        }

        if _llm_type == "Unknown" and _fn_type:
            # LLM couldn't classify — fall back to filename keyword.
            c["contract_type"] = _fn_type
            _type_override_note = (
                f"LLM returned 'Unknown'; type upgraded to '{_fn_type}' from "
                f"filename keyword as a last-resort fallback."
            )
        # NOTE (May 13 2026): a previous "MSA filename override" branch lived
        # here. It forced contract_type = "MSA" whenever the filename keyword
        # named a Master-Agreement variant, regardless of what the LLM
        # extracted from the body. That branch was REMOVED on May 13 2026
        # because the validator never explicitly authorised an exception to
        # the body-wins rule for MSA filenames; in fact the April 30 call
        # gave the opposite example (a file with "Amendment" in the filename
        # but a Master Agreement in the body — the validator's verdict was
        # MSA, body wins, exactly the opposite direction). The override was
        # an inference from the spreadsheet's behaviour rather than a stated
        # rule, and it directly contradicted what the validator said on the
        # record. With this branch gone, every filename / body disagreement
        # — MSA-flavoured or otherwise — resolves to the LLM's body verdict
        # with an audit-only note recorded in `type_override_note`.
        elif _llm_type in _GENERIC_LLM_TYPES and _fn_type in _SPECIFIC_ROOT_ELIGIBLE:
            # LLM gave a generic root-eligible answer; filename gives a more
            # specific (still root-eligible) subtype. Refining doesn't change
            # the hierarchy verdict (both → Miscellaneous landing) but preserves
            # the subtype label so reviewers can still distinguish Software /
            # Hardware / Services / etc. agreements in the output.
            c["contract_type"] = _fn_type
            _type_override_note = (
                f"LLM returned generic '{_llm_type}'; type refined to "
                f"'{_fn_type}' from filename keyword. Hierarchy verdict "
                f"unchanged (both are root-eligible)."
            )
        else:
            # LLM identified a specific type — that wins, even if filename disagrees.
            # This is the validator-confirmed rule (May 2026): when the LLM
            # reads the document body and identifies a specific type (Amendment,
            # MSA, Schedule, Sub-Amendment, SOW), trust the body. The filename
            # is only used for audit logging in this branch.
            c["contract_type"] = _llm_type
            if _fn_type and _fn_type != _llm_type:
                # Audit trail only — do NOT override.
                _type_override_note = (
                    f"Note: filename keyword suggests '{_fn_type}' but the "
                    f"document body identifies as '{_llm_type}'. Body wins "
                    f"(validator-confirmed rule, May 2026)."
                )

        # The strict-SOW filename guard that previously demoted any LLM-claimed
        # SOW lacking a "Statement of Work" filename keyword has been removed
        # for the same reason: the validator was explicit that file-name-driven
        # demotion is wrong. If the LLM reads the body and says it's a SOW, we
        # honour that. The few false-positive SOWs the LLM occasionally returned
        # will surface in review with the audit note above attached.
        c["type_override_note"]          = _type_override_note

        c["amendment_number"]            = meta.get("amendment_number")

        # ── Date extraction (two distinct dates now) ──────────────────
        # The LLM is asked to extract two dates:
        #   - signed_date     → when parties signed (USED FOR DISPLAY: tooltip,
        #                       Excel "Signed_Date" column, chart x-position)
        #   - effective_date  → when the document is effective (USED FOR
        #                       HIERARCHY RESOLUTION: parent references inside
        #                       contracts cite effective dates, not signing
        #                       dates, so we keep effective_date for matching)
        # Filename-encoded date is treated as the effective date for this
        # portfolio's naming convention (MM-DD-YYYY in the filename = effective
        # date). It serves as a fallback when the LLM didn't find an
        # effective_date, and as a sanity check that overrides the LLM's
        # effective_date if they disagree by >90 days (catches OCR mis-reads
        # where the LLM picked up a referenced parent's date).
        _llm_signed_date    = meta.get("signed_date")
        _llm_effective_date = meta.get("effective_date")
        _fn_date            = c.get("filename_date")

        # Effective date — drives hierarchy resolution
        _effective = _llm_effective_date or _fn_date
        _date_override_note = ""
        if _llm_effective_date and _fn_date:
            try:
                _ld = datetime.strptime(_llm_effective_date, "%Y-%m-%d").date()
                _fd = datetime.strptime(_fn_date,           "%Y-%m-%d").date()
                if abs((_ld - _fd).days) > 90:
                    _effective = _fn_date
                    _date_override_note = (
                        f"Filename date ({_fn_date}) preferred over LLM-extracted "
                        f"effective date ({_llm_effective_date}) — differ by "
                        f"{abs((_ld - _fd).days)} days."
                    )
            except (ValueError, TypeError):
                pass
        c["effective_date"]              = _effective
        c["date_override_note"]          = _date_override_note

        # Signed date — used for display. Falls back to effective date when
        # the LLM didn't surface a signing date (some scanned signature pages
        # are illegible; some documents — Order Forms, emails — have no
        # signature at all).
        c["signed_date"]                 = _llm_signed_date or _effective

        c["parties"]                     = meta.get("parties") or []
        c["parent_references"]           = meta.get("parent_references") or []
        c["supersedes_text"]             = meta.get("supersedes_text")
        # internal_doc_codes — kept in normalised form ({value, position}
        # dicts) so every downstream consumer (find_parent_by_doc_codes,
        # rescue logic, Excel export) sees a consistent shape regardless of
        # whether the cache entry was written before or after the May 12 2026
        # position-aware extraction prompt landed.
        c["internal_doc_codes"]          = _normalize_doc_code_entries(meta.get("internal_doc_codes"))

        # ── Section structure (May 22 2026 — three-level nested form) ────
        # Normalise from whichever shape the cache stored (new nested
        # section_structure OR legacy flat section_headers/products) into
        # the nested form. Legacy entries synthesise empty sub-header /
        # item lists; new entries carry the LLM's three-level extraction.
        section_struct = _normalize_section_structure(meta)
        c["section_structure"] = section_struct

        # Walk every level — top-level header product, sub-header product,
        # AND leaf-item text — to build the canonical and raw product sets.
        # The deepest level (items) is where the most specific Fiserv
        # product names live in this corpus, per the validator's-team
        # confirmation, so we explicitly include items in the canonicaliser.
        raw_product_names = list(_walk_section_products(section_struct))
        canonical_names   = [canonicalize_product(p) for p in raw_product_names]

        # Build the legacy-mirror flat arrays from the top-level entries so
        # any downstream consumer still reading section_headers /
        # section_header_products continues to work. (find_parent_by_products
        # is one of these — but it operates on products_in_headers, which we
        # populate from all three levels below.)
        legacy_headers  = [entry["header"] for entry in section_struct]
        legacy_products = [
            canonicalize_product(entry.get("product")) if entry.get("product") else None
            for entry in section_struct
        ]
        legacy_products_raw = [entry.get("product") for entry in section_struct]

        c["section_headers"]              = legacy_headers
        c["section_header_products"]      = legacy_products
        c["section_header_products_raw"]  = legacy_products_raw

        # Deduped, sorted list of distinct CANONICAL product names from
        # ALL levels of the structure (top-level header, sub-header, and
        # leaf items). Drops entries that didn't match the dictionary so
        # downstream consumers (find_parent_by_products, Excel reporting)
        # see only aligned names.
        c["products_in_headers"]          = sorted({
            p.strip() for p in canonical_names if isinstance(p, str) and p.strip()
        })
        # Parallel audit list — every distinct RAW product name the LLM
        # surfaced from any level, before canonicalisation. Useful to spot
        # dictionary gaps (real product names the LLM found but that v1
        # of the dictionary doesn't yet include).
        c["products_in_headers_raw"]      = sorted({
            p.strip() for p in raw_product_names if isinstance(p, str) and p.strip()
        })

        # Cache the human-readable rendering once so the Excel export and
        # HTML hover tooltip don't each have to recompute it.
        c["section_structure_text"]       = _render_section_structure_text(section_struct)
        c["is_active"]                   = meta.get("is_active")
        c["extraction_confidence"]       = meta.get("extraction_confidence", "low")
        c["extraction_confidence_score"] = _get_numeric_extraction_score(meta)
        c["extraction_failed"]           = meta.get("extraction_failed", False)

    # Diagnostic: show what contract types the LLM returned
    type_counts = defaultdict(int)
    for c in contracts_flat:
        type_counts[c["contract_type"]] += 1
    print(f"  Contract types found: {dict(type_counts)}")

    # Group by client for scoped parent matching (prevents cross-client errors)
    by_client = defaultdict(list)
    for c in contracts_flat:
        by_client[c["client"]].append(c)

    # Lookup by cache_key
    by_key = {c["cache_key"]: c for c in contracts_flat}

    # Initialize hierarchy fields
    for c in contracts_flat:
        c["parent_key"]                  = None
        c["hierarchy_confidence"]        = None
        c["hierarchy_level"]             = None
        c["hierarchy_notes"]             = ""
        c["hierarchy_rationale"]         = ""
        c["hierarchy_method"]            = None     # 'root' | 'miscellaneous' | 'reference' | 'doc_code' | 'products' | 'duplicate' | 'orphan'
        c["hierarchy_confidence_score"] = None     # int 0-100 (derived after method is known)
        c["hierarchy_confidence_bucket"] = None     # 'high' | 'medium' | 'low' | None
        c["is_ambiguous"]                = False   # set True when the resolver had a silent tie-break
        c["duplicate_of"]                = None    # set in Pass 3 when this row is a duplicate of a canonical sibling

    # ── Pass 1: identify roots ────────────────────────────────────────────
    # Validator-confirmed rules (Peoples Bank, May 2026):
    #   • Every Master Agreement is its own root, UNCONDITIONALLY — even when
    #     its body text references a prior MSA. A 2021 "Amended and Restated
    #     Master Agreement" referencing a 2016 MSA is its own root, not a
    #     child of the 2016 one.
    #   • Other root-eligible types (Services Agreement, Software Agreement,
    #     License Agreement, Network Agreement, Membership Agreement,
    #     Standalone) are roots ONLY if they have no resolvable parent
    #     reference and no shared internal doc code. If a parent IS found,
    #     they become children of that parent.
    #   • Child-only types (Amendment, Sub-Amendment, Schedule, Appendix,
    #     Attachment, SOW, Purchase Order, Order Form, etc.) are NEVER roots.
    for c in contracts_flat:
        ct = c["contract_type"]

        # Child-only types are deferred to Pass 2 — they cannot be roots.
        if ct in _CHILD_ONLY_TYPES:
            continue

        if ct == "MSA":
            # Unconditional root — assumption 4.
            c["hierarchy_level"]      = 0
            c["hierarchy_confidence"] = "high"
            c["hierarchy_method"]     = "root"
            c["hierarchy_notes"]      = "Parent — each Master Agreement is its own root"
            c["hierarchy_rationale"]  = (
                "Parent (MSA) — every Master Agreement is treated as its own root, "
                "even when its body references a prior MSA (validator-confirmed renewal-as-root rule)."
            )
            continue

        if ct in _ROOT_ELIGIBLE_TYPES:
            # Validator's call rule (May 2026) — verbatim:
            #   "Software agreements will always be a root agreement,
            #    software license agreements will always be a root agreement.
            #    Oh, they are not going to be a child agreement."
            # …with one important caveat the validator added a moment later:
            #   "And there's another scenario. […] FISA is getting into a
            #    software agreement. It is not just one document there,
            #    they're like a bundle of documents […] supplemental to the
            #    main agreement on the same date […] all have been tracked as
            #    service agreements. In that scenario, the main documents
            #    becomes a root and other ones, they become the part of that
            #    service agreement."
            #
            # So the validator's nuance is:
            #   ─ A Software/Services/Hardware/etc. Agreement that introduces
            #     a new relationship → root (Standalone).
            #   ─ A Software/Services Agreement whose body explicitly cites
            #     and amends/supplements another document IN THE DATASET →
            #     child of that document (a supplementary bundle member or
            #     an appendix that was mis-indexed as a "Services Agreement").
            #   ─ A Software/Services Agreement whose body cites a parent
            #     that is NOT in the dataset → still Standalone, NOT Orphan.
            #     The validator's "Orphan = child with missing parent" rule
            #     applies to child-only types only. For root-eligible types,
            #     the "always a root" rule wins: even when the body name-drops
            #     a prior agreement we can't find, the doc is its own root
            #     (mirrors the renewal-as-root MSA rule).
            #
            # Crucially, shared internal_doc_codes alone are NOT enough to
            # demote a root-eligible type to a child. The validator's
            # spreadsheet was explicit: the 17 cases where the previous
            # script-via-doc_code linkage was wrong all had description
            # "Standalone MSA" — i.e. the validator's verdict was Standalone.
            # Doc-code overlap is a weak repository-template signal here and
            # tends to mis-link siblings as parents.
            parent_key, conf, rationale, ambig = find_parent_by_references(c, by_client)
            # Bundle-exception gate (May 14 2026):
            # The validator's verbal rule was that non-MSA root-eligible types
            # (Software / Services / Hardware / etc. Agreements) are ALWAYS
            # roots, with one exception: a "bundle of documents" signed on
            # the SAME DATE where one is the main agreement and the others
            # are supplementary documents attached to it. Earlier code
            # generalised this to "any reference match → Child", which
            # collapsed years-apart renewals into spurious child links. We
            # now require the resolved parent to be within ±90 days of the
            # child for the bundle exception to fire; otherwise the doc
            # stays a Standalone (renewal-as-root) regardless of what the
            # body recital cites.
            is_bundle_member = False
            if parent_key:
                child_date  = parse_date_str(c.get("effective_date"))
                parent_obj  = next(
                    (o for o in by_client.get(c["client"], []) if o["cache_key"] == parent_key),
                    None,
                )
                parent_date = parse_date_str(parent_obj.get("effective_date")) if parent_obj else None
                if (child_date is not None
                        and parent_date is not None
                        and abs((child_date - parent_date).days) <= 90):
                    is_bundle_member = True

            if is_bundle_member:
                # Body cites a prior contract within ±90 days — treat as a
                # same-date bundle child / mis-indexed appendix.
                c["parent_key"]           = parent_key
                c["hierarchy_confidence"] = conf
                c["hierarchy_method"]     = "reference"
                c["hierarchy_rationale"]  = (rationale or "")
                c["is_ambiguous"]         = bool(ambig)
                continue
            # No body-text parent reference, or the referenced parent is
            # outside the ±90-day bundle window → unconditional root
            # (validator's "Software agreements will always be a root" rule;
            # the date-band restored matches the validator's "same-date
            # bundle" caveat without the broader generalisation).
            c["hierarchy_level"]      = 0
            c["hierarchy_confidence"] = "high"
            c["hierarchy_method"]     = "miscellaneous"
            c["hierarchy_notes"]      = (
                f"Parent ({ct}) — non-MSA root-eligible type, no in-window parent reference"
            )
            # Build the rationale: distinguish "no parent reference at all"
            # from "parent reference exists but resolves outside the bundle
            # window" — both lead to a Miscellaneous landing but the second
            # case is worth surfacing in the audit trail.
            if parent_key:
                _resolved_parent_fn = parent_obj["filename"] if parent_obj else "(unknown)"
                _days_gap = abs((child_date - parent_date).days) if (child_date and parent_date) else None
                c["hierarchy_rationale"] = (
                    f"Parent ({ct}) — body cites a prior contract "
                    f"({_resolved_parent_fn}) but the cited document's effective "
                    f"date is {_days_gap if _days_gap is not None else 'far'} day(s) "
                    f"away from this one, outside the ±90-day bundle-exception "
                    f"window. Per the validator-confirmed renewal-as-root rule, "
                    f"a non-MSA root-eligible type only becomes a child of a "
                    f"prior contract when the two were signed as a same-date "
                    f"bundle; otherwise it is its own root."
                )
            else:
                c["hierarchy_rationale"] = (
                    f"Parent ({ct}) — validator-confirmed root rule: "
                    f"\"Software agreements will always be a root agreement.\" No verbatim "
                    f"parent_reference / supersedes quote in the body. Shared doc-codes "
                    f"are NOT used to demote root-eligible types — doc-code-only "
                    f"linkages on root-eligible types are usually false positives."
                )
            continue

        # Unknown / unclassified types — fall through to Pass 2 and treat as
        # child-only (safer default: Orphan rather than Standalone).

    # ── Pass 2: resolve parents for non-roots ─────────────────────────────
    # Method priority (each step only runs when the previous one fails):
    #   (a) parent_references / supersedes_text  → method 'reference'
    #   (b) shared internal_doc_codes            → method 'doc_code'
    #   (c) product-overlap in section headers   → method 'products'
    #       (GRAY AREA — used only for child-only types where the body clearly
    #       cites a parent but we can't identify which root-eligible candidate
    #       it refers to. See find_parent_by_products for the conservative
    #       thresholds.)
    #
    # Date proximity is intentionally NOT used: in this corpus multiple MSAs
    # of similar dates are typically simultaneously active, so nearest-date
    # is unsafe (validator's explicit verbal rule).
    #
    # An earlier "active-MSA inference" step was REMOVED on May 14 2026 — it
    # was a soft form of date proximity (most-recent older MSA with a refusal
    # window), added to recover the validator's "MSA available" spreadsheet
    # cases. The spreadsheet later turned out to be unreliable and the rule
    # directly contradicted the verbal "no date proximity" guidance, so it
    # came back out.
    for c in contracts_flat:
        if c["hierarchy_level"] is not None or c.get("parent_key") is not None:
            continue  # already resolved as root or as a root-eligible child above

        parent_key, conf, rationale, ambig = find_parent_by_references(c, by_client)
        method = "reference" if parent_key else None

        # When the reference-finder refused due to a tie-break, preserve the
        # rationale string so we can surface it on the eventual Orphan record.
        # `ambig` from the references finder is True for two reasons:
        #   (a) parent_key was returned but tied with another candidate, OR
        #   (b) parent_key is None because the finder refused to pick.
        # Case (b) is identifiable by parent_key being None while rationale
        # carries the "Tie-break refused…" message.
        ref_tie_note = rationale if (parent_key is None and ambig and rationale) else None

        if not parent_key:
            parent_key, conf, rationale, ambig = find_parent_by_doc_codes(c, by_client)
            method = "doc_code" if parent_key else None

        # Capture an analogous "tie-break refused" note from product-overlap,
        # so the eventual Orphan rationale can explain what blocked the match.
        prod_tie_note = None
        if not parent_key:
            parent_key, conf, rationale, ambig = find_parent_by_products(c, by_client)
            if parent_key is None and ambig and rationale:
                prod_tie_note = rationale
            method = "products" if parent_key else method

        # Date proximity intentionally omitted — see comment above.

        if parent_key:
            c["parent_key"]           = parent_key
            c["hierarchy_confidence"] = conf
            c["hierarchy_method"]     = method
            c["hierarchy_rationale"]  = rationale or ""
            c["is_ambiguous"]         = bool(ambig)
        else:
            c["hierarchy_level"]      = -1
            c["hierarchy_confidence"] = None
            c["hierarchy_method"]     = "orphan"
            if ref_tie_note:
                # Tie-break refusal in the textual finder — distinct from the
                # generic "no parent text" orphan because the document DID cite
                # a parent, we just couldn't disambiguate among multiple
                # equal-rank candidates.
                c["hierarchy_notes"]     = "Orphan — parent reference is ambiguous (tie-break refused)"
                c["hierarchy_rationale"] = ref_tie_note
                c["is_ambiguous"]        = True
            elif prod_tie_note:
                # Tie-break refusal in product-overlap — gray-area ambiguity:
                # multiple root-eligible parents share the same product set.
                c["hierarchy_notes"]     = "Orphan — gray-area product overlap is ambiguous"
                c["hierarchy_rationale"] = prod_tie_note
                c["is_ambiguous"]        = True
            else:
                c["hierarchy_notes"]     = "Orphan — no parent signal found"
                c["hierarchy_rationale"] = (
                    "Orphan — no explicit parent reference, no shared document code, "
                    "and no product-overlap match against any root-eligible candidate. "
                    "Date proximity is not used per the validator-confirmed rule."
                )

    # BFS level assignment — iterate until stable (handles multi-level chains)
    for _ in range(20):
        progress = False
        for c in contracts_flat:
            if c["hierarchy_level"] is not None or c["parent_key"] is None:
                continue
            parent = by_key.get(c["parent_key"])
            if parent and parent["hierarchy_level"] is not None and parent["hierarchy_level"] >= 0:
                c["hierarchy_level"] = parent["hierarchy_level"] + 1
                if not c["hierarchy_notes"]:
                    c["hierarchy_notes"] = f"Parent: {parent['filename']}"
                progress = True
        if not progress:
            break

    # Anything still unresolved = circular reference or missing parent → orphan.
    #
    # Clear the now-stale parent_key on chain-break Orphans. Without this,
    # the row carries a parent_key pointing at a doc that is itself an Orphan,
    # which the Excel and HTML layers happily display in the Parent_Contract
    # column / hover tooltip — producing the "this Orphan has a parent?"
    # inconsistency. The matched-but-broken-chain parent is still preserved
    # in the hierarchy_rationale for trace / debugging.
    for c in contracts_flat:
        if c["hierarchy_level"] is None:
            broken_parent = by_key.get(c.get("parent_key") or "")
            broken_parent_fn = broken_parent["filename"] if broken_parent else None
            broken_method    = c.get("hierarchy_method")

            c["hierarchy_level"]      = -1
            c["hierarchy_confidence"] = "low"
            c["hierarchy_method"]     = "orphan"
            c["parent_key"]           = None    # clear stale parent linkage
            c["hierarchy_notes"]      = "Orphan — could not resolve parent chain"

            chain_break_note = (
                f"Orphan — a candidate parent was matched"
                + (f" via '{broken_method}'" if broken_method else "")
                + (f" ({broken_parent_fn})" if broken_parent_fn else "")
                + ", but that parent is itself an Orphan, so the chain breaks. "
                  "Parent linkage cleared; this contract is treated as a leaf "
                  "Orphan rather than a sub-chain of broken references."
            )
            if c.get("hierarchy_rationale"):
                # Preserve whatever method-specific rationale was set during
                # Pass 2 (e.g. the matched doc-code text or reference quote)
                # so the trace survives, then append the chain-break note.
                c["hierarchy_rationale"] = c["hierarchy_rationale"] + " | " + chain_break_note
            else:
                c["hierarchy_rationale"] = chain_break_note

    # NOTE (May 14 2026): The Pass 2b "Orphan-to-Standalone rescue" was
    # REMOVED here. It used to promote any Orphan whose LLM-extracted type
    # was child-only but whose filename keyword named a specific
    # root-eligible subtype (Software / Hardware / Services / Maintenance /
    # Marketing / Referral / Purchase / License / Network / Membership
    # Agreement) up to Standalone, overriding the LLM's body verdict with
    # the filename. The rationale at the time was that the validator's
    # spreadsheet treated those rows as "Standalone MSA". With the
    # spreadsheet now confirmed unreliable, that justification is gone,
    # and the rule directly contradicted the validator's verbal body-wins
    # rule: when the LLM reads the body and says it's a child-only type,
    # we trust the body. A row that ends up Orphan after Pass 2 stays
    # Orphan; the filename hint is not used to override the body.

    # ── Pass 3: Duplicate detection (validator-confirmed) ──────────────────
    # Real-world cause (verbatim from validator):
    #   "Fiserv gets into an agreement with a client. Citibank under that
    #    agreement is contracting for three different services. Each
    #    department signs the document, gets a copy, and uploads it on our
    #    repository at different times. That is the reason why you have
    #    duplicates."
    #
    # Instruction (verbatim):
    #   "Connect all the child agreements to one of the documents [the
    #    canonical], and mark the other two as the duplicate of that."
    #
    # Identification rules used here:
    #   1. Cluster root-eligible level-0 documents within the same client by
    #      (normalised contract_type, effective_date). A duplicate cluster
    #      requires at least 2 documents in the same (type, date) bucket.
    #   2. Within a cluster, parties must overlap ≥ 50% (Jaccard-style on
    #      lowercased names) — different counterparties on the same date are
    #      not duplicates of each other, just unrelated agreements that happen
    #      to share a calendar date.
    #   3. Canonical is the alphabetically-earliest filename inside the
    #      cluster — deterministic across runs, no LLM signal needed.
    #
    # Effects on duplicates:
    #   • hierarchy_status stays at "Standalone" (level 0) so they still show
    #     in roots — they ARE root agreements, just duplicate copies of one.
    #   • A `duplicate_of` field is set on each non-canonical, pointing at
    #     the canonical's cache_key.
    #   • The Notes column gets a "Duplicate of <filename>" annotation.
    #   • hierarchy_method is rewritten to "duplicate" (score base 90) so the
    #     downstream confidence/bucket reflects the dedup certainty, not the
    #     original "miscellaneous (base 95)" landing.
    #
    # Effects on children of duplicates:
    #   • Any contract whose parent_key points at a non-canonical duplicate
    #     is re-linked to point at the canonical instead, with a note
    #     explaining the redirection. This matches the validator's example:
    #     all amendments of the cluster should converge on one canonical MSA.
    duplicate_clusters = defaultdict(list)
    for c in contracts_flat:
        if c.get("hierarchy_level") != 0:
            continue
        ct  = c.get("contract_type")
        eff = c.get("effective_date")
        client = c.get("client")
        # Only cluster when all three keys are present — a missing key would
        # over-cluster everything with the same gap into one false duplicate group.
        if not (ct and eff and client):
            continue
        duplicate_clusters[(client, ct, eff)].append(c)

    canonical_for_duplicate: dict[str, str] = {}   # dup_key → canonical_key

    for (client, ct, eff), members in duplicate_clusters.items():
        if len(members) < 2:
            continue

        # Sub-cluster by exact party-set (May 20 2026):
        # The initial (client, contract_type, effective_date) bucket can mix
        # documents from different acquired entities — e.g. on Kearny Bank
        # a Fiserv-vs-Kearny-Federal MSA and a Fiserv-vs-Clifton MSA both
        # dated 2013-06-19 land in the same coarse bucket even though they
        # are unrelated agreements. Previously the code picked the
        # alphabetically-first member as the canonical and then checked
        # each subsequent member's parties against THAT canonical only,
        # which meant that when the canonical happened to be from a
        # different party-set, none of the remaining members got matched
        # to each other either.
        #
        # Fix: partition each coarse bucket into sub-clusters by exact
        # party-set, then process each sub-cluster independently. Members
        # with no extracted parties land in their own no-parties bucket
        # (rather than being silently fused with the first party-set seen,
        # which would re-introduce the cross-entity false-positive cluster).
        party_subclusters: dict[frozenset, list] = {}
        no_party_members: list = []
        for m in members:
            m_parties = frozenset(
                p.strip().lower() for p in (m.get("parties") or [])
                if isinstance(p, str) and p.strip()
            )
            if m_parties:
                party_subclusters.setdefault(m_parties, []).append(m)
            else:
                no_party_members.append(m)

        # Each party-set sub-cluster is its own duplicate cluster.
        subclusters_to_process = list(party_subclusters.values())
        if len(no_party_members) >= 2:
            # When two or more members share the (client, type, eff) bucket
            # but have no extracted parties, they're treated as one
            # no-parties sub-cluster — falling back to the type+date+client
            # signal, matching the original behaviour for the empty-parties
            # case.
            subclusters_to_process.append(no_party_members)

        for sub_members in subclusters_to_process:
            if len(sub_members) < 2:
                continue
            sub_sorted = sorted(sub_members, key=lambda x: x["filename"])
            canonical  = sub_sorted[0]
            for other in sub_sorted[1:]:
                other["duplicate_of"]      = canonical["cache_key"]
                other["hierarchy_method"]  = "duplicate"
                other["hierarchy_notes"]   = f"Duplicate of {canonical['filename']}"
                other["hierarchy_rationale"] = (
                    f"Duplicate-of-canonical (Pass 3): same client, same contract_type "
                    f"('{ct}'), same effective_date ({eff}), and identical parties to "
                    f"{canonical['filename']}. Canonical retained at root; this copy is "
                    f"marked as a duplicate. Children of either copy converge on the "
                    f"canonical so the chain stays single-rooted."
                )
                canonical_for_duplicate[other["cache_key"]] = canonical["cache_key"]

    # Re-link children whose parent_key now points at a non-canonical duplicate.
    for c in contracts_flat:
        pk = c.get("parent_key")
        if not pk:
            continue
        new_pk = canonical_for_duplicate.get(pk)
        if not new_pk:
            continue
        old_parent = by_key.get(pk)
        new_parent = by_key.get(new_pk)
        c["parent_key"] = new_pk
        redirect_note = (
            f"Parent redirected to canonical {new_parent['filename']} "
            f"(original match {old_parent['filename']} marked as duplicate in Pass 3)."
        )
        c["hierarchy_rationale"] = (
            c.get("hierarchy_rationale", "") + " | " + redirect_note
            if c.get("hierarchy_rationale") else redirect_note
        )

    # ── Numeric confidence scoring pass ────────────────────────────────────
    # Composite: average of the method base score and the LLM's extraction score.
    # Orphans get 0; roots get the method base as-is (there's no "parent" to score).
    for c in contracts_flat:
        method = c.get("hierarchy_method") or ("orphan" if c["hierarchy_level"] == -1 else None)
        base   = _METHOD_BASE_SCORES.get(method, 0)
        ext    = c.get("extraction_confidence_score", 0)

        if method == "orphan":
            score = 0
        elif method == "root":
            # Root confidence is driven by how certain we are this really is an MSA
            # (extraction score), combined with the "no parent found" method base.
            score = int(round((base + ext) / 2))
        else:
            score = int(round((base + ext) / 2))

        c["hierarchy_confidence_score"]  = score
        c["hierarchy_confidence_bucket"] = _score_to_bucket(score) if method != "orphan" else None

    # ── Ambiguous-bucket rule pass ─────────────────────────────────────────
    # Beyond the silent tie-breaks already flagged by the finders, also mark as
    # ambiguous any contract whose hierarchy was settled by the weakest method
    # (date proximity) AND whose extraction score itself is weak — both signals
    # are soft, so the placement is defensible but not authoritative.
    for c in contracts_flat:
        if c.get("is_ambiguous"):
            continue
        if (c.get("hierarchy_method") == "date"
                and c.get("extraction_confidence_score", 0) < _BUCKET_MEDIUM_MIN):
            c["is_ambiguous"] = True
            note = ("Ambiguous — resolved by weakest method (date proximity) "
                    "AND extraction score is below the Medium threshold.")
            c["hierarchy_rationale"] = (c["hierarchy_rationale"] + " | " + note
                                         if c.get("hierarchy_rationale") else note)

    # ── Surface date-override notes in the rationale ───────────────────────
    # When the filename-date sanity check replaced the LLM-extracted date,
    # carry that explanation into the rationale so reviewers can see it in
    # hover tooltips and Excel rows.
    for c in contracts_flat:
        note = c.get("date_override_note")
        if note:
            c["hierarchy_rationale"] = (c["hierarchy_rationale"] + " | " + note
                                         if c.get("hierarchy_rationale") else note)

    # ── Surface type-override notes in the rationale ───────────────────────
    # Same pattern as the date-override: if the filename keyword cross-check
    # overrode the LLM's contract_type, record that in the rationale.
    for c in contracts_flat:
        note = c.get("type_override_note")
        if note:
            c["hierarchy_rationale"] = (c["hierarchy_rationale"] + " | " + note
                                         if c.get("hierarchy_rationale") else note)

    # ── Compute the validator's hierarchy_status field ─────────────────────
    # Maps hierarchy_level → one of the validator-aligned statuses:
    # 'Parent' | 'Child' | 'Orphan' | 'Duplicate'.
    #
    # Validator's-team confirmation (May 22 2026): every level-0 root-eligible
    # document — whether the body identifies it as an MSA, Master Services
    # Agreement, Software Agreement, Services Agreement, Hardware Agreement,
    # Maintenance Agreement, etc. — should carry a single unified label.
    # Sid was explicit: "instead of having this bifurcation, we should have
    # this one definition that it should be an MSA". The previous bifurcation
    # between 'Original MSA' (MSA contract_type) and 'Standalone' (non-MSA
    # root-eligible) was the bifurcation he wanted removed.
    #
    # Subsequent decision (May 26 2026): the unified label was renamed from
    # 'Original MSA' to 'Parent' in BOTH the Excel and HTML outputs so the
    # data field matches the user-facing display label. The bifurcation Sid
    # wanted gone is still gone; only the label string has been updated.
    #
    # The granular contract_type field (Software Agreement, Services
    # Agreement, etc.) is still recorded faithfully from the body, so no
    # information is lost — only the hierarchy_status label is unified.
    for c in contracts_flat:
        lvl = c["hierarchy_level"]
        if c.get("duplicate_of"):
            c["hierarchy_status"] = "Duplicate"
        elif lvl == -1:
            c["hierarchy_status"] = "Orphan"
        elif lvl == 0:
            c["hierarchy_status"] = "Parent"
        else:
            c["hierarchy_status"] = "Child"

    return contracts_flat


# ============================================================
# PHASE 4: HTML VISUALIZATION
# ============================================================

def _level_to_y(level, orphan_y=-9):
    """Map hierarchy level to Y-axis position. Orphans placed at the given
    orphan_y (default -9 for the old static layout; build_visualization now
    overrides this with a value computed from the deepest non-orphan level
    present, so the Orphan strip is always at least 3 units below the
    deepest 'Others' row regardless of how deep the chains go)."""
    if level == -1:
        return orphan_y
    return -level


# Sentinel placeholder used for contracts whose date is missing.
# The visualization layer detects this value and plots those contracts inside
# a dedicated "Unknown Date" strip on the left of the x-axis (see
# build_visualization). Kept as a module constant so extraction and rendering
# agree on what "unknown" looks like on the timeline.
UNKNOWN_DATE_PLACEHOLDER = datetime(1900, 1, 1)


def _signed_date_to_dt(contract):
    """Get the signed date (the display date) as Python datetime, or
    UNKNOWN_DATE_PLACEHOLDER if missing. Falls back to effective_date when no
    signed date was extracted (already enforced upstream in resolve_hierarchy,
    but defensive here too in case a contract dict was constructed elsewhere)."""
    d = contract.get("signed_date") or contract.get("effective_date")
    if not d:
        return UNKNOWN_DATE_PLACEHOLDER
    try:
        return datetime.strptime(d, "%Y-%m-%d")
    except (ValueError, TypeError):
        return UNKNOWN_DATE_PLACEHOLDER


def _parent_filename(contract, all_contracts):
    pk = contract.get("parent_key")
    if not pk:
        return "None"
    parent = next((c for c in all_contracts if c["cache_key"] == pk), None)
    return parent["filename"] if parent else pk


def build_visualization(contracts, output_path, client_name=None):
    """Build interactive Plotly scatter plot mirroring Ryan's POC.

    Extras added in this version:
      (1) Dashed era-boundary lines at 2011, 2016, 2021 on top of the shaded bands.
      (2) Curved parent→child connectors (spline via a midpoint control point),
          colored by the child's hierarchy confidence.
      (3) Two dropdowns ('Color by …' and 'Show …') wired to Plotly updatemenus
          that restyle marker.color / visible on the category traces.
      (4) Two-row legend: a 'Confidence' row (dummy traces, color swatches) and
          a 'Document Type' row (real category traces, symbol swatches).
    """

    # Colours are keyed by the *numeric-score bucket* now, not the old
    # categorical confidence strings. The three-bucket scheme mirrors the
    # target screenshot's legend (≥70% / 40–69% / <40%).
    COLOR_MAP = {
        "high":    "#27ae60",   # green  — score ≥ 70
        "medium":  "#f39c12",   # amber  — score 40–69
        "low":     "#e74c3c",   # red    — score < 40
        None:      "#95a5a6",   # grey   — unknown / orphan
    }

    SYMBOL_MAP = {
        # Single root bucket — unified per validator's-team confirmation
        # (May 22 2026) and renamed from 'Original MSA' to 'Parent' across
        # BOTH the Excel data field and the HTML visualisation on
        # May 26 2026. MSA and non-MSA root-eligible types both use the
        # filled-star marker.
        "Parent":       "star",
        "Amendment":    "circle",
        "Sub-Amendment":"diamond",
        "Others":       "square",
        "Orphan":       "x",
    }

    # Era definitions — shared by shading, dashed boundaries, and the
    # "Color by: Era" dropdown option.
    ERA_BOUNDARIES = [
        datetime(2011, 1, 1),
        datetime(2016, 1, 1),
        datetime(2021, 1, 1),
    ]
    ERA_BANDS = [
        (datetime(1985, 1, 1), datetime(2011, 1, 1), "Era 1: Pre-2011",  "rgba(173,216,230,0.18)"),
        (datetime(2011, 1, 1), datetime(2016, 1, 1), "Era 2: 2011–2016", "rgba(144,238,144,0.18)"),
        (datetime(2016, 1, 1), datetime(2021, 1, 1), "Era 3: 2016–2021", "rgba(255,200,150,0.18)"),
        (datetime(2021, 1, 1), datetime(2027, 1, 1), "Era 4: 2021+",     "rgba(230,200,230,0.18)"),
    ]
    ERA_COLOR_MAP = {
        "Era 1": "#6baed6",
        "Era 2": "#74c476",
        "Era 3": "#fd8d3c",
        "Era 4": "#9e9ac8",
    }

    TYPE_COLOR_MAP = {
        "MSA":           "#2c7bb6",
        "Amendment":     "#d7191c",
        "Sub-Amendment": "#fdae61",
        "SOW":           "#abd9e9",
        "Schedule":      "#ffffbf",
        # 'Miscellaneous' (renamed from 'Standalone' May 26 2026) — the
        # catch-all colour for root-eligible docs lacking a specific subtype.
        "Miscellaneous": "#999999",
        "Unknown":       "#cccccc",
    }

    def era_of(dt):
        if dt < ERA_BOUNDARIES[0]:
            return "Era 1"
        if dt < ERA_BOUNDARIES[1]:
            return "Era 2"
        if dt < ERA_BOUNDARIES[2]:
            return "Era 3"
        return "Era 4"

    # Detect whether any contract has a missing effective_date. If so, we
    # reserve a narrow "Unknown Date" strip just inside the left edge of the
    # visible range (1976–1979) and relocate unknown-date points into it, so
    # they remain visible when the x-axis starts at 1980.
    has_unknown_dates = any(
        _signed_date_to_dt(c) == UNKNOWN_DATE_PLACEHOLDER for c in contracts
    )
    UNKNOWN_STRIP_X0   = datetime(1976, 1, 1)
    UNKNOWN_STRIP_X1   = datetime(1979, 6, 1)
    UNKNOWN_STRIP_MID  = datetime(1977, 9, 1)   # where unknown points get plotted
    # When unknowns exist we widen the visible range to 1975 so the strip fits;
    # otherwise we keep the strict 1980 start the user requested.
    xaxis_range_start  = datetime(1975, 1, 1) if has_unknown_dates else datetime(1980, 1, 1)

    # ── Dynamic Y-axis layout ────────────────────────────────────────────
    # The static -9 anchor used by the old _level_to_y caused deep-chain rows
    # (Level 6, Level 7, …) plotted at y=-6 / -7 to crowd up against the
    # Orphan row at y=-9, which made it look like Orphans had connectors when
    # they were really Others overlapping.
    #
    # Fix: compute the deepest non-orphan level present, then anchor the
    # Orphan strip 3 units below that. We also enforce a minimum max-level of
    # 5 so that on small datasets (no deep chains) the chart still has the
    # familiar Level 1 / Level 2 / … / Level 5 / Orphan ladder.
    _nonorphan_levels = [c["hierarchy_level"] for c in contracts
                         if isinstance(c.get("hierarchy_level"), int) and c["hierarchy_level"] >= 0]
    MAX_LEVEL_SHOWN  = max(5, max(_nonorphan_levels) if _nonorphan_levels else 5)
    ORPHAN_Y         = -(MAX_LEVEL_SHOWN + 3)         # always ≥3-unit gap
    ORPHAN_DIVIDER_Y = ORPHAN_Y + 1.5                 # dashed line above strip

    def level_y(level):
        """Local y-mapper that respects the dynamic ORPHAN_Y."""
        return _level_to_y(level, orphan_y=ORPHAN_Y)

    def _plot_x(contract):
        """Return the x-axis datetime used for plotting this contract:
        its real effective_date, or the Unknown-strip midpoint if missing."""
        dt = _signed_date_to_dt(contract)
        return UNKNOWN_STRIP_MID if dt == UNKNOWN_DATE_PLACEHOLDER else dt

    fig = go.Figure()

    # "Unknown Date" strip — only rendered when at least one contract lacks a date.
    if has_unknown_dates:
        fig.add_vrect(
            x0=UNKNOWN_STRIP_X0, x1=UNKNOWN_STRIP_X1,
            fillcolor="rgba(220,220,220,0.55)", line_width=0,
            annotation_text="Unknown Date",
            annotation_position="top left",
            annotation_font_size=10,
            annotation_font_color="rgba(90,90,90,0.95)"
        )
        # Solid divider separating the Unknown strip from the real timeline
        fig.add_vline(
            x=UNKNOWN_STRIP_X1,
            line=dict(color="rgba(120,120,120,0.7)", width=1.2, dash="dot"),
        )

    # (1a) Era shading bands
    for x0, x1, label, color in ERA_BANDS:
        fig.add_vrect(
            x0=x0, x1=x1,
            fillcolor=color, line_width=0,
            annotation_text=label,
            annotation_position="top left",
            annotation_font_size=10,
            annotation_font_color="rgba(100,100,100,0.7)"
        )

    # (1b) Dashed era-boundary lines on top of the shading
    for boundary in ERA_BOUNDARIES:
        fig.add_vline(
            x=boundary,
            line=dict(color="rgba(200,80,80,0.55)", width=1.5, dash="dash"),
        )

    # (1c) Orphan-strip visual separation. A faint grey band behind the
    # Orphan row plus a dashed divider just above it make it unambiguous
    # which markers are Orphans even when deep-chain "Others" rows extend
    # toward the bottom of the chart. Previously a Level-6 or Level-7
    # marker plotted just one or two units above the Orphan row could be
    # mistaken for an Orphan with a connector.
    fig.add_hrect(
        y0=ORPHAN_Y - 0.7,
        y1=ORPHAN_Y + 0.7,
        fillcolor="rgba(248,206,204,0.18)",   # same hue as the Orphan fill in Excel
        line_width=0,
    )
    fig.add_hline(
        y=ORPHAN_DIVIDER_Y,
        line=dict(color="rgba(120,120,120,0.55)", width=1.2, dash="dash"),
    )

    # (2) Parent → child connector lines — curved spline via a midpoint control
    #     point, colored by the child's hierarchy confidence. We track each
    #     connector's child category / child confidence and the three per-mode
    #     colours so the 'Color by' and 'Show' dropdowns can restyle / hide
    #     connectors alongside the category markers.
    def _level_to_group(level, contract_type=None):
        # All level-0 docs map to a single 'Parent' group on the chart.
        # MSA, Master Services Agreement, Software / Hardware / Services /
        # Internet Banking / Access / Escrow / Subscription Agreements, etc.
        # are all treated as one root family for the visualisation. The
        # Excel Hierarchy_Status data field also reads 'Parent' (renamed
        # from 'Original MSA' on May 26 2026 so Excel matches HTML).
        # contract_type is kept as a parameter for callers that still pass
        # it but is otherwise unused.
        if level == 0:    return "Parent"
        if level == 1:    return "Amendment"
        if level == 2:    return "Sub-Amendment"
        if level > 2:     return "Others"
        return "Orphan"

    connector_trace_indices    = []
    connector_colors_by_conf   = []
    connector_colors_by_type   = []
    connector_colors_by_era    = []
    connector_child_groups     = []
    connector_child_confs      = []
    connector_child_cache_keys = []   # parallel — child contract's cache_key, used by Products filter

    for c in contracts:
        if not c["parent_key"] or c["hierarchy_level"] < 0:
            continue
        parent = next((x for x in contracts if x["cache_key"] == c["parent_key"]), None)
        if not parent:
            continue
        px, py = _plot_x(parent), level_y(parent["hierarchy_level"])
        cx, cy = _plot_x(c),       level_y(c["hierarchy_level"])
        # Midpoint pulled slightly below both endpoints produces a downward bow.
        mx = px + (cx - px) / 2
        my = min(py, cy) - 0.4

        # Connector colour follows the *numeric-score bucket* so connectors
        # stay visually consistent with the bucketed markers. We still track
        # the raw categorical confidence for the hover and filter logic.
        child_bucket  = c.get("hierarchy_confidence_bucket")
        child_conf    = child_bucket   # what the Filter-by-Confidence dropdown keys on
        child_type    = c.get("contract_type")
        child_real_dt = _signed_date_to_dt(c)
        child_group   = _level_to_group(c["hierarchy_level"], c.get("contract_type"))

        edge_color_conf = COLOR_MAP.get(child_bucket, "#95a5a6")
        edge_color_type = TYPE_COLOR_MAP.get(child_type, "#cccccc")
        edge_color_era  = ERA_COLOR_MAP.get(era_of(child_real_dt), "#cccccc")

        connector_trace_indices.append(len(fig.data))
        connector_colors_by_conf.append(edge_color_conf)
        connector_colors_by_type.append(edge_color_type)
        connector_colors_by_era.append(edge_color_era)
        connector_child_groups.append(child_group)
        connector_child_confs.append(child_conf)
        connector_child_cache_keys.append(c["cache_key"])

        fig.add_trace(go.Scatter(
            x=[px, mx, cx],
            y=[py, my, cy],
            mode="lines",
            line=dict(color=edge_color_conf, width=1.2, shape="spline", smoothing=1.3),
            opacity=0.45,
            showlegend=False,
            hoverinfo="skip",
        ))

    # Contract groups for legend. The `is_ambiguous` flag is still computed
    # and exported to Excel, but it no longer forms its own visual bucket —
    # ambiguous contracts fall under their hierarchy-level group.
    groups = [
        ("Parent",        [c for c in contracts if c["hierarchy_level"] == 0]),   # unified root bucket — Excel Hierarchy_Status also reads 'Parent' (renamed May 26 2026)
        ("Amendment",     [c for c in contracts if c["hierarchy_level"] == 1]),
        ("Sub-Amendment", [c for c in contracts if c["hierarchy_level"] == 2]),
        ("Others",        [c for c in contracts if c["hierarchy_level"]  > 2]),
        ("Orphan",        [c for c in contracts if c["hierarchy_level"] == -1]),
    ]

    # Category traces — precompute parallel color arrays for each "Color by …"
    # mode, AND per-point confidence labels so the new "Filter by Confidence"
    # dropdown can fade non-matching markers via marker.opacity.
    category_trace_indices   = []
    category_group_names     = []
    category_confidences     = []      # list of lists, one per category trace
    category_cache_keys      = []      # parallel to xs/ys — used by Products filter
    category_contract_types  = []      # list of lists — used by the Parent-type filter
    colors_by_confidence_all = []
    colors_by_type_all       = []
    colors_by_era_all        = []

    for group_name, group_contracts in groups:
        if not group_contracts:
            continue

        xs, ys, colors_conf, colors_type, colors_era, hover_texts = [], [], [], [], [], []
        trace_confidences = []
        for c in group_contracts:
            real_dt = _signed_date_to_dt(c)          # real date (or UNKNOWN placeholder) for era bucketing
            plot_dt = _plot_x(c)                        # x-coord used for plotting
            xs.append(plot_dt)
            ys.append(level_y(c["hierarchy_level"]))
            bucket = c.get("hierarchy_confidence_bucket")
            colors_conf.append(COLOR_MAP.get(bucket, "#95a5a6"))
            colors_type.append(TYPE_COLOR_MAP.get(c.get("contract_type"), "#cccccc"))
            # Era coloring uses the *real* date; unknowns fall into Era 1 by default,
            # but their hover text still shows "unknown" so the user can disambiguate.
            colors_era.append(ERA_COLOR_MAP.get(era_of(real_dt), "#cccccc"))
            # Filter-by-Confidence keys on the bucket so "High only" etc. behave correctly.
            trace_confidences.append(bucket)
            # Section structure — render the nested 3-level form in the
            # tooltip. Top-level headers as bullets; sub-headers indented
            # one step; leaf items packed onto the sub-header's line.
            # Trims long lists so the tooltip stays readable.
            _structure = c.get("section_structure") or []
            if _structure:
                _lines = []
                _top_shown = 0
                for entry in _structure:
                    if _top_shown >= 6:
                        _lines.append(f"… (+{len(_structure) - _top_shown} more sections)")
                        break
                    _top_shown += 1
                    h = entry.get("header") or ""
                    p = entry.get("product")
                    if isinstance(p, str) and p.strip():
                        _lines.append(f"<b>• {h}</b> <i>[{p.strip()}]</i>")
                    else:
                        _lines.append(f"<b>• {h}</b>")
                    subs = entry.get("subheaders") or []
                    _sub_shown = 0
                    for sub in subs:
                        if _sub_shown >= 4:
                            _lines.append(f"&nbsp;&nbsp;&nbsp;&nbsp;… (+{len(subs) - _sub_shown} more sub-sections)")
                            break
                        _sub_shown += 1
                        t  = sub.get("text") or ""
                        sp = sub.get("product")
                        items = [it for it in (sub.get("items") or []) if isinstance(it, str) and it.strip()]
                        line = f"&nbsp;&nbsp;&nbsp;&nbsp;— {t}"
                        if isinstance(sp, str) and sp.strip():
                            line += f" <i>[{sp.strip()}]</i>"
                        if items:
                            shown_items = items[:5]
                            tail = f" …(+{len(items) - 5})" if len(items) > 5 else ""
                            line += " :: " + ", ".join(shown_items) + tail
                        _lines.append(line)
                _hdrs_html = "<br>" + "<br>".join(_lines)
            else:
                _hdrs_html = " (none extracted)"

            # Product list summary line — distinct products mentioned across
            # all this contract's section headers.
            _products = c.get("products_in_headers") or []
            _products_html = ", ".join(_products) if _products else "(none)"

            hover_texts.append(
                f"<b>{c['filename']}</b><br>"
                f"Client: {c['client']}<br>"
                f"Type: {c['contract_type']}"
                + (f" #{c['amendment_number']}" if c.get("amendment_number") else "") + "<br>"
                f"Signed Date: {c.get('signed_date') or 'unknown'}<br>"
                f"Effective Date: {c.get('effective_date') or 'unknown'}<br>"
                f"Hierarchy Level: {c['hierarchy_level']}<br>"
                f"Parent: {_parent_filename(c, contracts)}<br>"
                f"Hierarchy Score: {c.get('hierarchy_confidence_score', 0)} / 100 "
                f"({bucket or 'unknown'})<br>"
                f"Extraction Score: {c.get('extraction_confidence_score', 0)} / 100<br>"
                f"Resolution Method: {c.get('hierarchy_method') or 'unknown'}<br>"
                f"Products in Headers: {_products_html}<br>"
                f"Section Headers:{_hdrs_html}<br>"
                f"Notes: {c.get('hierarchy_notes', '')}"
            )

        category_trace_indices.append(len(fig.data))
        category_group_names.append(group_name)
        category_confidences.append(trace_confidences)
        category_cache_keys.append([c["cache_key"] for c in group_contracts])
        category_contract_types.append([c.get("contract_type") or "" for c in group_contracts])
        colors_by_confidence_all.append(colors_conf)
        colors_by_type_all.append(colors_type)
        colors_by_era_all.append(colors_era)

        # (4) Real category trace, placed in the "category" legend group so
        #     it renders under a 'Document Type' sub-header.
        fig.add_trace(go.Scatter(
            x=xs, y=ys,
            mode="markers",
            marker=dict(
                size=11,
                color=colors_conf,   # default: Color by Confidence
                symbol=SYMBOL_MAP.get(group_name, "circle"),
                line=dict(width=1, color="white")
            ),
            name=f"{group_name} ({len(group_contracts)})",
            legendgroup="category",
            legendgrouptitle_text="Document Type",
            text=hover_texts,
            hovertemplate="%{text}<extra></extra>",
        ))

    # (4) Confidence legend entries — dummy traces with no data, purely to
    #     render a 'Confidence' color legend alongside the category symbols.
    #     Labels now show the numeric-score bucket thresholds driving the colours.
    CONFIDENCE_LEGEND = [
        ("High (≥70%)",      COLOR_MAP["high"]),
        ("Medium (40–69%)",  COLOR_MAP["medium"]),
        ("Low (<40%)",       COLOR_MAP["low"]),
        ("Unknown",          COLOR_MAP[None]),
    ]
    for label, color in CONFIDENCE_LEGEND:
        fig.add_trace(go.Scatter(
            x=[None], y=[None],
            mode="markers",
            marker=dict(size=11, color=color, symbol="circle",
                        line=dict(width=1, color="white")),
            name=label,
            legendgroup="confidence",
            legendgrouptitle_text="Confidence",
            showlegend=True,
            hoverinfo="skip",
        ))

    # Combined trace index list used by the "Color by" and "Show" dropdowns
    # so they can restyle category markers and connector edges in one call.
    all_indices = category_trace_indices + connector_trace_indices
    n_cat       = len(category_trace_indices)
    n_conn      = len(connector_trace_indices)

    # (3) "Color by" dropdown — swaps marker.color on category traces AND
    #     line.color on connector traces so edges track the active colour mode.
    #     Each trace slot in all_indices gets both a marker.color entry and a
    #     line.color entry; Plotly applies whichever is relevant for that trace
    #     type and silently ignores the other.
    _TRANSPARENT = "rgba(0,0,0,0)"

    def _color_args(cat_marker_colors, conn_line_colors):
        marker_list = list(cat_marker_colors) + [_TRANSPARENT] * n_conn
        line_list   = [_TRANSPARENT] * n_cat + list(conn_line_colors)
        return [
            {"marker.color": marker_list, "line.color": line_list},
            all_indices,
        ]

    color_by_buttons = [
        dict(
            label="Color by: Confidence",
            method="restyle",
            args=_color_args(colors_by_confidence_all, connector_colors_by_conf),
        ),
        dict(
            label="Color by: Contract Type",
            method="restyle",
            args=_color_args(colors_by_type_all, connector_colors_by_type),
        ),
        dict(
            label="Color by: Era",
            method="restyle",
            args=_color_args(colors_by_era_all, connector_colors_by_era),
        ),
    ]

    # (3) "Show" dropdown — toggles visibility on category traces ('legendonly'
    #     keeps the entry clickable in the legend but hides the points) AND on
    #     connector traces whose child category is being hidden, so no stub
    #     edges are left dangling.
    def _show_args(keep):
        cat_vis  = [True if gn in keep else "legendonly" for gn in category_group_names]
        conn_vis = [True if cg in keep else False        for cg in connector_child_groups]
        return [{"visible": cat_vis + conn_vis}, all_indices]

    _ALL_GROUPS = {"Parent", "Amendment", "Sub-Amendment", "Others", "Orphan"}
    show_buttons = [
        dict(label="Show: All",                method="restyle", args=_show_args(_ALL_GROUPS)),
        dict(label="Show: Parents only",       method="restyle",
             args=_show_args({"Parent"})),
        dict(label="Show: Amendments only",    method="restyle",
             args=_show_args({"Amendment", "Sub-Amendment"})),
        dict(label="Show: Hide Orphans",       method="restyle",
             args=_show_args(_ALL_GROUPS - {"Orphan"})),
    ]

    # (3-new) "Filter by Confidence" dropdown — replaces the previously
    #     cosmetic confidence legend with a real filter. Non-matching category
    #     markers are faded to opacity 0.08 (kept visible but de-emphasised so
    #     they still show context), and non-matching connectors are faded to
    #     a very light grey on their line.color. Both effects use restyle so
    #     they compose cleanly with "Color by" (which will overwrite connector
    #     line.color on its next click) and with "Show" (which uses visible).
    _FADE_CAT_OPACITY = 0.08
    _FADE_CONN_LINE   = "rgba(180,180,180,0.08)"

    def _filter_args(keep_confs):
        """keep_confs: set of confidence values to keep visible; None = show all.
        None may appear inside the set to match contracts whose confidence is
        literally None (no signal)."""
        if keep_confs is None:
            cat_opacity_per_trace = [[1.0] * len(confs) for confs in category_confidences]
            conn_line_colors      = list(connector_colors_by_conf)
        else:
            cat_opacity_per_trace = [
                [1.0 if cf in keep_confs else _FADE_CAT_OPACITY for cf in confs]
                for confs in category_confidences
            ]
            conn_line_colors = [
                connector_colors_by_conf[i] if connector_child_confs[i] in keep_confs
                else _FADE_CONN_LINE
                for i in range(n_conn)
            ]
        marker_opacity_list = cat_opacity_per_trace + [1.0] * n_conn
        line_color_list     = [_TRANSPARENT] * n_cat + conn_line_colors
        return [
            {"marker.opacity": marker_opacity_list, "line.color": line_color_list},
            all_indices,
        ]

    filter_by_conf_buttons = [
        dict(label="Confidence: All",              method="restyle", args=_filter_args(None)),
        dict(label="Confidence: High (≥70%)",      method="restyle", args=_filter_args({"high"})),
        dict(label="Confidence: Medium (40–69%)",  method="restyle", args=_filter_args({"medium"})),
        dict(label="Confidence: Low (<40%)",       method="restyle", args=_filter_args({"low"})),
        dict(label="Confidence: High + Medium",    method="restyle", args=_filter_args({"high", "medium"})),
        dict(label="Confidence: Unknown",          method="restyle", args=_filter_args({None})),
    ]

    # ── Parent-type filter ────────────────────────────────────────────────
    # The "Parent" bucket on the chart hides a wide variety of underlying
    # contract types — MSA, Software Agreement, Hardware Agreement, Services
    # Agreement, License Agreement, Purchase Agreement, Maintenance Agreement,
    # Marketing Agreement, Network Agreement, Referral Agreement, Internet
    # Banking Agreement, Access Agreement, Escrow Agreement, Subscription
    # Agreement, Miscellaneous, etc. This dropdown lets a reviewer bifurcate
    # the Parent bucket by the underlying contract_type so they can see,
    # for example, just the Hardware Agreements at the root row, with every
    # other Parent marker (and every non-Parent marker like Amendment /
    # Sub-Amendment / Orphan) left untouched.
    #
    # Implementation mirrors the Filter-by-Confidence pattern: a per-marker
    # marker.opacity restyle. Non-matching Parent markers fade to
    # _FADE_CAT_OPACITY; non-matching markers in other groups are kept at
    # opacity 1.0 (because the filter scope is Parents only). Connectors are
    # left at full visibility — they are coloured by the child's verdict and
    # would be confusing if also faded by parent type.
    parent_group_idx = (
        category_group_names.index("Parent") if "Parent" in category_group_names else None
    )

    # Count of each contract_type among Parent markers, for the dropdown labels.
    from collections import Counter
    parent_type_counts = Counter()
    if parent_group_idx is not None:
        for ct in category_contract_types[parent_group_idx]:
            if ct:
                parent_type_counts[ct] += 1
    distinct_parent_types = sorted(parent_type_counts.keys(),
                                    key=lambda t: (-parent_type_counts[t], t))

    def _parent_type_args(keep_type):
        """keep_type: contract_type string to keep at full opacity within
        the Parent bucket; None = show every Parent at full opacity.
        Non-Parent markers (Amendment / Sub-Amendment / Others / Orphan)
        are unaffected — their opacity is always 1.0 under this filter."""
        cat_opacity_per_trace = []
        for i, name in enumerate(category_group_names):
            if name == "Parent" and keep_type is not None:
                cat_opacity_per_trace.append(
                    [1.0 if ct == keep_type else _FADE_CAT_OPACITY
                     for ct in category_contract_types[i]]
                )
            else:
                cat_opacity_per_trace.append(
                    [1.0] * len(category_contract_types[i])
                )
        marker_opacity_list = cat_opacity_per_trace + [1.0] * n_conn
        return [
            {"marker.opacity": marker_opacity_list},
            all_indices,
        ]

    parent_type_buttons = [
        dict(label="Parent type: All",
             method="restyle",
             args=_parent_type_args(None)),
    ]
    for ct in distinct_parent_types:
        n = parent_type_counts[ct]
        # Truncate over-long contract_type labels so the dropdown doesn't
        # blow out the chart width — same approach as the Products filter.
        label = ct if len(ct) <= 32 else ct[:29] + "…"
        parent_type_buttons.append(dict(
            label=f"Parent type: {label} ({n})",
            method="restyle",
            args=_parent_type_args(ct),
        ))

    # ── Product-lineage filter ────────────────────────────────────────────
    # Build a map: product → set of cache_keys whose section_header_products
    # contain that product. The lineage of a product is therefore every
    # contract in this set; chronologically it spans from the oldest
    # signed_date among those contracts to the latest. Selecting a product
    # in the dropdown fades all NON-lineage contracts (markers + connectors)
    # so the chart visually collapses to just that product's hierarchy.
    product_to_cache_keys = defaultdict(set)
    for c in contracts:
        for p in (c.get("products_in_headers") or []):
            if isinstance(p, str) and p.strip():
                product_to_cache_keys[p.strip()].add(c["cache_key"])

    # Sort products by occurrence count (most-referenced first) so the
    # dropdown surfaces the highest-signal entries at the top. Cap the
    # dropdown at 30 entries so the menu stays usable even for long-tail
    # portfolios; if you need to filter by a less common product, run a
    # one-line cache lookup instead (see the README).
    _MAX_PRODUCTS_IN_DROPDOWN = 30
    sorted_products = sorted(
        product_to_cache_keys.keys(),
        key=lambda p: (-len(product_to_cache_keys[p]), p.lower()),
    )[:_MAX_PRODUCTS_IN_DROPDOWN]

    # ── Lineage-line overlays ─────────────────────────────────────────────
    # For each product in the dropdown we add a single Scatter trace that
    # connects every contract in the product's lineage in chronological
    # order. The trace data is baked in at HTML-generation time (so the
    # arrays aren't duplicated per dropdown button); visibility is toggled
    # at runtime via trace-level `opacity`:
    #   - opacity 0.0  → line is invisible (default state for every trace)
    #   - opacity 0.85 → line is shown for the currently-selected product
    # This avoids resizing the underlying x/y data on each click, and it
    # leaves the `visible` attribute alone so the Show filter still
    # composes with Products.
    _LINEAGE_LINE_COLOR = "rgb(75,40,170)"     # solid indigo, alpha controlled by trace opacity
    _LINEAGE_LINE_ON    = 0.85
    _LINEAGE_LINE_OFF   = 0.0

    product_line_indices = []
    for p in sorted_products:
        keys = product_to_cache_keys[p]
        # Sort the product's contracts chronologically by signed_date.
        # Contracts with missing dates plot at UNKNOWN_STRIP_MID via _plot_x,
        # so they cluster at the left edge of the lineage line.
        lineage_contracts = sorted(
            (c for c in contracts if c["cache_key"] in keys),
            key=lambda c: _plot_x(c),
        )
        lin_xs = [_plot_x(c) for c in lineage_contracts]
        lin_ys = [level_y(c["hierarchy_level"]) for c in lineage_contracts]
        product_line_indices.append(len(fig.data))
        fig.add_trace(go.Scatter(
            x=lin_xs, y=lin_ys,
            mode="lines+markers",
            line=dict(color=_LINEAGE_LINE_COLOR, width=2.5, dash="dot",
                      shape="spline", smoothing=0.6),
            marker=dict(size=8, color=_LINEAGE_LINE_COLOR,
                        symbol="circle-open", line=dict(width=2, color=_LINEAGE_LINE_COLOR)),
            opacity=_LINEAGE_LINE_OFF,        # hidden until the dropdown selects this product
            showlegend=False,
            hoverinfo="skip",
            name=f"Lineage · {p}",
        ))

    n_product_lines = len(product_line_indices)
    # Combined index list used by the Products dropdown so its restyle
    # touches everything: category traces (marker.opacity), connector
    # traces (line.color), AND the per-product lineage lines (opacity).
    all_indices_with_product = all_indices + product_line_indices

    def _product_args(product_name):
        """Return restyle args for a Products-dropdown click.
        product_name is None → "(no filter)" — restores cat/conn defaults
                               and hides every lineage line.
        product_name is a string → fades non-lineage markers & connectors,
                               and turns on ONLY that product's lineage line."""
        if product_name is None:
            cat_opacity_per_trace = [[1.0] * len(keys) for keys in category_cache_keys]
            conn_line_colors      = list(connector_colors_by_conf)
            line_opacities        = [_LINEAGE_LINE_OFF] * n_product_lines
        else:
            keep = product_to_cache_keys.get(product_name, set())
            cat_opacity_per_trace = [
                [1.0 if k in keep else _FADE_CAT_OPACITY for k in trace_keys]
                for trace_keys in category_cache_keys
            ]
            conn_line_colors = [
                connector_colors_by_conf[i] if connector_child_cache_keys[i] in keep
                else _FADE_CONN_LINE
                for i in range(n_conn)
            ]
            line_opacities = [
                _LINEAGE_LINE_ON if sorted_products[i] == product_name else _LINEAGE_LINE_OFF
                for i in range(n_product_lines)
            ]

        # marker.opacity: per-point arrays for cat traces, scalar 1.0 for
        # conn traces (no markers — no-op visually), scalar 1.0 for product
        # lines (their visibility is driven by trace-level `opacity`).
        marker_opacity_list = (
            cat_opacity_per_trace
            + [1.0] * n_conn
            + [1.0] * n_product_lines
        )
        # line.color: transparent for cat (mode="markers"), per-edge for conn,
        # constant indigo for product lines (alpha already in the color).
        line_color_list = (
            [_TRANSPARENT] * n_cat
            + conn_line_colors
            + [_LINEAGE_LINE_COLOR] * n_product_lines
        )
        # opacity (trace-level): preserved at 1.0 for cat (default), 0.45
        # for conn (matches their initial value — no-op), and toggled per
        # product line. We deliberately do NOT touch `visible`, so any prior
        # Show filter survives.
        opacity_list = (
            [1.0] * n_cat
            + [0.45] * n_conn
            + line_opacities
        )

        return [
            {
                "marker.opacity": marker_opacity_list,
                "line.color":     line_color_list,
                "opacity":        opacity_list,
            },
            all_indices_with_product,
        ]

    products_buttons = [
        dict(
            label="Products: (no filter)",
            method="restyle",
            args=_product_args(None),
        )
    ]
    for p in sorted_products:
        n = len(product_to_cache_keys[p])
        # Truncate long product names so the dropdown doesn't blow out
        # to the edge of the figure.
        label = p if len(p) <= 40 else p[:37] + "…"
        products_buttons.append(dict(
            label=f"Products: {label} ({n})",
            method="restyle",
            args=_product_args(p),
        ))

    # Resolve chart title: prefer CLIENT_TITLE_OVERRIDES, fall back to raw
    # folder name, fall back to the old generic string if no client was passed.
    if client_name:
        display_name = CLIENT_TITLE_OVERRIDES.get(client_name, client_name)
        title_text   = f"{display_name} - Document Hierarchy"
    else:
        title_text = "Document Hierarchy"

    fig.update_layout(
        title=dict(
            text=title_text,
            font=dict(size=18),
            x=0.5, xanchor="center",
            yref="container",
            y=0.97, yanchor="top",
        ),
        xaxis=dict(
            title="Timeline",
            # Visible range starts at 1980 by default (per user request). If any
            # contract has an unknown effective_date, the range is widened to 1975
            # so the dedicated "Unknown Date" strip (1976–1979) fits inside it.
            range=[xaxis_range_start, datetime(2028, 1, 1)],
        ),
        yaxis=dict(
            title="Hierarchy Level",
            # Tick ladder is built from the actual deepest level in this run
            # (clamped to a minimum of 5 for visual consistency on small
            # datasets), plus the Orphan strip anchored at ORPHAN_Y. Levels 6+
            # that previously rendered as unlabelled markers between Level 5
            # and Orphan now get their own labelled ticks.
            tickvals=(
                [0]
                + [-i for i in range(1, MAX_LEVEL_SHOWN + 1)]
                + [ORPHAN_Y]
            ),
            ticktext=(
                ["Parent"]
                + [f"Level {i}" for i in range(1, MAX_LEVEL_SHOWN + 1)]
                + ["Orphan"]
            ),
            # Pin the visible range so the dashed divider + Orphan strip stay
            # comfortably inside the plot area rather than getting clipped.
            range=[ORPHAN_Y - 1.2, 0.6],
            zeroline=True, zerolinecolor="rgba(0,0,0,0.1)"
        ),
        hovermode="closest",
        plot_bgcolor="white",
        paper_bgcolor="white",
        # Chart height scales with the deepest non-orphan level so each tick
        # row keeps roughly the same visual height. At 5 levels the chart is
        # the original 780px; at 21 levels (Peoples Bank's longest chain) it
        # grows to ~1380px so rows aren't crammed against each other and the
        # Orphan strip stays clearly separated. Capped at 1600 to prevent
        # absurd heights on pathological chains.
        height=min(1600, max(780, 360 + 50 * (MAX_LEVEL_SHOWN + 3))),
        legend=dict(
            orientation="v",
            yanchor="top",   y=1.0,
            xanchor="left",  x=1.02,
            font=dict(size=11),
            groupclick="toggleitem",
        ),
        # Top margin enlarged to host the title plus a 2-row dropdown grid.
        # Row 1 (top):    Color by  +  Show
        # Row 2 (middle): Filter by Confidence  +  Products
        # Row 3 (bottom): Parent type  (full-width, spans the row alone — it
        #                                can have many bank-specific subtypes
        #                                so the labels are kept wide)
        # The 3-row layout was introduced when the Parent-type filter
        # landed; the previous 2-row layout couldn't fit a fifth dropdown
        # without overlapping labels.
        margin=dict(l=80, r=220, t=220, b=60),
        updatemenus=[
            # ─── Row 1 (top) ───────────────────────────────────────────
            dict(
                type="dropdown",
                buttons=color_by_buttons,
                direction="down",
                showactive=True,
                x=0.00, xanchor="left",
                y=1.18, yanchor="bottom",
                bgcolor="white",
                bordercolor="rgba(0,0,0,0.25)",
                pad=dict(l=4, r=4, t=4, b=4),
            ),
            dict(
                type="dropdown",
                buttons=show_buttons,
                direction="down",
                showactive=True,
                x=0.50, xanchor="left",
                y=1.18, yanchor="bottom",
                bgcolor="white",
                bordercolor="rgba(0,0,0,0.25)",
                pad=dict(l=4, r=4, t=4, b=4),
            ),
            # ─── Row 2 (middle) ────────────────────────────────────────
            dict(
                type="dropdown",
                buttons=filter_by_conf_buttons,
                direction="down",
                showactive=True,
                x=0.00, xanchor="left",
                y=1.10, yanchor="bottom",
                bgcolor="white",
                bordercolor="rgba(0,0,0,0.25)",
                pad=dict(l=4, r=4, t=4, b=4),
            ),
            # Products filter — defaults to "(no filter)" so the original
            # hierarchy diagram is what users see on first load.
            dict(
                type="dropdown",
                buttons=products_buttons,
                direction="down",
                showactive=True,
                x=0.50, xanchor="left",
                y=1.10, yanchor="bottom",
                bgcolor="white",
                bordercolor="rgba(0,0,0,0.25)",
                pad=dict(l=4, r=4, t=4, b=4),
            ),
            # ─── Row 3 (bottom) ────────────────────────────────────────
            # Parent-type filter — bifurcates the unified Parent bucket
            # into its underlying contract_type subtypes (MSA, Hardware
            # Agreement, Software Agreement, etc.). Defaults to "All".
            dict(
                type="dropdown",
                buttons=parent_type_buttons,
                direction="down",
                showactive=True,
                x=0.00, xanchor="left",
                y=1.02, yanchor="bottom",
                bgcolor="white",
                bordercolor="rgba(0,0,0,0.25)",
                pad=dict(l=4, r=4, t=4, b=4),
            ),
        ],
    )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    # include_plotlyjs=True embeds the JS — file works offline, no CDN needed
    fig.write_html(str(output_path), include_plotlyjs=True)
    print(f"  Saved: {Path(output_path).resolve()}")


# ============================================================
# PHASE 5: EXCEL EXPORT
# ============================================================

def export_excel(contracts, output_path):
    """Export hierarchy to color-coded Excel workbook."""

    # ── Display-time label mapping (Excel-only) ─────────────────────────────
    # Per user request (May 26 2026 follow-ups): the Excel output uses a
    # mix of the new "Miscellaneous" label and the legacy "standalone"
    # label depending on which column / sheet the reviewer is looking at:
    #
    #   • Contract_Type column        → "Miscellaneous"  (data layer, no remap)
    #   • Hierarchy_Method column     → "standalone"     (remap via _excel_method)
    #   • Classification Grid bracket → "Standalone"     (remap in grid builder)
    #
    # The data layer (`c["contract_type"]`, `c["hierarchy_method"]`) keeps
    # "Miscellaneous" / "miscellaneous" for HTML consistency; this writer
    # applies only the targeted remaps the user asked for.
    def _excel_method(m):
        return "standalone" if m == "miscellaneous" else m

    rows = []
    # Build a cache_key → filename lookup so the Duplicate_Of column can show
    # the human-readable filename of each canonical instead of the raw key.
    _key_to_filename = {c["cache_key"]: c["filename"] for c in contracts}
    for c in contracts:
        dup_key = c.get("duplicate_of")
        dup_of  = _key_to_filename.get(dup_key, "") if dup_key else ""
        rows.append({
            "Filename":           c["filename"],
            "Client":             c["client"],
            "Contract_Type":      c["contract_type"],
            "Signed_Date":        c.get("signed_date") or "",
            "Effective_Date":     c.get("effective_date") or "",
            "Amendment_No":       c.get("amendment_number") or "",
            "Parent_Contract":    _parent_filename(c, contracts),
            "Duplicate_Of":       dup_of,
            "Hierarchy_Level":    c["hierarchy_level"],
            # Hierarchy_Status mirrors the validator's ContractPoint Status
            # column: 'Parent' | 'Child' | 'Orphan' | 'Duplicate'.
            # ('Parent' covers every level-0 root regardless of body subtype
            #  — MSA, Master Services Agreement, Software Agreement,
            #  Hardware Agreement, Internet Banking / Access / Escrow /
            #  Subscription Agreement, etc. The unified label was confirmed
            #  by the validator's team on May 22 2026 and renamed from
            #  'Original MSA' to 'Parent' on May 26 2026 so Excel matches
            #  the HTML display label. 'Duplicate' is Pass-3 flagged
            #  multi-department-upload copies.)
            "Hierarchy_Status":   c.get("hierarchy_status") or "",
            "Hierarchy_Method":   _excel_method(c.get("hierarchy_method") or ""),
            "Hierarchy_Score":    c.get("hierarchy_confidence_score") if c.get("hierarchy_confidence_score") is not None else "",
            "Hierarchy_Bucket":   c.get("hierarchy_confidence_bucket") or "",
            "Hierarchy_Conf":     c.get("hierarchy_confidence") or "",
            "Extraction_Score":   c.get("extraction_confidence_score", 0),
            "Extraction_Conf":    c.get("extraction_confidence") or "",
            "Is_Ambiguous":       "yes" if c.get("is_ambiguous") else "no",
            "Parties":            "; ".join(c.get("parties") or []),
            # Section_Headers shows the nested 3-level section structure
            # as a single multi-line string:
            #     • Schedule A: ASP Services Exhibit
            #         — Account Processing Services [Premier Account Processing]
            #             :: Premier Account Processing, Cleartouch Account Processing
            #         — ATM/EFT Services :: ATM Driving, ATM Driving - Host Connection Fees
            #     • Termination
            # The bracketed products are CANONICAL dictionary names; the
            # ":: …" comma list shows the LLM-extracted leaf-item names from
            # each sub-header (the deepest level — typically where the most
            # specific Fiserv product names appear).
            "Section_Headers":    c.get("section_structure_text") or "",
            "Products_in_Headers":     ", ".join(c.get("products_in_headers") or []),
            # Audit column: the LLM's pre-canonicalisation product names —
            # everything the LLM surfaced from this contract's section headers
            # before dictionary alignment. Useful for spotting dictionary
            # gaps (real products the LLM finds that aren't yet in v*.xlsx).
            "Products_in_Headers_Raw": ", ".join(c.get("products_in_headers_raw") or []),
            "Rationale":          c.get("hierarchy_rationale") or "",
            "Notes":              c.get("hierarchy_notes") or "",
        })

    df = pd.DataFrame(rows)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(str(output_path), index=False, engine="openpyxl")

    # Style the workbook
    wb = load_workbook(str(output_path))
    ws = wb.active

    LEVEL_COLORS = {
        0:  "D5E8D4",   # green  — Parent (unified root: MSA + non-MSA root-eligibles)
        1:  "DAE8FC",   # blue   — Amendment
        2:  "FFF2CC",   # yellow — Sub-Amendment
        3:  "FFE6CC",   # orange — Deeper
        -1: "F8CECC",   # red    — Orphan
    }
    AMBIGUOUS_FILL = "E8E0F4"   # lavender — overrides LEVEL_COLORS when Is_Ambiguous=yes
    DUPLICATE_FILL = "E6F2FF"   # pale blue — Pass-3 duplicates of a canonical root

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Color rows by hierarchy level. Precedence (highest first):
    #   1. Duplicate row (validator-confirmed Pass-3 cluster)
    #   2. Ambiguous row (silent tie-break detected)
    #   3. Hierarchy level
    level_col_idx  = list(df.columns).index("Hierarchy_Level") + 1
    amb_col_idx    = list(df.columns).index("Is_Ambiguous")    + 1
    status_col_idx = list(df.columns).index("Hierarchy_Status") + 1
    for row in ws.iter_rows(min_row=2):
        level_val  = row[level_col_idx  - 1].value
        amb_val    = row[amb_col_idx    - 1].value
        status_val = row[status_col_idx - 1].value
        try:
            level = int(level_val)
        except (TypeError, ValueError):
            level = None
        if status_val == "Duplicate":
            hex_color = DUPLICATE_FILL
        elif amb_val == "yes":
            hex_color = AMBIGUOUS_FILL
        else:
            hex_color = LEVEL_COLORS.get(level, "FFFFFF") if level is not None else "FFFFFF"
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for cell in row:
            cell.fill = fill

    # ── Classification Grid sheet ─────────────────────────────────────────
    # Editable 5×4 grid — Document Type (rows) × Confidence Level (cols).
    # Each cell is pre-filled with:
    #   • the current count of contracts for that (type, confidence) cell
    #   • a breakdown by resolution method (reference / doc_code / date / root / orphan)
    #   • the static logic that determined classification into this cell
    _add_classification_grid_sheet(wb, contracts)

    wb.save(str(output_path))
    print(f"  Saved: {Path(output_path).resolve()}")


def _add_classification_grid_sheet(wb, contracts):
    """Append a 'Classification Grid' sheet to the workbook. Rows = document
    types; columns = confidence buckets. Each cell contains the live count, a
    per-method breakdown for that cell, and a short description of the logic
    the script uses to classify contracts into that (type, confidence) bucket.
    All cells are editable — reviewers can over-write them with their own notes."""

    # Single root row — 'Parent' is the unified label for every level-0
    # root-eligible doc (MSA, Master Services Agreement, Software / Hardware /
    # Services / Internet Banking / Access / Escrow / Subscription / etc.
    # Agreements), per the validator's-team May 22 2026 confirmation. The
    # label was renamed from 'Original MSA' to 'Parent' on May 26 2026 so
    # the Excel data field matches the HTML display label.
    DOC_TYPES = ["Parent", "Amendment", "Sub-Amendment", "Others", "Orphan"]
    CONF_LEVELS = [
        ("High (≥70%)",     "high"),
        ("Medium (40–69%)", "medium"),
        ("Low (<40%)",      "low"),
        ("Unknown",         None),
    ]

    def _doc_type_of(c):
        lvl = c["hierarchy_level"]
        if lvl == 0:    return "Parent"   # unified root bucket
        if lvl == 1:    return "Amendment"
        if lvl == 2:    return "Sub-Amendment"
        if lvl  > 2:    return "Others"
        return "Orphan"

    # Count + method-breakdown per (doc_type, conf_bucket)
    counts     = {d: {label: 0 for label, _ in CONF_LEVELS} for d in DOC_TYPES}
    method_map = {d: {label: defaultdict(int) for label, _ in CONF_LEVELS} for d in DOC_TYPES}
    # Display-time remap (Classification Grid only, May 26 2026 follow-up):
    # the bracketed per-method breakdown under each count should show
    # 'Standalone' (capitalised) instead of the internal 'miscellaneous'
    # method identifier. Other method names ('root', 'reference', etc.)
    # pass through unchanged.
    def _grid_method_label(m):
        return "Standalone" if m == "miscellaneous" else m

    for c in contracts:
        dt     = _doc_type_of(c)
        bucket = c.get("hierarchy_confidence_bucket")
        label  = next((lbl for lbl, b in CONF_LEVELS if b == bucket), "Unknown")
        counts[dt][label] += 1
        method_map[dt][label][_grid_method_label(c.get("hierarchy_method") or "unknown")] += 1

    # Static logic text per cell — explains why a contract lands there.
    # Scoring recap: composite = round((method_base + extraction_score) / 2),
    # then bucketed: ≥70 High, 40–69 Medium, <40 Low, orphans → Unknown.
    # Method bases: root 95, reference 90, doc_code 60, date 30, orphan 0.
    LOGIC = {
        # Unified root bucket — every level-0 root-eligible doc (MSA,
        # Master Services Agreement, Software / Hardware / Services /
        # Maintenance / Internet Banking / Access / Escrow / Subscription
        # / etc. Agreements) per validator's-team May 22 2026 confirmation.
        # Previously split between 'Original MSA/Renewal Amendment' and
        # 'Standalone'; the unified label was renamed from 'Original MSA'
        # to 'Parent' on May 26 2026 so Excel matches HTML.
        ("Parent", "High (≥70%)"):
            "Any root-eligible document with no resolvable parent reference and "
            "no shared document code. Method 'root' (base 95 for MSA) or "
            "'standalone' (base 95 for non-MSA root-eligible types). Composite "
            "≥70 for any extraction score ≥ 45.",
        ("Parent", "Medium (40–69%)"):
            "Root-eligible doc whose LLM extraction score came in <45 (weak OCR / "
            "unclear title page). Placement as a root is still correct.",
        ("Parent", "Low (<40%)"):
            "Not reachable — method base 95 dominates any non-negative extraction score.",
        ("Parent", "Unknown"):
            "Not applicable — roots always receive a numeric bucket.",

        ("Amendment", "High (≥70%)"):
            "Parent resolved via textual parent_reference or supersedes quote "
            "(method 'reference', base 90) with extraction ≥50, OR via shared "
            "internal document code (method 'doc_code', base 60) with extraction ≥80.",
        ("Amendment", "Medium (40–69%)"):
            "Reference match with weak OCR (ext <50), OR doc-code match with "
            "moderate extraction, OR date-proximity fallback (base 30) with "
            "very strong extraction (ext ≥50).",
        ("Amendment", "Low (<40%)"):
            "Date-proximity fallback (base 30) with ext <50 — weakest method + "
            "weak data. Reviewer sanity-check recommended.",
        ("Amendment", "Unknown"):
            "Not applicable — only orphans receive Unknown.",

        ("Sub-Amendment", "High (≥70%)"):
            "Child of an Amendment, resolved via explicit reference (base 90) "
            "with strong extraction, OR via shared doc codes (base 60) with ≥80 extraction.",
        ("Sub-Amendment", "Medium (40–69%)"):
            "Reference or doc-code match with moderate extraction, OR date-proximity "
            "to an older Amendment with high extraction.",
        ("Sub-Amendment", "Low (<40%)"):
            "Date-proximity to nearest older Amendment with weak extraction.",
        ("Sub-Amendment", "Unknown"):
            "Not applicable — only orphans receive Unknown.",

        ("Others", "High (≥70%)"):
            "Level 3+ nested contract (amendment-to-sub-amendment, SOW or Purchase "
            "Order under a Sub-Amendment, etc.), resolved via explicit reference "
            "or shared doc codes with strong extraction.",
        ("Others", "Medium (40–69%)"):
            "Deep-level resolution via any method combined with moderate extraction.",
        ("Others", "Low (<40%)"):
            "Date-proximity fallback at level 3+. Chains this deep warrant manual review.",
        ("Others", "Unknown"):
            "Not applicable — only orphans receive Unknown.",

        ("Orphan", "High (≥70%)"):
            "Not applicable.",
        ("Orphan", "Medium (40–69%)"):
            "Not applicable.",
        ("Orphan", "Low (<40%)"):
            "Not applicable.",
        ("Orphan", "Unknown"):
            "No parent could be resolved via parent_references, shared doc codes, "
            "or date proximity. Bucket is set to None by design; score is 0.",
    }

    # Build the sheet
    ws = wb.create_sheet("Classification Grid")
    ws.append(["Document Type"] + [label for label, _ in CONF_LEVELS])

    for dt in DOC_TYPES:
        row = [dt]
        for label, _ in CONF_LEVELS:
            cnt       = counts[dt][label]
            breakdown = method_map[dt][label]
            bd_str    = ", ".join(f"{v} via {k}" for k, v in sorted(breakdown.items()))
            logic     = LOGIC.get((dt, label), "")
            cell_txt  = f"Count: {cnt}"
            if bd_str:
                cell_txt += f"\n({bd_str})"
            cell_txt += f"\n\nLogic: {logic}"
            row.append(cell_txt)
        ws.append(row)

    # Styling
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Header row
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # Row label column (A) — bold, light-grey fill
    label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r in range(2, 2 + len(DOC_TYPES)):
        c = ws.cell(row=r, column=1)
        c.font      = Font(bold=True)
        c.fill      = label_fill
        c.alignment = Alignment(vertical="center")
        c.border    = border

    # Data cells — wrap text, top-aligned, thin borders
    for r in range(2, 2 + len(DOC_TYPES)):
        for col in range(2, 2 + len(CONF_LEVELS)):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            c.border    = border

    # Dimensions: tall rows, wide data columns
    ws.column_dimensions["A"].width = 18
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].width = 48
    ws.row_dimensions[1].height = 32
    for r in range(2, 2 + len(DOC_TYPES)):
        ws.row_dimensions[r].height = 130

    # Freeze the header row + label column for easier navigation
    ws.freeze_panes = "B2"


# ============================================================
# MAIN
# ============================================================

def main():
    METRICS.start()   # run timer starts at 0 here; stops when we print the summary at the end
    print("=" * 62)
    print("  CONTRACT HIERARCHY ANALYZER — Fiserv Legal Ops POC")
    print("=" * 62)

    # Guard: catch missing configuration (everything is read from env / .env).
    if OPENAI_BACKEND == "fiserv":
        if not FOUNDATION_API_URL:
            print("\nERROR: OPENAI_BACKEND=fiserv but FOUNDATION_API_URL is not set.")
            print("Set FOUNDATION_API_URL (and FISERV_* values) in your .env file.")
            print("See .env.example and VDI_SETUP.md for the Fiserv setup.")
            return
    elif not OPENAI_API_KEY:
        print("\nERROR: No OpenAI API key found.")
        print("Set OPENAI_API_KEY in a .env file next to this script, or export it in your shell.")
        print("Copy .env.example to .env and paste your key. See VDI_SETUP.md for step-by-step help.")
        return

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    # Resolve client filter: CLI args override the ONLY_CLIENTS config.
    # Each CLI arg after the script name is treated as one client folder name.
    only_clients = list(sys.argv[1:]) if len(sys.argv) > 1 else ONLY_CLIENTS

    # ── Phase 1: Scan ──────────────────────────────────────
    if only_clients:
        print(f"\n[1/5] Scanning '{CONTRACTS_ROOT}' (restricted to: {only_clients})...")
    else:
        print(f"\n[1/5] Scanning '{CONTRACTS_ROOT}'...")
    contracts = scan_contracts(CONTRACTS_ROOT, only_clients=only_clients)
    if not contracts:
        print("No contracts found. Check your CONTRACTS_ROOT path and folder structure.")
        return

    client_names = sorted(set(c["client"] for c in contracts))
    print(f"  Found {len(contracts)} PDFs across {len(client_names)} client(s): {', '.join(client_names)}")

    # ── Phase 2: LLM Extraction ────────────────────────────
    cache       = load_cache(CACHE_FILE)

    # Stale-cache rescue pre-pass — two independent triggers.
    #
    # Trigger 1 — "missing parent_references" rescue:
    #   Evict entries for child-type contracts whose initial extraction
    #   returned an empty parent_references list AND haven't already been
    #   re-extracted with the strengthened prompt. This is the validator's
    #   "MSA information available in Child Agreement" category — amendments
    #   / schedules / SOWs / etc. whose body text DOES name a parent but the
    #   LLM's first pass missed it. With the strengthened parent_references
    #   prompt now in place, evicting these entries causes them to be
    #   re-extracted naturally below, picking up the parents the first pass
    #   missed.
    #
    # Trigger 2 — "type-mismatch" rescue (validator-recommended, May 2026):
    #   Evict entries where the LLM's cached contract_type strongly
    #   disagrees with a confident filename keyword (Master Agreement,
    #   Software Agreement, Hardware Agreement, etc.). A single mis-typed
    #   root agreement orphans every downstream amendment that cites it, so
    #   surgically re-extracting these 5-10 entries fixes dozens of resolver
    #   failures without paying the API cost of a full re-run. See
    #   _is_type_mismatched_extraction for the conservative trigger logic.
    #
    # Trigger 3 — "flat doc-codes" rescue (May 12 2026):
    #   Evict entries whose `internal_doc_codes` is still in the legacy
    #   flat-string format (no position labels). The position-aware
    #   matching landed on May 12 — see _is_flat_doc_codes_extraction.
    #
    # Trigger 4 — "flat section structure" rescue (May 22 2026):
    #   Evict entries whose section data is still in the legacy flat
    #   `section_headers` + `section_header_products` format. The new
    #   nested 3-level `section_structure` (top-level header → sub-headers
    #   → leaf items) captures the deeper product-name signal that the
    #   flat format missed. Re-extracting under the new prompt populates
    #   the deeper levels and meaningfully expands `products_in_headers`.
    #   See _is_flat_section_structure_extraction.
    if RESCUE_MISSING_PARENT_REFS:
        parent_stale_keys     = [k for k, v in cache.items() if _is_stale_child_extraction(v)]
        type_stale_keys       = [k for k, v in cache.items() if _is_type_mismatched_extraction(v, k)]
        doc_codes_stale_keys  = [k for k, v in cache.items() if _is_flat_doc_codes_extraction(v)]
        section_stale_keys    = [k for k, v in cache.items() if _is_flat_section_structure_extraction(v)]

        # Honour client filter — don't evict entries for clients we're not
        # processing in this run (avoids forcing API calls for unrelated data).
        if only_clients:
            allowed_prefixes = {f"{name}/" for name in only_clients}
            parent_stale_keys    = [k for k in parent_stale_keys    if any(k.startswith(p) for p in allowed_prefixes)]
            type_stale_keys      = [k for k in type_stale_keys      if any(k.startswith(p) for p in allowed_prefixes)]
            doc_codes_stale_keys = [k for k in doc_codes_stale_keys if any(k.startswith(p) for p in allowed_prefixes)]
            section_stale_keys   = [k for k in section_stale_keys   if any(k.startswith(p) for p in allowed_prefixes)]

        # Combine and dedupe — a single entry may match multiple triggers
        # (e.g. a mis-typed root agreement with flat-string doc-codes), but
        # it only needs to be evicted once.
        all_stale = list(dict.fromkeys(
            parent_stale_keys + type_stale_keys + doc_codes_stale_keys + section_stale_keys
        ))

        if all_stale:
            n_parent  = len(parent_stale_keys)
            n_type    = len(type_stale_keys)
            n_codes   = len(doc_codes_stale_keys)
            n_section = len(section_stale_keys)
            print(f"\n[Rescue] Evicting {len(all_stale)} stale cache entries for re-extraction "
                  f"with strengthened prompt:")
            print(f"    parent-references trigger: {n_parent}  | "
                  f"type-mismatch trigger: {n_type}  | "
                  f"doc-codes-format trigger: {n_codes}  | "
                  f"section-structure trigger: {n_section}")
            for k in all_stale[:8]:
                ct = (cache[k] or {}).get("contract_type", "?")
                short = k.split("/", 1)[1] if "/" in k else k
                triggers = []
                if k in parent_stale_keys:    triggers.append("missing-refs")
                if k in type_stale_keys:      triggers.append("type-mismatch")
                if k in doc_codes_stale_keys: triggers.append("flat-doc-codes")
                if k in section_stale_keys:   triggers.append("flat-sections")
                print(f"    [{ct}] {short[:70]}  ({', '.join(triggers)})")
            if len(all_stale) > 8:
                print(f"    …and {len(all_stale) - 8} more")
            for k in all_stale:
                del cache[k]
            save_cache(cache, CACHE_FILE)

    cached_n    = sum(1 for c in contracts if c["cache_key"] in cache)
    to_process  = len(contracts) - cached_n
    print(f"\n[2/5] Extracting metadata via LLM ({cached_n} cached, {to_process} to process)...")

    for i, contract in enumerate(contracts):
        prefix = f"  [{i + 1}/{len(contracts)}] {contract['client']}/{contract['filename']}"

        if contract["cache_key"] in cache:
            contract["metadata"] = cache[contract["cache_key"]]
            print(f"{prefix} — cached")
        else:
            print(f"{prefix} — calling API...", end="", flush=True)
            meta = extract_metadata(contract, cache)
            contract["metadata"] = meta
            save_cache(cache, CACHE_FILE)
            if meta.get("extraction_failed"):
                print(" FAILED (see error above, contract marked as Unknown)")
            else:
                pages = meta.get("num_pages", "?")
                conf  = (meta.get("extraction_confidence") or "low").upper()
                print(f" done ({pages} pages, confidence: {conf})")
            # Delay between API calls (skip delay after last call)
            if i < len(contracts) - 1 and not meta.get("extraction_failed"):
                time.sleep(API_DELAY)

    # ── Phases 3–5: Per-client hierarchy + outputs ────────
    # Each client gets its own output subfolder with its own HTML and Excel.
    # Hierarchy resolution is scoped per client (contracts only match parents within same client).
    print(f"\n[3–5] Resolving hierarchy and building outputs per client...")

    by_client_raw = defaultdict(list)
    for c in contracts:
        by_client_raw[c["client"]].append(c)

    for client_name in client_names:
        client_contracts = by_client_raw[client_name]
        client_dir = Path(OUTPUT_DIR) / client_name
        client_dir.mkdir(parents=True, exist_ok=True)

        print(f"\n  ── {client_name} ({len(client_contracts)} contracts) ──")

        # Phase 3: Hierarchy resolution (scoped to this client only)
        client_contracts = resolve_hierarchy(client_contracts)

        roots   = sum(1 for c in client_contracts if c["hierarchy_level"] == 0)
        level1  = sum(1 for c in client_contracts if c["hierarchy_level"] == 1)
        level2  = sum(1 for c in client_contracts if c["hierarchy_level"] == 2)
        deeper  = sum(1 for c in client_contracts if c["hierarchy_level"] > 2)
        orphans = sum(1 for c in client_contracts if c["hierarchy_level"] == -1)
        ambig   = sum(1 for c in client_contracts if c.get("is_ambiguous"))
        bk_high = sum(1 for c in client_contracts if c.get("hierarchy_confidence_bucket") == "high")
        bk_med  = sum(1 for c in client_contracts if c.get("hierarchy_confidence_bucket") == "medium")
        bk_low  = sum(1 for c in client_contracts if c.get("hierarchy_confidence_bucket") == "low")

        print(f"    Parent:                           {roots}")
        print(f"    Amendments (L1):                  {level1}")
        print(f"    Sub-Amend  (L2):                  {level2}")
        print(f"    Others (L3+):                     {deeper}")
        print(f"    Orphans:         {orphans}")
        print(f"    Ambiguous:       {ambig}")
        print(f"    Score buckets   — High(≥70): {bk_high}  Medium(40–69): {bk_med}  Low(<40): {bk_low}")

        # Phase 4: HTML
        build_visualization(
            client_contracts,
            client_dir / "contracts_hierarchy.html",
            client_name=client_name,
        )

        # Phase 5: Excel
        export_excel(client_contracts, client_dir / "contracts_hierarchy.xlsx")

    # ── Done ───────────────────────────────────────────────
    print(f"\n{'=' * 62}")
    print("  DONE. Per-client outputs are in the 'output' folder:")
    for client_name in client_names:
        print(f"    output/{client_name}/contracts_hierarchy.html")
        print(f"    output/{client_name}/contracts_hierarchy.xlsx")
    print(f"    Cache: output/extraction_cache.json  (re-run is instant)")
    print(f"{'=' * 62}")

    # Run-metric summary (tokens / calls / model / wall-clock). Printed last so
    # the run-time figure reflects the full successful run.
    print_run_metrics()


if __name__ == "__main__":
    main()
